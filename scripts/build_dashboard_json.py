"""
build_dashboard_json.py
-----------------------
업체가 보낸 .xlsx 파일을 읽어서 data/dashboard.json 으로 변환한다.

규칙:
  - data/raw/ 폴더 내에서 가장 최근에 수정된 .xlsx 파일을 입력으로 사용
  - 엑셀 시트:
      "일일평가"  → daily 배열
      "에러로그"  → errors 배열
    (시트명이 정확히 일치하지 않으면 부분 매칭으로 찾는다)
  - 출력: data/dashboard.json (UTF-8, 한글 그대로)

읽기 방식:
  - 1차: openpyxl (빠름, 일반 xlsx 전용)
  - 2차 폴백: xlwings (Excel 실행, DRM 보호된 xlsx도 처리 가능)
    → Windows + Excel 설치 필요. openpyxl이 zipfile.BadZipFile로 실패하면 자동 전환.

기대 컬럼:
  일일평가 시트  : 평가일, 입실인원, 주평가내용, 일일평가, 일일에러, 연속성공, 비고
  에러로그 시트  : No, 발생일, 시각, 회차, 코드, 유형, 상세, 원인, 조치, 결과, 삼성 담당자, 업체 담당자

'시각' 처리:
  업체가 custom h:mm 서식으로 시각을 보내면 openpyxl이 시간을 '하루의 분수'(0~1 float)로
  넘겨 0.xxxxx 형태로 깨질 수 있다. _cell_to_time() 가 이를 h:mm 으로 복원한다.
"""

from __future__ import annotations

import json
import math
import re
import zipfile
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
PROJECTS_ROOT = ROOT / "data" / "projects"
REGISTRY_PATH = ROOT / "data" / "projects.json"
PORTFOLIO_PATH = ROOT / "data" / "portfolio.json"
LIBRARY_PATH = ROOT / "data" / "library.json"

# 과제별 경로 — _set_project()가 설정 (멀티 과제: data/projects/<id>/ 기준)
RAW_DIR = OUT_PATH = CONFIG_PATH = MGMT_PATH = None


def _set_project(pid: str):
    """모든 입출력 경로를 해당 과제 폴더로 지정한다."""
    global RAW_DIR, OUT_PATH, CONFIG_PATH, MGMT_PATH
    base = PROJECTS_ROOT / pid
    if not base.is_dir():
        raise SystemExit(f"과제 폴더가 없습니다: {base}")
    RAW_DIR = base / "raw"
    OUT_PATH = base / "dashboard.json"
    CONFIG_PATH = base / "config.json"
    MGMT_PATH = base / "REPORT.xlsx"
    if not MGMT_PATH.exists() and (base / "SEC_REPORT.xlsx").exists():
        MGMT_PATH = base / "SEC_REPORT.xlsx"  # 구명 하위호환


def _load_registry():
    if not REGISTRY_PATH.exists():
        raise SystemExit(f"과제 레지스트리가 없습니다: {REGISTRY_PATH}")
    return json.loads(REGISTRY_PATH.read_text(encoding="utf-8"))


# ── 컬럼 헤더 매핑 ─────────────────────────────────────────────
# 한글 헤더에서 공백/괄호 제거 후 매칭한다 (오타 방지)
DAILY_FIELD_ALIASES = {
    "date":      ["평가일", "일자", "date"],
    "personnel": ["입실인원", "입실자", "인원", "personnel"],
    "activity":  ["주평가내용", "평가내용", "내용", "activity"],
    "total":     ["일일평가", "평가횟수", "사이클", "total"],
    "errors":    ["일일에러", "에러", "errors"],
    "streak":    ["연속성공", "연속", "streak"],
    "hours":     ["가동시간", "런시간", "hours"],   # Pilot~: 무정지 런(시간 잣대) 산출용 (선택)
    "notes":     ["비고", "메모", "notes"],
}
ERROR_FIELD_ALIASES = {
    "no":     ["no", "번호", "순번", "no."],
    "date":   ["발생일", "일자", "date"],
    "time":   ["시각", "시간", "time"],
    "cycle":  ["회차", "사이클", "cycle"],
    "code":   ["코드", "code"],
    "type":   ["유형", "타입", "type"],
    "detail": ["상세", "상세내용", "detail"],
    "cause":  ["원인", "cause"],
    # Pilot~: 근본원인 분류 축 (설계/부품/제작·조립/SW/시험환경·자재/운영·조작) — 선택 컬럼.
    #   POC 4분류의 상위호환 세분화 (docs/RECORD_SCHEMA.md §3). 공통 레코드 스토어의 cause 로 매핑.
    "cause_class": ["원인분류", "근본원인분류"],
    "action": ["조치", "조치사항", "action"],
    "result": ["결과", "조치결과", "result"],
    # 삼성 담당자 / 업체 담당자 — 둘 다 "담당"을 포함하므로 구체적인 후보를 먼저 둔다.
    # owner_sec(삼성)를 owner(업체)보다 먼저 매핑해 교차 매칭을 방지.
    "owner_sec": ["삼성담당자", "삼성담당", "삼성", "secowner", "sec"],
    "owner":     ["업체담당자", "업체담당", "협력사담당", "업체", "협력사", "vendor", "담당", "owner"],
    # Pilot~: 발생 시점 형상(버전) — "구버전 고장" 입증 수단 (docs/RECORD_SCHEMA.md #9)
    "sw_ver": ["sw버전", "sw버젼", "소프트웨어버전", "swver"],
    "hw_ver": ["hw버전", "hw버젼", "하드웨어버전", "hwver"],
    # 업체가 입력하는 확장 자료 (선택) — "더 상세" 모달에서만 표시.
    #   detail_more : 긴 설명 텍스트
    #   images      : 사진 파일명(쉼표/줄바꿈 구분). 실제 파일은 data/errors/ 폴더에 둔다.
    "detail_more": ["상세설명", "추가상세", "상세자료", "추가설명", "detailmore"],
    "images":      ["사진", "이미지", "첨부파일", "첨부", "파일명", "image", "photo", "attachment"],
}

# ── POC 전용 시트 (템플릿① POC 모드 — 보고 부담 최소화: 필수 5필드) ──
ISSUE_FIELD_ALIASES = {
    "id":       ["이슈id", "이슈번호", "id", "no"],
    "mode":     ["고장모드", "모드", "유형"],
    "severity": ["심각도", "등급"],
    "cause4":   ["원인분류", "4분류", "분류"],
    "status":   ["상태"],
    "date":     ["발생일", "일자"],
    # 종결일·무발생검증은 POC 선택 필드 — 있으면 수렴 추이·폐루프 표시에 사용 (docs/RECORD_SCHEMA.md §1)
    "closedDate": ["종결일", "종결날짜"],
    "verify":   ["무발생검증", "무발생", "검증진행"],
    # 확산/운영: 호기/라인(함대 층화)·다운타임(비용 Pareto) — 해당 단계 필수 (docs/RECORD_SCHEMA.md #14)
    "unit":     ["호기", "호기/라인"],
    "downtime": ["다운타임(분)", "다운타임"],
    "detail":   ["상세", "내용"],
    "images":   ["사진", "이미지"],
}
# 확산: 호기별 양산 퀄 (설치→SAT→축약 무고장 런)
UNIT_FIELD_ALIASES = {
    "unit":      ["호기"],
    "line":      ["라인"],
    "installDate": ["설치일"],
    "sat":       ["sat", "sat판정"],
    "runH":      ["축약런", "런h", "축약런(h)"],
    "runTarget": ["목표(h)", "목표h", "목표"],
    "status":    ["상태"],
    "notes":     ["비고"],
}
# 운영: 월간 RAM 지표 (통합관제 산출)
MONTHLY_FIELD_ALIASES = {
    "month":    ["월", "month"],
    "avail":    ["가동률", "가동률(%)"],
    "mtbf":     ["mtbf", "mtbf(h)"],
    "mttr":     ["mttr", "mttr(분)", "mttr(m)"],
    "downtime": ["다운타임(h)", "다운타임"],
    "alarms":   ["알람수", "알람"],
    "promoted": ["승격수", "fracas승격", "승격"],
}
# 운영: CIP (개선과제)
CIP_FIELD_ALIASES = {
    "id":     ["cipid", "cip", "id"],
    "topic":  ["과제", "개선과제", "주제"],
    "target": ["대상모드", "대상"],
    "status": ["상태"],
    "effect": ["기대효과", "효과"],
}
RUN_FIELD_ALIASES = {
    "date":   ["일자", "날짜", "평가일"],
    "hours":  ["런시간", "가동시간", "시간"],
    "errors": ["에러수", "에러"],
    "notes":  ["비고", "메모"],
}
ABN_FIELD_ALIASES = {
    "scenario": ["시나리오", "항목"],
    "recovery": ["복구시간", "복구"],
    "verdict":  ["판정", "결과"],
    "notes":    ["비고"],
}

DAILY_SHEET_KEYWORDS  = ["일일평가", "일일", "daily"]
ERRORS_SHEET_KEYWORDS = ["에러로그", "에러", "error"]
ISSUES_SHEET_KEYWORDS = ["이슈로그", "이슈", "issue"]
RUNLOG_SHEET_KEYWORDS = ["런기록", "런로그", "run"]
ABN_SHEET_KEYWORDS    = ["비정상평가", "비정상", "abnormal"]
UNITQ_SHEET_KEYWORDS  = ["호기퀄", "호기", "unitqual"]
MONTHLY_SHEET_KEYWORDS = ["월간지표", "월간", "monthly"]
CIP_SHEET_KEYWORDS    = ["cip", "개선과제", "개선"]


def _norm(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", "", str(s)).lower()


def _find_name(names: list[str], keywords: list[str]) -> str | None:
    """시트명 리스트에서 keywords 중 하나라도 포함된 첫 이름을 반환."""
    norm_map = {n: _norm(n) for n in names}
    for kw in keywords:
        kw_n = _norm(kw)
        for name in names:
            if kw_n in norm_map[name]:
                return name
    return None


def _build_column_map(header_row, aliases: dict[str, list[str]]) -> dict[str, int]:
    """헤더 행을 보고 {필드명: 컬럼인덱스} 매핑을 만든다.
    1차 정확 일치 → 2차 부분 일치 — '처분' vs '처분ID' 같은 접두 충돌을 방지한다."""
    norm_cells = [_norm(c) for c in header_row]
    col_map: dict[str, int] = {}
    for field, candidates in aliases.items():
        for cand in candidates:                     # 1차: 정확 일치 우선
            cand_n = _norm(cand)
            for idx, cell in enumerate(norm_cells):
                if cell == cand_n:
                    col_map[field] = idx
                    break
            if field in col_map:
                break
        if field in col_map:
            continue
        for cand in candidates:                     # 2차: 부분 일치
            cand_n = _norm(cand)
            for idx, cell in enumerate(norm_cells):
                if cand_n and cand_n in cell:
                    col_map[field] = idx
                    break
            if field in col_map:
                break
    return col_map


def _find_header_row(rows: list[list], aliases: dict[str, list[str]]) -> int:
    """선두 최대 10행에서 헤더처럼 보이는 행 인덱스를 찾는다.
    vendor가 상단에 제목/병합 안내를 넣었을 때 1행이 아닐 수 있어서 자동 감지.
    헤더 후보 매칭 개수가 가장 많은 행을 선택. 매칭이 전혀 없으면 0(첫행) fallback.
    """
    best_idx, best_hits = 0, 0
    candidates_flat: list[str] = []
    for cands in aliases.values():
        candidates_flat.extend(_norm(c) for c in cands if c)
    for idx, row in enumerate(rows[:10]):
        if row is None:
            continue
        norm_cells = [_norm(c) for c in row]
        hits = 0
        for cell in norm_cells:
            if not cell:
                continue
            for cand_n in candidates_flat:
                if cell == cand_n or (cand_n and cand_n in cell):
                    hits += 1
                    break
        if hits > best_hits:
            best_idx, best_hits = idx, hits
    return best_idx


_DATE_TEXT_RE = re.compile(
    r"^\s*(\d{2,4})[\s\.\-/년]+(\d{1,2})[\s\.\-/월]+(\d{1,2})\s*일?\s*$"
)


def _try_normalize_date(s: str) -> str:
    """텍스트로 들어온 날짜를 YYYY-MM-DD 로 정규화. 매칭 안 되면 원문 반환.
    지원 포맷: 2026-06-01, 2026/06/01, 2026.6.1, 26-6-1, 2026년 6월 1일 등.
    """
    m = _DATE_TEXT_RE.match(s)
    if not m:
        return s
    y, mo, d = m.group(1), m.group(2), m.group(3)
    if len(y) == 2:
        y = f"20{y}"
    try:
        return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
    except ValueError:
        return s


_EXCEL_EPOCH = datetime(1899, 12, 30)   # Excel 1900 leap-year 버그 보정 포함


def _cell_to_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, time):
        return v.strftime("%H:%M")
    # Excel 직렬 날짜 — 셀 서식이 '일반/숫자'면 float로 넘어옴.
    # 20000~80000 범위면 1954~2118년 사이 → 날짜로 해석. (bool은 int 서브클래스라 명시적으로 제외)
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            f = float(v)
            if 20000 < f < 80000:
                return (_EXCEL_EPOCH + timedelta(days=f)).strftime("%Y-%m-%d")
        except (OverflowError, ValueError):
            pass
    s = str(v).strip()
    # 날짜처럼 보이면 정규화 시도 (다양한 구분자/생략 연도 등 대응)
    return _try_normalize_date(s)


def _cell_to_int(v) -> int:
    if v is None or v == "":
        return 0
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


_TIME_TEXT_RE = re.compile(r"^\s*(?:(?:오전|오후|am|pm)\s*)?(\d{1,2})\s*[:시]\s*(\d{1,2})")


def _cell_to_time(v) -> str:
    """에러로그 '시각' 셀 → 'h:mm' 문자열.
    업체가 custom h:mm 서식으로 보내면 openpyxl이 시간을 '하루의 분수'(0~1 float)로
    넘겨 셀 값이 0.xxxxx 로 깨지는 문제를 보정한다. (업체 입력 방식 h:mm 에 맞춰 복원)
    """
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return f"{v.hour}:{v.minute:02d}"
    if isinstance(v, time):
        return f"{v.hour}:{v.minute:02d}"
    # 숫자 — Excel 시간 직렬값. 정수부=날짜, 소수부=하루 중 비율. 소수부만 분으로 환산.
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            frac = float(v) % 1.0
        except (OverflowError, ValueError):
            frac = None
        if frac is not None:
            total_min = round(frac * 1440) % 1440   # 분 단위 반올림 + 자정 롤오버 보정
            h, m = divmod(total_min, 60)
            return f"{h}:{m:02d}"
    # 문자열 — 이미 'HH:MM'/'오후 2:32' 등으로 들어온 경우 h:mm 로 정리해서 반환
    s = str(v).strip()
    mt = _TIME_TEXT_RE.match(s)
    if mt:
        h = int(mt.group(1))
        ampm = re.match(r"^\s*(오후|pm)", s, re.IGNORECASE)
        if ampm and h < 12:
            h += 12
        return f"{h}:{int(mt.group(2)):02d}"
    return s


def _split_images(v) -> list[str]:
    """'사진' 셀 → 파일명 리스트. 쉼표/줄바꿈/세미콜론 구분. 빈 값은 제거.
    실제 이미지 파일은 data/errors/ 폴더에 같은 이름으로 둔다(프론트가 그 경로로 로드)."""
    if v is None:
        return []
    parts = re.split(r"[,\n;]+", str(v))
    return [p.strip() for p in parts if p.strip()]


def _safe_idx(row: list, idx: int):
    """row가 짧으면 None 반환 — vendor가 컬럼 수를 다르게 보낼 때 IndexError 방지."""
    return row[idx] if idx is not None and 0 <= idx < len(row) else None


def _parse_daily(rows: list[list]) -> list[dict]:
    if not rows:
        return []
    header_idx = _find_header_row(rows, DAILY_FIELD_ALIASES)
    header = rows[header_idx]
    body = rows[header_idx + 1:]
    cmap = _build_column_map(header, DAILY_FIELD_ALIASES)
    if "date" not in cmap or "total" not in cmap:
        raise SystemExit(
            f"[일일평가] 시트에서 필수 컬럼(평가일/일일평가)을 찾지 못했습니다. "
            f"감지된 헤더(행 {header_idx + 1}): {[str(h) for h in header]}"
        )

    out = []
    for row in body:
        if row is None or all(c is None or c == "" for c in row):
            continue
        date_val = _cell_to_str(_safe_idx(row, cmap["date"]))
        if not date_val:
            continue
        # YYYY-MM-DD 형식이 아니면 skip (헤더 잔여 행이나 잘못된 행 한 번 더 거름)
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", date_val):
            continue
        out.append({
            "date":      date_val,
            "personnel": _cell_to_str(_safe_idx(row, cmap.get("personnel"))),
            "activity":  _cell_to_str(_safe_idx(row, cmap.get("activity"))),
            "total":     max(0, _cell_to_int(_safe_idx(row, cmap["total"]))),
            "errors":    max(0, _cell_to_int(_safe_idx(row, cmap.get("errors")))),
            "streak":    max(0, _cell_to_int(_safe_idx(row, cmap.get("streak")))),
            "hours":     max(0, _cell_to_int(_safe_idx(row, cmap.get("hours")))),
            "notes":     _cell_to_str(_safe_idx(row, cmap.get("notes"))),
        })
    out.sort(key=lambda r: r["date"])
    return out


def _parse_errors(rows: list[list]) -> list[dict]:
    if not rows:
        return []
    header_idx = _find_header_row(rows, ERROR_FIELD_ALIASES)
    header = rows[header_idx]
    body = rows[header_idx + 1:]
    cmap = _build_column_map(header, ERROR_FIELD_ALIASES)
    out = []
    for row in body:
        if row is None or all(c is None or c == "" for c in row):
            continue
        no_val = _cell_to_int(_safe_idx(row, cmap.get("no")))
        date_val = _cell_to_str(_safe_idx(row, cmap.get("date")))
        if not no_val and not date_val:
            continue
        out.append({
            "no":        no_val,
            "date":      date_val,
            "time":      _cell_to_time(_safe_idx(row, cmap.get("time"))),
            "cycle":     _cell_to_int(_safe_idx(row, cmap.get("cycle"))),
            "code":      _cell_to_str(_safe_idx(row, cmap.get("code"))),
            "type":      _cell_to_str(_safe_idx(row, cmap.get("type"))),
            "detail":    _cell_to_str(_safe_idx(row, cmap.get("detail"))),
            "cause":     _cell_to_str(_safe_idx(row, cmap.get("cause"))),
            "cause_class": _cell_to_str(_safe_idx(row, cmap.get("cause_class"))),
            "action":    _cell_to_str(_safe_idx(row, cmap.get("action"))),
            "result":    _cell_to_str(_safe_idx(row, cmap.get("result"))),
            "owner_sec": _cell_to_str(_safe_idx(row, cmap.get("owner_sec"))),
            "owner":     _cell_to_str(_safe_idx(row, cmap.get("owner"))),
            # Pilot~: 발생 시점 형상 (선택 컬럼 — 없으면 빈 값)
            "sw_ver":    _cell_to_str(_safe_idx(row, cmap.get("sw_ver"))),
            "hw_ver":    _cell_to_str(_safe_idx(row, cmap.get("hw_ver"))),
            # 확장 자료(선택): 더 상세 모달에서만 사용
            "detailMore": _cell_to_str(_safe_idx(row, cmap.get("detail_more"))),
            "images":     _split_images(_safe_idx(row, cmap.get("images"))),
        })
    out.sort(key=lambda r: r.get("no", 0))
    return out


def _pick_latest_xlsx() -> Path:
    xlsxs = sorted(
        [p for p in RAW_DIR.glob("*.xlsx") if not p.name.startswith("~$")],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not xlsxs:
        raise SystemExit(f"data/raw/ 폴더에 .xlsx 파일이 없습니다. ({RAW_DIR})")
    if len(xlsxs) > 1:
        print(f"[build] ⚠ data/raw/ 에 xlsx 파일이 {len(xlsxs)}개 있습니다. 최신(mtime) 파일을 사용합니다:")
        for p in xlsxs:
            ts = datetime.fromtimestamp(p.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
            marker = "→" if p == xlsxs[0] else " "
            print(f"        {marker} {p.name}  (mtime {ts})")
        print(f"        다른 파일을 쓰려면 불필요한 파일을 제거하거나 사용할 파일을 다시 저장(터치)하세요.")
    return xlsxs[0]


# ── 시트 로딩: openpyxl 우선, 실패하면 xlwings ────────────────────
def _load_via_openpyxl(src: Path) -> tuple[list[list], list[list], list[str]]:
    from openpyxl import load_workbook
    wb = load_workbook(src, data_only=True)
    names = wb.sheetnames
    daily_name  = _find_name(names, DAILY_SHEET_KEYWORDS)
    errors_name = _find_name(names, ERRORS_SHEET_KEYWORDS)

    if daily_name is None:
        raise SystemExit(f"'일일평가' 시트를 찾지 못했습니다. 시트 목록: {names}")

    daily_rows  = [list(r) for r in wb[daily_name].iter_rows(values_only=True)]
    errors_rows = (
        [list(r) for r in wb[errors_name].iter_rows(values_only=True)]
        if errors_name else []
    )
    return daily_rows, errors_rows, names


def _xlwings_sheet_rows(sheet) -> list[list]:
    rng = sheet.used_range
    val = rng.value
    if val is None:
        return []
    if not isinstance(val, list):
        return [[val]]
    if val and not isinstance(val[0], list):
        # 1D 결과 — 단일 행 or 단일 열
        if rng.rows.count == 1:
            return [val]
        return [[v] for v in val]
    return val


def _load_via_xlwings(src: Path) -> tuple[list[list], list[list], list[str]]:
    try:
        import xlwings as xw
    except ImportError as e:
        raise SystemExit(
            "xlwings 미설치. DRM 보호 파일을 읽으려면 'pip install xlwings' 후 재시도. "
            "(Windows + Excel 설치 필수)"
        ) from e

    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    try:
        wb = app.books.open(str(src), update_links=False, read_only=True)
        try:
            names = [s.name for s in wb.sheets]
            daily_name  = _find_name(names, DAILY_SHEET_KEYWORDS)
            errors_name = _find_name(names, ERRORS_SHEET_KEYWORDS)

            if daily_name is None:
                raise SystemExit(f"'일일평가' 시트를 찾지 못했습니다. 시트 목록: {names}")

            daily_rows  = _xlwings_sheet_rows(wb.sheets[daily_name])
            errors_rows = (
                _xlwings_sheet_rows(wb.sheets[errors_name])
                if errors_name else []
            )
            return daily_rows, errors_rows, names
        finally:
            wb.close()
    finally:
        app.quit()


def _load_workbook_rows(src: Path) -> tuple[list[list], list[list]]:
    try:
        daily_rows, errors_rows, _ = _load_via_openpyxl(src)
        return daily_rows, errors_rows
    except zipfile.BadZipFile:
        # DRM 래핑 추정 — openpyxl은 zip 구조가 아니라고 거부함
        print("[build] openpyxl 실패 (DRM 추정) → xlwings로 Excel 통한 재시도")
        daily_rows, errors_rows, _ = _load_via_xlwings(src)
        return daily_rows, errors_rows


# ── 전체 시트 로더 (POC 등 시트 구성이 다른 단계용) — openpyxl → xlwings 폴백 ──
def _load_all_sheets(src: Path) -> dict[str, list[list]]:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(src, data_only=True)
        return {n: [list(r) for r in wb[n].iter_rows(values_only=True)] for n in wb.sheetnames}
    except zipfile.BadZipFile:
        print("[build] openpyxl 실패 (DRM 추정) → xlwings로 Excel 통한 재시도")
        import xlwings as xw
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        try:
            wb = app.books.open(str(src), update_links=False, read_only=True)
            try:
                return {s.name: _xlwings_sheet_rows(s) for s in wb.sheets}
            finally:
                wb.close()
        finally:
            app.quit()


def _pick_sheet(sheets: dict[str, list[list]], keywords: list[str]) -> list[list]:
    name = _find_name(list(sheets.keys()), keywords)
    return sheets.get(name, []) if name else []


def _parse_generic(rows: list[list], aliases: dict, required: str) -> list[dict]:
    """별칭 사전 기반 범용 시트 파서 — required 필드가 빈 행은 스킵."""
    if not rows:
        return []
    header_idx = _find_header_row(rows, aliases)
    cmap = _build_column_map(rows[header_idx], aliases)
    out = []
    for row in rows[header_idx + 1:]:
        if row is None or all(c is None or c == "" for c in row):
            continue
        rec = {}
        for field in aliases:
            v = _safe_idx(row, cmap.get(field))
            if field in ("hours", "errors"):
                rec[field] = max(0, _cell_to_int(v))
            elif field == "images":
                rec[field] = _split_images(v)
            elif field == "date":
                rec[field] = _cell_to_str(v)
            else:
                rec[field] = _cell_to_str(v)
        if not rec.get(required):
            continue
        out.append(rec)
    return out


# ── 관리 데이터(config + SEC_REPORT.xlsx) 로딩 ─────────────────────────
def _load_config() -> dict:
    if not CONFIG_PATH.exists():
        return {}
    try:
        return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    except Exception as e:
        print(f"[build] ⚠ config.json 읽기 실패: {e}")
        return {}


CODE_FIELDS = {"code": ["코드"], "type": ["유형"], "severity": ["등급", "심각도"], "desc": ["설명"]}
ACTION_FIELDS = {"id": ["조치id"], "code": ["대상코드"], "action": ["조치내용"],
                 "owner": ["담당"], "due": ["목표일"], "status": ["상태"],
                 # 검증시작: 누적 cycle(숫자) 또는 검증 시작일(날짜) 둘 다 허용.
                 #   날짜를 적으면 빌드가 그날까지의 누적 Cycle로 자동 환산한다.
                 "verifyStart": ["검증시작cycle", "검증시작일", "검증시작", "검증일", "조치완료일"]}
# 판정대장 (양산 시범 평가~): 관련/비관련 합동판정 — docs/CRITERIA.md §4
ADJUD_FIELDS = {"id": ["사건id", "사건"], "target": ["대상에러no", "대상에러", "대상"],
                "verdict": ["판정"], "attribution": ["귀책분류", "귀책"],
                "evidence": ["증거"], "agreed": ["합의상태", "합의"], "date": ["판정일"]}
# 처분대장 (인증 준비): 오픈 건 전건 처분(종결예정/carry-over/waiver) → Known Issues Register 단일본
#   — docs/PROCESS.md §2.4 (기한·책임자 서명)
DISPO_FIELDS = {"id": ["처분id", "id"], "target": ["대상id", "대상에러no", "대상에러", "대상코드", "대상"],
                "dispo": ["처분"], "reason": ["사유"], "due": ["기한", "목표일"],
                "owner": ["오너", "책임자", "담당"], "agreed": ["합의", "서명", "합의상태"]}


def _norm_sev(v) -> str:
    s = _cell_to_str(v).lower()
    if "crit" in s or "치명" in s:
        return "Critical"
    if "maj" in s or "중대" in s:
        return "Major"
    if "min" in s or "경미" in s:
        return "Minor"
    return _cell_to_str(v) or "Minor"


def _read_mgmt_rows(rows, fields) -> list[dict]:
    if not rows:
        return []
    hidx = _find_header_row(rows, fields)
    cmap = _build_column_map(rows[hidx], fields)
    out = []
    for row in rows[hidx + 1:]:
        if row is None or all(c is None or c == "" for c in row):
            continue
        out.append({f: _safe_idx(row, cmap.get(f)) for f in fields})
    return out


def _load_mgmt_rows_openpyxl() -> tuple[list[list], list[list], list[list]]:
    from openpyxl import load_workbook
    wb = load_workbook(MGMT_PATH, data_only=True)
    names = wb.sheetnames
    cm = _find_name(names, ["코드마스터", "코드", "code"])
    am = _find_name(names, ["조치검증", "조치", "action"])
    jm = _find_name(names, ["판정대장", "판정", "adjud"])
    dm = _find_name(names, ["처분대장", "처분", "dispo"])
    code_rows   = [list(r) for r in wb[cm].iter_rows(values_only=True)] if cm else []
    action_rows = [list(r) for r in wb[am].iter_rows(values_only=True)] if am else []
    adjud_rows  = [list(r) for r in wb[jm].iter_rows(values_only=True)] if jm else []
    dispo_rows  = [list(r) for r in wb[dm].iter_rows(values_only=True)] if dm else []
    return code_rows, action_rows, adjud_rows, dispo_rows


def _load_mgmt_rows_xlwings() -> tuple[list[list], list[list]]:
    try:
        import xlwings as xw
    except ImportError as e:
        raise SystemExit(
            "xlwings 미설치. DRM 보호된 SEC_REPORT.xlsx를 읽으려면 'pip install xlwings' 후 재시도. "
            "(Windows + Excel 설치 필수)"
        ) from e

    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    try:
        wb = app.books.open(str(MGMT_PATH), update_links=False, read_only=True)
        try:
            names = [s.name for s in wb.sheets]
            cm = _find_name(names, ["코드마스터", "코드", "code"])
            am = _find_name(names, ["조치검증", "조치", "action"])
            jm = _find_name(names, ["판정대장", "판정", "adjud"])
            dm = _find_name(names, ["처분대장", "처분", "dispo"])
            code_rows   = _xlwings_sheet_rows(wb.sheets[cm]) if cm else []
            action_rows = _xlwings_sheet_rows(wb.sheets[am]) if am else []
            adjud_rows  = _xlwings_sheet_rows(wb.sheets[jm]) if jm else []
            dispo_rows  = _xlwings_sheet_rows(wb.sheets[dm]) if dm else []
            return code_rows, action_rows, adjud_rows, dispo_rows
        finally:
            wb.close()
    finally:
        app.quit()


def _load_mgmt() -> tuple[list[dict], list[dict], list[dict], list[dict]]:
    if not MGMT_PATH.exists():
        print("[build] REPORT.xlsx 없음 — 코드마스터/조치 없이 진행 (generate_mgmt_template.py로 생성)")
        return [], [], [], []
    try:
        code_rows, action_rows, adjud_rows, dispo_rows = _load_mgmt_rows_openpyxl()
    except zipfile.BadZipFile:
        # DRM 래핑 추정 — openpyxl은 zip 구조가 아니라고 거부함 (사내 보안문서 저장 시)
        print("[build] REPORT.xlsx openpyxl 실패 (DRM 추정) → xlwings로 Excel 통한 재시도")
        code_rows, action_rows, adjud_rows, dispo_rows = _load_mgmt_rows_xlwings()
    raw_c = _read_mgmt_rows(code_rows, CODE_FIELDS)
    raw_a = _read_mgmt_rows(action_rows, ACTION_FIELDS)
    raw_j = _read_mgmt_rows(adjud_rows, ADJUD_FIELDS)
    codes = [{"code": _cell_to_str(c["code"]), "type": _cell_to_str(c["type"]),
              "severity": _norm_sev(c["severity"]), "desc": _cell_to_str(c["desc"])}
             for c in raw_c if _cell_to_str(c["code"])]
    actions = []
    for a in raw_a:
        if not _cell_to_str(a["id"]):
            continue
        # 검증시작 칸: 날짜(YYYY-MM-DD)면 verifyStartDate로 넘겨 _compute에서 누적 Cycle로
        # 환산하고, 숫자면 기존처럼 verifyStart(cycle)로 직접 사용한다.
        vs_str = _cell_to_str(a["verifyStart"])
        if re.match(r"^\d{4}-\d{2}-\d{2}$", vs_str):
            vs_cycle, vs_date = 0, vs_str
        else:
            vs_cycle, vs_date = _cell_to_int(a["verifyStart"]), ""
        actions.append({"id": _cell_to_str(a["id"]), "code": _cell_to_str(a["code"]),
                        "action": _cell_to_str(a["action"]), "owner": _cell_to_str(a["owner"]),
                        "due": _cell_to_str(a["due"]), "status": _cell_to_str(a["status"]),
                        "verifyStart": vs_cycle, "verifyStartDate": vs_date})
    adjud = [{"id": _cell_to_str(j["id"]), "target": _cell_to_str(j["target"]),
              "verdict": _cell_to_str(j["verdict"]), "attribution": _cell_to_str(j["attribution"]),
              "evidence": _cell_to_str(j["evidence"]), "agreed": _cell_to_str(j["agreed"]),
              "date": _cell_to_str(j["date"])}
             for j in raw_j if _cell_to_str(j["id"])]
    raw_d = _read_mgmt_rows(dispo_rows, DISPO_FIELDS)
    dispositions = [{"id": _cell_to_str(d["id"]), "target": _cell_to_str(d["target"]),
                     "dispo": _cell_to_str(d["dispo"]), "reason": _cell_to_str(d["reason"]),
                     "due": _cell_to_str(d["due"]), "owner": _cell_to_str(d["owner"]),
                     "agreed": _cell_to_str(d["agreed"])}
                    for d in raw_d if _cell_to_str(d["id"])]
    return codes, actions, adjud, dispositions


def _iso_week(date_str: str):
    try:
        y, w, _ = date.fromisoformat(date_str).isocalendar()
        return (y, w)
    except Exception:
        return None


def _occ_band(n: int) -> str:
    return "빈발" if n >= 6 else ("보통" if n >= 3 else "드묾")


_PRIORITY = {
    ("Critical", "드묾"): "Medium", ("Critical", "보통"): "High", ("Critical", "빈발"): "High",
    ("Major", "드묾"): "Low", ("Major", "보통"): "Medium", ("Major", "빈발"): "High",
    ("Minor", "드묾"): "Low", ("Minor", "보통"): "Low", ("Minor", "빈발"): "Medium",
}


# ── 파생 지표 계산 (신뢰성 대시보드 v5) ──────────────────────────
def _compute(daily, errors, config, codes, actions, now=None) -> dict:
    acc = config.get("acceptance") or {}
    target = acc.get("targetCycle") or config.get("project", {}).get("target", 360)
    mtbf_t = acc.get("mtbfTargetCycle", 100)
    conf_lv = acc.get("confidenceLevel", 0.80)
    verify_cy = acc.get("verifyCycle", 200)
    recur_lim = acc.get("recurrenceLimit", 0)

    cum_total = sum(d["total"] for d in daily)
    cum_err = sum(d["errors"] for d in daily)
    success = max(0, cum_total - cum_err)
    # ── 연속 성공 누적 재계산 + 에러 위치 산출 (업체 세그먼트 양식 대응) ──
    # 업체는 하루를 에러 기준으로 여러 행(세그먼트)으로 쪼개 적고, '연속성공'은 그 세그먼트의
    # 당일 연속 성공 값이다(누적 아님). 그래서 연속성공 컬럼을 그대로 믿지 않고
    #   · 누적 연속 = (total − 에러)를 행 순서대로 더하되, 에러가 있는 행 끝에서 0으로 리셋
    #   · 에러 위치(누적 Cycle) = 성공(연속달성) 뒤에 에러가 온다고 보고 직접 산출
    # 으로 재계산한다. 현재 횟수(에러버짓 리셋)도 이 위치를 쓴다(에러로그 회차 비의존).
    run = 0; streak_max = 0; _cum = 0
    err_cycles = []
    for d in daily:                       # daily는 날짜순(동일 날짜는 입력 순서 유지)
        succ = max(0, d["total"] - d["errors"])
        peak = run + succ                 # 세그먼트 내 성공이 앞쪽 → 끊기기 직전 누적 연속
        streak_max = max(streak_max, peak)
        for k in range(d["errors"]):
            err_cycles.append(_cum + succ + 1 + k)
        run = 0 if d["errors"] else peak  # 에러 있는 행은 끝에서 리셋
        d["_runStreak"] = run
        _cum += d["total"]
    streak_cur = run
    if len(err_cycles) != len(errors):
        print(f"[build] ⚠ 일일평가 에러 합({len(err_cycles)}) ≠ 에러로그 행수({len(errors)}) "
              f"— 현재횟수는 일일평가 기준으로 산출")

    # ── 고장(failure) 판정 ──────────────────────────────────────
    # 에러 ≠ 고장. 에러로그의 코드/유형에 failureKeyword("고장")가 포함된 항목만 '진짜 고장'.
    # MTBF·신뢰수준(무고장 시험)은 에러버짓·연속성공과 무관하게 '고장' 기준으로만 리셋된다
    #  → 일반 에러가 나도 무고장 연속 Cycle은 계속 올라가고, '고장'이 찍힌 순간에만 0으로 끊긴다.
    fail_kw = acc.get("failureKeyword", "고장")
    def _is_failure(e):
        return bool(fail_kw) and fail_kw in f"{e.get('code','')} {e.get('type','')}"
    failures = [e for e in errors if _is_failure(e)]
    failure_count = len(failures)
    last_fail_cycle = max((e["cycle"] for e in failures if e.get("cycle")), default=0)
    failure_free = max(0, cum_total - last_fail_cycle)   # 마지막 고장 이후 누적 Cycle(고장 0건이면 전체)
    mtbf_cur = round(cum_total / failure_count) if failure_count else cum_total
    # 기간 에러율(%) = 에러(인터럽트) 발생률. 목표는 config.acceptance.errRateTargetPct(최근 4주 기준 <N%).
    err_rate_cur = round(cum_err / cum_total * 100, 1) if cum_total else 0
    err_rate_t = acc.get("errRateTargetPct")

    # ── 에러 버짓 리셋 모델 ──────────────────────────────────────
    # 한 '시도(attempt)'는 연속 target Cycle + 에러 ≤ error_limit 를 동시에 만족해야 합격.
    # error_limit 를 초과하는 에러(= error_limit+1번째)가 나오면 그 시도는 실패 →
    # 해당 에러 지점부터 새 시도 시작(진행 Cycle·버짓 0). 진행률은 '현재 시도' 기준.
    error_limit = acc.get("errorLimit", config.get("project", {}).get("errorLimit", 3))
    attempt_start = 0          # 현재 시도가 시작된 누적 Cycle
    budget_used = 0            # 현재 시도의 누적 에러 (0~error_limit)
    reset_cycles = []          # 버짓 초과로 시도가 리셋된 누적 Cycle 위치들
    for c in err_cycles:                       # 일일평가 세그먼트에서 산출한 에러 위치
        if budget_used + 1 > error_limit:      # 다음 에러가 한도 초과 → 시도 실패·리셋
            reset_cycles.append(c)
            attempt_start = c                  # 이 에러 지점부터 새 시도 (이 에러는 실패로 소진)
            budget_used = 0
        else:
            budget_used += 1
    budget_resets = len(reset_cycles)          # 버짓 초과로 시도가 리셋된 횟수
    attempt_cycles = max(0, cum_total - attempt_start)   # 현재 시도 진행 Cycle

    # 주차별 누적연속 + 안정성(에러율↓/MTBF↑)
    weekly = {}
    for d in sorted(daily, key=lambda x: x["date"]):
        wk = _iso_week(d["date"])
        if wk is None:
            continue
        w = weekly.setdefault(wk, {"total": 0, "errors": 0, "lastStreak": 0})
        w["total"] += d["total"]; w["errors"] += d["errors"]
        w["lastStreak"] = d.get("_runStreak", 0)   # 재계산한 누적 연속(컬럼 아님)
    for d in daily:
        d.pop("_runStreak", None)                  # 내부 계산용 필드 — 출력 JSON에는 남기지 않음
    # 주차별 고장 건수(MTBF=Mean Cycles Between Failures 산출용 — 에러 아님)
    fail_wk = {}
    for e in failures:
        wk = _iso_week(e.get("date", "")) if e.get("date") else None
        if wk is not None:
            fail_wk[wk] = fail_wk.get(wk, 0) + 1
    weekly_list = []
    run_t = run_e = run_f = 0
    for i, (key, w) in enumerate(sorted(weekly.items()), start=1):
        prev_t = run_t                                        # 주 시작 시점 누적 Cycle
        run_t += w["total"]; run_e += w["errors"]; run_f += fail_wk.get(key, 0)
        # cumStreak = '현재 시도 진행 Cycle'(버짓 기준). 진행률 헤드라인(progress.cum)과
        # 동일한 리셋 모델을 써서 주 끝 시점 값을 산출한다 → 최신 주 = attempt_cycles 로 일치.
        wk_start = max((rc for rc in reset_cycles if rc <= run_t), default=0)
        cum_attempt = max(0, run_t - wk_start)
        reset_in_week = any(prev_t < rc <= run_t for rc in reset_cycles)  # 이 주에 버짓 리셋 발생?
        try:
            week_start = date.fromisocalendar(key[0], key[1], 1).isoformat()
        except Exception:
            week_start = ""
        weekly_list.append({
            "week": f"W{i}", "weekStart": week_start, "cumStreak": cum_attempt,
            "reset": reset_in_week,
            "weekSuccess": max(0, w["total"] - w["errors"]), "errors": w["errors"],
            "errRate": round(run_e / run_t * 100, 1) if run_t else 0,
            "mtbf": round(run_t / run_f) if run_f else run_t,
        })

    # ── 기간별 에러율 안정화 (에러발생확률↓ → 안정화 입증) ────────────
    # '에러 건수'가 아니라 '정규화 에러율(건/100Cycle)'을 기간별로 봐서 추세가 내려가는지(안정화)를 본다.
    #  · rate    = 그 기간 실측 에러율(건/100Cy)  → 막대
    #  · cumRate = 시작~그 기간 누적 평균 에러율   → 안정화 추세선
    # 기본 월별(config ui.steps.errRateBin="month"). 표본이 짧으면 "week"로 주별 추세 확인 가능.
    err_bin = str(config.get("ui", {}).get("steps", {}).get("errRateBin", "month")).lower()
    buckets, order = {}, []
    for d in sorted(daily, key=lambda x: x["date"]):
        ds = str(d["date"])[:10]
        key = _iso_week(ds) if err_bin == "week" else ds[:7]   # 주(y,w) 또는 'YYYY-MM'
        if key is None:
            continue
        if key not in buckets:
            buckets[key] = {"cycles": 0, "errors": 0, "d0": ds, "d1": ds}
            order.append(key)
        buckets[key]["cycles"] += d["total"]
        buckets[key]["errors"] += d["errors"]
        buckets[key]["d0"] = min(buckets[key]["d0"], ds)
        buckets[key]["d1"] = max(buckets[key]["d1"], ds)
    errrate_list = []
    er_t = er_e = 0
    _md = lambda s: f"{int(s[5:7])}/{int(s[8:10])}"          # 'YYYY-MM-DD' → 'M/D'
    # 주차 라벨을 '주차별 연속추이' 차트와 동일하게: 프로젝트 시작주(월요일) 기준 주차 번호
    _proj_start = config.get("project", {}).get("startDate")
    def _proj_week(iso_key):
        if not (err_bin == "week" and _proj_start and isinstance(iso_key, tuple)):
            return None
        try:
            bmon = date.fromisocalendar(iso_key[0], iso_key[1], 1)
            s = date.fromisoformat(_proj_start)
            smon = s - timedelta(days=s.weekday())
            return (bmon - smon).days // 7 + 1
        except Exception:
            return None
    for i, key in enumerate(order):
        b = buckets[key]
        er_t += b["cycles"]; er_e += b["errors"]
        _wn = _proj_week(key)
        label = (f"{_wn}주차" if _wn else f"{i + 1}주차") if err_bin == "week" else f"{int(key[5:7])}월"
        rng = _md(b["d0"]) if b["d0"] == b["d1"] else f"{_md(b['d0'])}~{_md(b['d1'])}"
        errrate_list.append({
            "period": label, "range": rng, "cycles": b["cycles"], "errors": b["errors"],
            "rate": round(b["errors"] / b["cycles"] * 100, 1) if b["cycles"] else 0,
            "cumRate": round(er_e / er_t * 100, 1) if er_t else 0,
            "mtbi": round(b["cycles"] / b["errors"]) if b["errors"] else b["cycles"],
        })

    # ── 최근 N주 롤링 윈도우 에러율 (업데이트일 기준 'as of now') ──────────
    # 양산 전환 시 월별 편차 보상: 캘린더로 쪼개지 않고 업데이트일에서 N주를 뒤로 묶어 본다.
    #   · 윈도우가 비면(데이터가 오래됨) 최신 데이터일 기준으로 폴백
    #   · 윈도우 Cycle이 recentMinCycles 미만이면 lowSample=True (표본부족 표시)
    _steps = config.get("ui", {}).get("steps", {})
    recent_weeks = int(_steps.get("recentWeeks", 4) or 4)
    recent_min = int(_steps.get("recentMinCycles", 20) or 0)
    anchor = (now or datetime.now(timezone.utc)).date()
    daily_dates = [d["date"][:10] for d in daily if d.get("date")]

    def _window(hi):
        lo = hi - timedelta(days=recent_weeks * 7 - 1)
        cyc = err = 0
        for d in daily:
            if lo.isoformat() <= str(d["date"])[:10] <= hi.isoformat():
                cyc += d["total"]; err += d["errors"]
        return lo, cyc, err

    win_from, rw_cyc, rw_err = _window(anchor)
    win_to, anchored_on = anchor, "update"
    if rw_cyc == 0 and daily_dates:                 # 업데이트일 기준 윈도우가 비면 최신 데이터일로 폴백
        win_to = date.fromisoformat(max(daily_dates))
        win_from, rw_cyc, rw_err = _window(win_to)
        anchored_on = "lastData"
    # 라벨용 시작일: 윈도우 내 '실제 데이터' 최초일 (윈도우가 데이터 이전까지 뻗어 과대표기되는 것 방지)
    _inwin = [d for d in daily_dates if win_from.isoformat() <= d <= win_to.isoformat()]
    disp_from = min(_inwin) if _inwin else win_from.isoformat()
    recent_window = {
        "weeks": recent_weeks,
        "fromDate": disp_from, "toDate": win_to.isoformat(),
        "cycles": rw_cyc, "errors": rw_err,
        "rate": round(rw_err / rw_cyc * 100, 1) if rw_cyc else 0,
        "mtbi": round(rw_cyc / rw_err) if rw_err else rw_cyc,
        "lowSample": bool(recent_min and rw_cyc < recent_min),
        "anchoredOn": anchored_on,
    }

    # 신뢰수준 (무고장 시험): C = 1 − e^(−n/MTBF목표)
    def required(c):
        return round(mtbf_t * (-math.log(1 - c))) if 0 < c < 1 else 0
    conf_cur = 1 - math.exp(-failure_free / mtbf_t) if mtbf_t else 0
    conf_table = [{"c": round(c * 100), "required": required(c)} for c in (0.80, 0.87, 0.90)]

    # 코드 집계 + 코드마스터(등급) 조인
    sev_of = {c["code"]: c["severity"] for c in codes}
    type_of = {c["code"]: c["type"] for c in codes}
    cnt = {}
    for e in errors:
        if e["code"]:
            cnt[e["code"]] = cnt.get(e["code"], 0) + 1

    top5 = sorted(cnt.items(), key=lambda kv: -kv[1])[:5]
    top5_by_code = [{"code": c, "type": type_of.get(c, ""), "count": n,
                     "severity": sev_of.get(c, "Minor"), "recur": n > 1} for c, n in top5]

    sev_dist = {"Critical": 0, "Major": 0, "Minor": 0}
    for e in errors:
        s = sev_of.get(e["code"], "Minor")
        sev_dist[s] = sev_dist.get(s, 0) + 1
    sev_dist["total"] = len(errors)

    tcnt = {}
    for e in errors:
        t = e["type"] or "기타"
        tcnt[t] = tcnt.get(t, 0) + 1
    pareto = []
    run = 0; tot = sum(tcnt.values()) or 1
    for t, n in sorted(tcnt.items(), key=lambda kv: -kv[1]):
        run += n
        pareto.append({"type": t, "count": n, "cumPct": round(run / tot * 100)})

    matrix = []
    for c, n in sorted(cnt.items(), key=lambda kv: -kv[1]):
        sev = sev_of.get(c, "Minor"); ob = _occ_band(n)
        matrix.append({"code": c, "type": type_of.get(c, ""), "severity": sev,
                       "count": n, "occ": ob, "priority": _PRIORITY.get((sev, ob), "Low")})

    # 재발 집계(recur_count)는 조치/검증 판정 뒤에서 계산한다 — 검증완료 코드를 제외해야 하므로.

    # 조치/검증 (조치 후 verify_cy Cycle 무발생 → 검증완료 자동판정)
    # 검증시작일(날짜)을 적은 경우 그 날짜까지의 누적 Cycle로 환산해 vs로 사용한다.
    def _cycle_at_date(dstr):
        return sum(d["total"] for d in daily if d["date"] <= dstr) if dstr else 0

    act_out = []
    open_critical = verified = relevant = 0
    for a in actions:
        c = a["code"]; vs = a["verifyStart"] or _cycle_at_date(a.get("verifyStartDate", ""))
        later_same = any(e["code"] == c and e["cycle"] > vs for e in errors) if vs else False
        no_fail = 0 if (later_same or not vs) else max(0, cum_total - vs)
        if later_same:
            result = "재발"
        elif vs and no_fail >= verify_cy:
            result = "검증완료"
        elif vs and no_fail > 0:
            result = "검증중"
        else:
            result = "조치중"
        # 수동 오버라이드: '상태' 칸에 '검증완료'라고 적으면 자동 판정과 무관하게 검증완료로 인정.
        #   (담당자가 200Cy 자동 도달 전에 직접 검증종결 처리할 때 사용)
        manual_done = "검증완료" in _cell_to_str(a.get("status"))
        if manual_done:
            result = "검증완료"
        sev = sev_of.get(c, "Minor")
        if sev in ("Critical", "Major"):
            relevant += 1
            if result == "검증완료":
                verified += 1
        if sev == "Critical" and result != "검증완료":
            open_critical += 1
        act_out.append({**a, "severity": sev, "type": type_of.get(c, ""),
                        "verifyStartCycle": vs,
                        "noFailCycles": no_fail, "verifyTarget": verify_cy,
                        "verifyProgress": 100 if manual_done else (round(min(1, no_fail / verify_cy) * 100) if verify_cy else 0),
                        "verifyResult": result, "verifyManual": manual_done})
    verify_closed_rate = round(verified / relevant * 100) if relevant else 100

    # ── 재발 집계 (검증완료 후 리셋) ──────────────────────────────
    # 같은 코드가 2회 이상 = 재발, 코드별 (횟수-1)의 합. 단, 조치의 verifyResult 가
    # '검증완료'(조치 후 verify_cy Cycle 무재발)인 코드는 근본원인 해결·검증종결로 보고
    # 재발 카운트에서 제외(리셋)한다. 검증 이후 다시 터진 '재발' 코드는 검증완료가 아니므로 유지된다.
    resolved_codes = {a["code"] for a in act_out if a.get("verifyResult") == "검증완료"}
    recur_count = sum(n - 1 for c, n in cnt.items() if n > 1 and c not in resolved_codes)
    recur_items = [{"code": c, "type": type_of.get(c, ""), "count": n}
                   for c, n in cnt.items() if n > 1 and c not in resolved_codes]
    recur_cleared = [{"code": c, "type": type_of.get(c, ""), "count": n}
                     for c, n in cnt.items() if n > 1 and c in resolved_codes]

    def st(ok, prog=False):
        return "pass" if ok else ("prog" if prog else "fail")
    # id = 안정적 식별자(순서·라벨을 config.ui.acceptance.criteria 에서 바꿔도 값/판정 매핑 유지).
    criteria = [
        {"id": "complete", "key": "완주", "value": f"{attempt_cycles}/{target}",
         "status": st(attempt_cycles >= target, prog=attempt_cycles < target)},
        {"id": "mtbf", "key": f"MTBF≥{mtbf_t} @{int(conf_lv * 100)}%", "value": f"MTBF {mtbf_cur} / 신뢰수준 {round(conf_cur * 100)}%",
         "status": st(conf_cur >= conf_lv)},
        {"id": "openCritical", "key": "미해결 Critical=0", "value": f"{open_critical}건",
         "status": st(open_critical <= acc.get("criticalOpenLimit", 0))},
        {"id": "recur", "key": f"재발≤{recur_lim}", "value": f"{recur_count}건",
         "status": st(recur_count <= recur_lim)},
        {"id": "verifyClose", "key": "전 결함 검증종결", "value": f"{verified}/{relevant}",
         "status": st(relevant and verified == relevant, prog=verified < relevant)},
    ]
    passed = sum(1 for c in criteria if c["status"] == "pass")

    if recur_count >= 1 or open_critical >= 1:
        op_grade = "주의"
    elif recur_count == 0 and open_critical == 0 and verify_closed_rate >= 80:
        op_grade = "양호"
    else:
        op_grade = "보통"

    return {
        "metrics": {
            "progress": {"cum": attempt_cycles, "target": target,
                         "pct": round(attempt_cycles / target * 100, 1) if target else 0},
            "errorBudget": {"used": budget_used, "limit": error_limit,
                            "resets": budget_resets, "lifetimeErrors": cum_err,
                            "dailyAvg": round(cum_err / len(daily), 2) if daily else 0,
                            "weeklyAvg": round(cum_err / len(weekly_list), 2) if weekly_list else 0},
            "successRate": round(success / cum_total * 100, 1) if cum_total else 0,
            "success": success, "errorsTotal": cum_err,
            "mtbf": {"current": mtbf_cur, "target": mtbf_t},
            "errRateCur": err_rate_cur, "errRateTarget": err_rate_t,
            "streak": {"current": streak_cur, "max": streak_max},
            "failure": {"count": failure_count, "lastCycle": last_fail_cycle,
                        "freeCycles": failure_free, "keyword": fail_kw},
            "weekly": weekly_list,
            "errRate": errrate_list,
            "recentWindow": recent_window,
            "throughput": {"total": cum_total,
                           "daily": round(cum_total / len(daily), 1) if daily else 0},
            "confidence": {"level": conf_lv, "current": round(conf_cur, 3),
                           "currentPct": round(conf_cur * 100), "currentCycles": failure_free,
                           "requiredForLevel": required(conf_lv), "table": conf_table},
        },
        "failure": {"top5ByCode": top5_by_code, "paretoByType": pareto,
                    "severityDist": sev_dist, "matrix": matrix},
        "actions": act_out,
        "recurrence": {"count": recur_count,
                       "rate": round(recur_count / len(errors) * 100, 1) if errors else 0,
                       "items": recur_items,
                       "cleared": recur_cleared},
        "acceptance": {"criteria": criteria, "passed": passed, "total": len(criteria)},
        "opReliability": {"grade": op_grade, "recur": recur_count,
                          "openCritical": open_critical, "verifyClosedRate": verify_closed_rate},
    }


# ── 단계별 계산: POC (통계 금지 — 4분류·Pareto·무고장 런·비정상평가) ──
FOURWAY_MAP = [
    ("concept", "① 컨셉 리스크", ["컨셉"]),
    ("design",  "② 설계 개선",   ["설계"]),
    ("impl",    "③ 구현(SW) 버그", ["구현", "sw", "버그"]),
    ("env",     "④ 시험환경 요인", ["시험환경", "환경", "시험"]),
]
CLOSED_STATUSES = ("종결", "완료", "검증완료")


# ── 공통 레코드 스토어 (docs/RECORD_SCHEMA.md §1 — 3원칙 ①의 구현) ──────
# 전 과제·전 단계 공통 형식의 records[] 를 dashboard.json 에 병기한다.
#   · 배관(수집→분류→폐루프→재발 링크)은 전 단계 하나, 화면(렌즈)만 단계별.
#   · 단계가 올라도 스키마는 같고 '필수 필드'만 늘어난다 — 검증은 경고(빌드 차단 없음).
#   · 이관 리허설: --validate-stage <상위 단계> 로 현 대장을 다음 단계 잣대로 사전 점검.
RECORD_REQUIRED = {
    "poc":   ["id", "mode", "severity", "cause", "status"],
    "pilot": ["id", "mode", "severity", "cause", "status", "date", "cycle", "detail", "swVer"],
    "mass":  ["id", "mode", "severity", "cause", "status", "date", "cycle", "detail", "swVer",
              "images", "verdict"],
    "spread": ["id", "mode", "severity", "cause", "status", "date", "detail", "unit"],
    "ops":   ["id", "mode", "severity", "cause", "status", "date", "detail", "unit"],
}
RECORD_FIELD_LABEL = {
    "id": "ID", "mode": "고장모드", "severity": "심각도", "cause": "원인분류", "status": "상태",
    "date": "발생일", "cycle": "누적Cy", "detail": "상세", "swVer": "SW버전",
    "images": "증거(사진)", "verdict": "판정(관련/비관련)", "unit": "호기/라인",
}
# 무발생검증(#10)·재발링크(#11)는 파생/조건부 필드라 공란 검증 대상에서 제외한다:
#   재발 링크는 선행 레코드가 있을 때만 존재하고, 무발생검증은 '검증중' 상태에서만 의미가 있다.
#   (검증중인데 무발생검증이 비면 아래 _validate_records 가 별도로 경고)


def _records_from_issues(issues: list[dict]) -> list[dict]:
    """POC 이슈로그 → 공통 레코드. (recurOf 는 _compute_poc 가 주입한 값을 승계)"""
    recs = []
    for i in issues:
        recs.append({
            "id": i.get("id") or "", "mode": i.get("mode") or "", "modeCode": "",
            "severity": i.get("severity") or "",
            "cause": i.get("cause4") or "",          # 단계 축: POC 4분류
            "status": i.get("status") or "",
            "date": i.get("date") or "", "time": "", "cycle": None,
            "detail": i.get("detail") or "", "action": "",
            "swVer": "", "hwVer": "",
            "verify": i.get("verify") or "", "recurLink": i.get("recurOf") or "",
            "closedDate": i.get("closedDate") or "",
            "images": i.get("images") or [], "verdict": "",
            "unit": i.get("unit") or "", "downtime": i.get("downtime") or "",
        })
    return recs


def _records_from_errors(errors: list[dict], codes: list[dict],
                         actions: list[dict], adjud: list[dict] | None = None) -> list[dict]:
    """Pilot·양산 에러로그 → 공통 레코드. errors 순서를 유지한다(프론트 인덱스 정렬용).
    상태는 코드별 조치검증(검증완료→종결/검증중/조치중)에서 유도 — 화면 Top5 '현황'과 동일 근사."""
    sev_of = {c.get("code"): c.get("severity") for c in (codes or [])}
    act_of: dict = {}
    for a in (actions or []):
        act_of.setdefault(a.get("code"), a)
    verd_of: dict = {}
    for r in (adjud or []):
        verd_of[_norm(str(r.get("target") or ""))] = r.get("verdict") or ""
    recs = []
    for e in errors:
        a = act_of.get(e.get("code")) or {}
        vr = str(a.get("verifyResult") or a.get("status") or "")
        if "검증완료" in vr or vr in CLOSED_STATUSES:
            status = "종결"
        elif "재발" in vr:
            status = "재분석"
        elif "검증" in vr:
            status = "검증중"
        elif e.get("action") or a:
            status = "조치중"
        else:
            status = "신규"
        verify = ""
        if a.get("noFailCycles") is not None and a.get("verifyTarget"):
            verify = f"{a['noFailCycles']}/{a['verifyTarget']}Cy"
        no = e.get("no")
        rid = f"E-{no:03d}" if isinstance(no, int) and no else str(no or "")
        verdict = verd_of.get(_norm(str(no or ""))) or verd_of.get(_norm(e.get("code") or ""), "")
        recs.append({
            "id": rid, "mode": e.get("type") or "", "modeCode": e.get("code") or "",
            "severity": sev_of.get(e.get("code")) or "",
            "cause": e.get("cause_class") or "",     # 단계 축: 근본원인 분류 (선택 컬럼)
            "causeText": e.get("cause") or "",
            "status": status,
            "date": e.get("date") or "", "time": e.get("time") or "", "cycle": e.get("cycle"),
            "detail": e.get("detail") or "", "action": e.get("action") or "",
            "swVer": e.get("sw_ver") or "", "hwVer": e.get("hw_ver") or "",
            "verify": verify, "recurLink": "", "closedDate": "",
            "images": e.get("images") or [], "verdict": verdict, "unit": e.get("unit") or "",
        })
    _annotate_recur_links(recs)
    return recs


def _annotate_recur_links(records: list[dict]):
    """동일 고장모드 선행 레코드 ID 주입 (재발 = 동일 모드 재출현, docs/CRITERIA.md §5)."""
    seen: dict[str, str] = {}
    for r in sorted(records, key=lambda x: (x.get("date") or "9999", str(x.get("id") or ""))):
        k = r.get("modeCode") or r.get("mode") or ""
        if k and k in seen and not r.get("recurLink"):
            r["recurLink"] = seen[k]
        if k:
            seen[k] = r.get("id") or ""


def _status_dist_of(records: list[dict]) -> dict:
    dist = {"new": 0, "acting": 0, "verifying": 0, "closed": 0}
    for r in records:
        dist[_status_bucket(r.get("status"))] += 1
    return dist


def _validate_records(records: list[dict], stage: str, tag: str = "") -> list[str]:
    """단계별 필수 필드 충족 검증 — 필드별 미충족 건수를 경고 한 줄로 집계.
    반환: 경고 문자열 리스트 (이관 리허설 리포트에서 재사용)."""
    req = RECORD_REQUIRED.get(stage) or RECORD_REQUIRED["poc"]
    gaps = []
    for f in req:
        missing = [r for r in records
                   if not (len(r.get(f) or []) if f == "images" else str(r.get(f) or "").strip())]
        if missing:
            ids = ", ".join(str(m.get("id") or "?") for m in missing[:3])
            gaps.append(f"{RECORD_FIELD_LABEL.get(f, f)} 없음 {len(missing)}건 ({ids}{' …' if len(missing) > 3 else ''})")
    if stage != "poc":
        nover = [r for r in records if _status_bucket(r.get("status")) == "verifying" and not r.get("verify")]
        if nover:
            gaps.append(f"검증중인데 무발생검증(n/목표) 미기재 {len(nover)}건")
    label = tag or f"{stage} 잣대"
    for g in gaps:
        print(f"[build] ⚠ 레코드 검증({label}): {g}")
    if not gaps and tag:
        print(f"[build] ✓ 레코드 검증({label}): 필수 필드 전 건 충족")
    return gaps


def _pareto(items: list[dict], key: str) -> list[dict]:
    """고장모드별 건수 내림차순 + 누적% (수정개발 우선순위)."""
    counts: dict[str, int] = {}
    for it in items:
        k = it.get(key) or "(미분류)"
        counts[k] = counts.get(k, 0) + 1
    rows = sorted(counts.items(), key=lambda kv: -kv[1])
    total = sum(counts.values()) or 1
    out, cum = [], 0
    for mode, n in rows:
        cum += n
        out.append({"mode": mode, "count": n, "cumPct": round(cum / total * 100)})
    return out


def _run_gauge(runlog: list[dict], target: float) -> dict:
    """무고장 런 게이지: 연속 무에러/무정지 시간. 에러 발생일 = 리셋(그날 시간까지 진행 후 0부터)."""
    runlog = sorted(runlog, key=lambda r: r.get("date") or "")
    cum = 0.0        # 현재 시도 누적 시간
    total = 0.0      # 전체 누적 가동 시간
    resets = []      # [{atHours(전체 누적 기준 위치), date}]
    for r in runlog:
        h = r.get("hours") or 0
        total += h
        cum += h
        if (r.get("errors") or 0) > 0:
            resets.append({"at": round(total, 1), "date": r.get("date") or "", "note": r.get("notes") or ""})
            cum = 0.0
    pct = round(min(100.0, cum / target * 100), 1) if target else 0
    return {"cum": round(cum, 1), "target": target, "pct": pct,
            "attempt": len(resets) + 1, "resets": resets, "totalHours": round(total, 1)}


def _recur_by_mode(items: list[dict], key: str) -> dict:
    counts: dict[str, int] = {}
    for it in items:
        k = it.get(key) or ""
        if k:
            counts[k] = counts.get(k, 0) + 1
    recur = [{"mode": k, "count": n} for k, n in counts.items() if n > 1]
    return {"count": len(recur), "items": recur}


def _status_bucket(s: str) -> str:
    """이슈 상태 → 폐루프 4상태 (신규→조치중→검증중→종결). 표기가 흔들려도 포함어로 흡수."""
    s = (s or "").strip()
    if s in CLOSED_STATUSES:
        return "closed"
    if "검증" in s:
        return "verifying"
    if "조치" in s or "분석" in s or "진행" in s:
        return "acting"
    return "new"


def _poc_trend(issues: list[dict]) -> list[dict]:
    """주차별 누적 발견 vs 누적 종결 (월요일 시작) — '수렴하고 있다'의 그림.
    종결일 없는 종결 건은 마지막 주에 계상하고 경고한다 (추이 총계 = issueStats와 일치 보장)."""
    def monday(dstr: str):
        dt = datetime.strptime(dstr, "%Y-%m-%d").date()
        return dt - timedelta(days=dt.weekday())

    dated = [i for i in issues if i.get("date")]
    if not dated:
        return []
    closed_issues = [i for i in issues if _status_bucket(i.get("status")) == "closed"]
    no_cdate = sum(1 for i in closed_issues if not i.get("closedDate"))
    if no_cdate:
        print(f"[build] ⚠ 종결 이슈 {no_cdate}건에 종결일 없음 — 수렴 추이의 마지막 주에 계상")
    all_dates = [i["date"] for i in dated] + [i["closedDate"] for i in closed_issues if i.get("closedDate")]
    w0, w1 = min(map(monday, all_dates)), max(map(monday, all_dates))
    weeks = []
    w = w0
    while w <= w1:
        weeks.append(w)
        w += timedelta(days=7)
    found_by: dict = {}
    closed_by: dict = {}
    for i in dated:
        k = monday(i["date"])
        found_by[k] = found_by.get(k, 0) + 1
    for i in closed_issues:
        k = monday(i["closedDate"]) if i.get("closedDate") else weeks[-1]
        closed_by[k] = closed_by.get(k, 0) + 1
    out, cf, cc = [], 0, 0
    for n, w in enumerate(weeks):
        cf += found_by.get(w, 0)
        cc += closed_by.get(w, 0)
        out.append({"week": n + 1, "weekStart": w.isoformat(), "found": cf, "closed": cc})
    return out


def _compute_poc(issues: list[dict], runlog: list[dict], abnormal: list[dict], config: dict) -> dict:
    run_cfg = config.get("run") or {}
    target = float(run_cfg.get("target") or 72)
    run = _run_gauge(runlog, target)
    fourway = []
    for key, label, kws in FOURWAY_MAP:
        subset = [i for i in issues
                  if any(kw in _norm(i.get("cause4") or "") for kw in [_norm(k) for k in kws])]
        closed = sum(1 for i in subset if (i.get("status") or "") in CLOSED_STATUSES)
        fourway.append({"key": key, "label": label, "count": len(subset), "closed": closed})
    mapped = sum(f["count"] for f in fourway)
    if mapped < len(issues):   # 분류 누락분은 시험환경이 아니라 별도 표기 대신 구현으로 오해 없게 카운트만 경고
        print(f"[build] ⚠ 이슈 {len(issues) - mapped}건의 원인분류가 4분류에 매칭되지 않음")
    closed_total = sum(1 for i in issues if (i.get("status") or "") in CLOSED_STATUSES)

    # 재발 링크: 동일 고장모드 선행 레코드 ID 주입 (재발 정의 = 동일 모드 재출현, docs/CRITERIA.md §5)
    seen_mode: dict[str, str] = {}
    for i in sorted(issues, key=lambda x: (x.get("date") or "9999", x.get("id") or "")):
        m = i.get("mode") or ""
        if m and m in seen_mode:
            i["recurOf"] = seen_mode[m]
        if m:
            seen_mode[m] = i.get("id") or ""

    # 폐루프 상태 분포 (신규/조치중/검증중/종결)
    status_dist = {"new": 0, "acting": 0, "verifying": 0, "closed": 0}
    for i in issues:
        status_dist[_status_bucket(i.get("status"))] += 1

    return {
        "metrics": {"progress": {"cum": run["cum"], "target": run["target"], "pct": run["pct"]}},
        "run": run,
        "fourway": fourway,
        "pareto": _pareto(issues, "mode"),
        "recurrence": _recur_by_mode(issues, "mode"),
        "issueStats": {"total": len(issues), "closed": closed_total, "open": len(issues) - closed_total},
        "statusDist": status_dist,
        "trend": _poc_trend(issues),
        "abnormal": abnormal,
    }


# ── 단계별 계산: Pilot (추세 — MCBF 성장·Pareto·시정조치 규율·형상) ──
def _compute_pilot(daily: list[dict], errors: list[dict], config: dict,
                   codes: list[dict], actions: list[dict]) -> dict:
    run_cfg = config.get("run") or {}
    target = float(run_cfg.get("target") or 300)
    growth_target = run_cfg.get("growthTarget") or (config.get("acceptance") or {}).get("mtbfTargetCycle")

    # 무정지 런(시간): 일일평가 hours 합산, 에러 발생일 리셋
    runlog = [{"date": d["date"], "hours": d.get("hours") or 0,
               "errors": d.get("errors") or 0, "notes": d.get("notes") or ""} for d in daily]
    run = _run_gauge(runlog, target)

    # 주차별 MCBF 성장 (누적 cycles / 누적 errors) — 월요일 시작 주 단위
    growth = []
    if daily:
        cum_cy, cum_err = 0, 0
        week_of = {}
        for d in daily:
            dt = datetime.strptime(d["date"], "%Y-%m-%d").date()
            monday = dt - timedelta(days=dt.weekday())
            week_of.setdefault(monday, []).append(d)
        for i, monday in enumerate(sorted(week_of)):
            for d in week_of[monday]:
                cum_cy += d.get("total") or 0
                cum_err += d.get("errors") or 0
            growth.append({"week": i + 1, "weekStart": monday.isoformat(),
                           "mcbf": round(cum_cy / max(cum_err, 1))})

    # 형상(버전) 이력: 에러로그의 sw_ver 등장 순서 (버전 배포 시점 근사)
    versions, seen = [], set()
    for e in errors:
        v = e.get("sw_ver") or ""
        if v and v not in seen:
            seen.add(v)
            versions.append({"ver": v, "date": e.get("date") or ""})

    # 시정조치 규율: 검증완료율 (모든 수정 → 검증 런)
    closed = sum(1 for a in actions if (a.get("status") or "") in CLOSED_STATUSES
                 or (a.get("verifyResult") or "") == "검증완료")
    act_rate = {"total": len(actions), "closed": closed,
                "pct": round(closed / len(actions) * 100) if actions else 0}

    return {
        "metrics": {"progress": {"cum": run["cum"], "target": run["target"], "pct": run["pct"]}},
        "run": run,
        "growth": growth,
        "growthTarget": growth_target,
        "pareto": _pareto(errors, "type"),
        "recurrence": _recur_by_mode(errors, "type"),
        "versions": versions,
        "actionRate": act_rate,
    }


# ── 단계별 계산: 확산 (원인계층·호기별 층화 — 설계성 고장은 전 함대 리스크) ──
CAUSE_LAYER_MAP = [
    ("design",  "① 설계",       ["설계"]),
    ("build",   "② 제작·조립",  ["제작", "조립"]),
    ("install", "③ 설치·시공",  ["설치", "시공"]),
    ("oper",    "④ 운영·환경",  ["운영", "환경", "조작"]),
]


def _num(v, default=0.0) -> float:
    try:
        return float(str(v).replace(",", "").replace("%", ""))
    except (TypeError, ValueError):
        return default


def _annotate_issue_recur(issues: list[dict]):
    """동일 고장모드 선행 레코드 ID(recurOf) 주입 — POC와 동일 규칙 (CRITERIA §5)."""
    seen: dict[str, str] = {}
    for i in sorted(issues, key=lambda x: (x.get("date") or "9999", x.get("id") or "")):
        m = i.get("mode") or ""
        if m and m in seen:
            i["recurOf"] = seen[m]
        if m:
            seen[m] = i.get("id") or ""


def _cause_layer(issues: list[dict]) -> list[dict]:
    out = []
    for key, label, kws in CAUSE_LAYER_MAP:
        subset = [i for i in issues
                  if any(kw in _norm(i.get("cause4") or "") for kw in [_norm(k) for k in kws])]
        closed = sum(1 for i in subset if (i.get("status") or "") in CLOSED_STATUSES)
        out.append({"key": key, "label": label, "count": len(subset), "closed": closed})
    return out


def _compute_spread(issues: list[dict], units: list[dict], config: dict) -> dict:
    _annotate_issue_recur(issues)
    for u in units:
        u["runH"] = _num(u.get("runH"))
        u["runTarget"] = _num(u.get("runTarget")) or _num((config.get("run") or {}).get("target")) or 48
    sat_done = sum(1 for u in units if "PASS" in str(u.get("sat") or "").upper())
    run_done = sum(1 for u in units if u["runH"] >= u["runTarget"] and "PASS" in str(u.get("sat") or "").upper())
    total = len(units)
    fleet = {"total": total, "satDone": sat_done, "qualified": run_done,
             "pct": round(run_done / total * 100, 1) if total else 0}
    # 호기별 이슈 층화 (어느 호기에 몰리는가 — "이 호기만의 병" 판별)
    unit_dist: dict[str, int] = {}
    for i in issues:
        u = i.get("unit") or "(미기재)"
        unit_dist[u] = unit_dist.get(u, 0) + 1
    layers = _cause_layer(issues)
    design_open = [i for i in issues
                   if any(kw in _norm(i.get("cause4") or "") for kw in ["설계"])]
    records = _records_from_issues(issues)
    return {
        "metrics": {"progress": {"cum": run_done, "target": total, "pct": fleet["pct"]}},
        "fleet": fleet, "units": units,
        "causeLayer": layers,
        "escalations": [{"id": i.get("id"), "mode": i.get("mode"), "status": i.get("status"),
                         "detail": i.get("detail")} for i in design_open],
        "unitDist": [{"unit": k, "count": v} for k, v in sorted(unit_dist.items())],
        "pareto": _pareto(issues, "mode"),
        "recurrence": _recur_by_mode(issues, "mode"),
        "issueStats": {"total": len(issues),
                       "closed": sum(1 for i in issues if (i.get("status") or "") in CLOSED_STATUSES),
                       "open": sum(1 for i in issues if (i.get("status") or "") not in CLOSED_STATUSES)},
        "records": records,
        "statusDist": _status_dist_of(records),
    }


# ── 단계별 계산: 운영/관제 (월간 RAM · 알람→FRACAS 승격 · 다운타임 Pareto · CIP) ──
def _compute_ops(issues: list[dict], monthly: list[dict], cip: list[dict], config: dict) -> dict:
    _annotate_issue_recur(issues)
    months = []
    for m in monthly:
        months.append({"month": str(m.get("month") or ""), "avail": _num(m.get("avail")),
                       "mtbf": _num(m.get("mtbf")), "mttr": _num(m.get("mttr")),
                       "downtime": _num(m.get("downtime")),
                       "alarms": int(_num(m.get("alarms"))), "promoted": int(_num(m.get("promoted")))})
    cur = months[-1] if months else {}
    acc = config.get("acceptance") or {}
    avail_target = _num(acc.get("availTargetPct"), 98.0)
    alarms_total = sum(m["alarms"] for m in months)
    promoted_total = sum(m["promoted"] for m in months)
    # 다운타임 Pareto — "어떤 고장부터 없애는 게 경제적인가" (건수가 아니라 손실 시간 순)
    down_by: dict[str, float] = {}
    for i in issues:
        down_by[i.get("mode") or "(미분류)"] = down_by.get(i.get("mode") or "(미분류)", 0) + _num(i.get("downtime"))
    down_total = sum(down_by.values()) or 1
    down_pareto, cumv = [], 0.0
    for mode, mins in sorted(down_by.items(), key=lambda kv: -kv[1]):
        cumv += mins
        down_pareto.append({"mode": mode, "minutes": round(mins), "cumPct": round(cumv / down_total * 100)})
    records = _records_from_issues(issues)
    return {
        "metrics": {"progress": {"cum": cur.get("avail", 0), "target": avail_target,
                                 "pct": round(min(100.0, cur.get("avail", 0) / avail_target * 100), 1) if avail_target else 0}},
        "ram": {"months": months, "current": cur, "availTarget": avail_target},
        "alarms": {"total": alarms_total, "promoted": promoted_total,
                   "rate": round(promoted_total / alarms_total * 100, 1) if alarms_total else 0},
        "downPareto": down_pareto,
        "cip": cip,
        "causeLayer": _cause_layer(issues),
        "pareto": _pareto(issues, "mode"),
        "recurrence": _recur_by_mode(issues, "mode"),
        "issueStats": {"total": len(issues),
                       "closed": sum(1 for i in issues if (i.get("status") or "") in CLOSED_STATUSES),
                       "open": sum(1 for i in issues if (i.get("status") or "") not in CLOSED_STATUSES)},
        "records": records,
        "statusDist": _status_dist_of(records),
    }


def build_project(pid: str):
    _set_project(pid)
    config = _load_config()
    stage = config.get("stage", "mass")
    if stage == "poc":
        return _build_poc(pid, config)
    if stage == "pilot":
        return _build_pilot(pid, config)
    if stage == "spread":
        return _build_spread(pid, config)
    if stage == "ops":
        return _build_ops(pid, config)
    return _build_mass(pid, config)


def _now_iso():
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _write_out(pid: str, out: dict, tail: str):
    OUT_PATH.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[build:{pid}] 출력: {OUT_PATH.relative_to(ROOT)}  ({tail})")


def _codes_from_sheets(sheets: dict) -> list[dict]:
    """raw 엑셀 안의 코드마스터 시트 → codes[] (개발 단계: 관리엑셀 없이 이슈로그 파일에 병기)."""
    rows = _pick_sheet(sheets, ["코드마스터", "코드", "code"])
    if not rows:
        return []
    raw = _read_mgmt_rows(rows, CODE_FIELDS)
    return [{"code": _cell_to_str(c["code"]), "type": _cell_to_str(c["type"]),
             "severity": _norm_sev(c["severity"]), "desc": _cell_to_str(c["desc"])}
            for c in raw if _cell_to_str(c["code"])]


def _backfill_mode_codes(codes: list[dict], *record_lists) -> int:
    """modeCode 공란 레코드에 코드마스터 type(=모드명) 역매핑으로 표준 코드 주입."""
    t2c = {_norm(c["type"]): c["code"] for c in codes if c.get("type")}
    n = 0
    for lst in record_lists:
        for r in lst or []:
            if not r.get("modeCode"):
                code = t2c.get(_norm(r.get("mode") or ""))
                if code:
                    r["modeCode"] = code
                    n += 1
    return n


def _build_poc(pid: str, config: dict):
    src = _pick_latest_xlsx()
    print(f"[build:{pid}] 입력 파일: {src.name} (stage=poc)")
    sheets = _load_all_sheets(src)
    issues   = _parse_generic(_pick_sheet(sheets, ISSUES_SHEET_KEYWORDS), ISSUE_FIELD_ALIASES, "id")
    runlog   = _parse_generic(_pick_sheet(sheets, RUNLOG_SHEET_KEYWORDS), RUN_FIELD_ALIASES, "date")
    abnormal = _parse_generic(_pick_sheet(sheets, ABN_SHEET_KEYWORDS), ABN_FIELD_ALIASES, "scenario")
    codes    = _codes_from_sheets(sheets)
    computed = _compute_poc(issues, runlog, abnormal, config)
    records = _records_from_issues(issues)   # _compute_poc 이후 (recurOf 주입 승계)
    if codes:
        n = _backfill_mode_codes(codes, issues, records)
        if n:
            print(f"[build:{pid}] 코드마스터 역매핑: modeCode {n}건 주입")
    _validate_records(records, "poc")
    out = {"generatedAt": _now_iso(), "source": src.name, "config": config,
           "issues": issues, "runlog": runlog, "records": records, "codes": codes, **computed}
    _write_out(pid, out, f"이슈 {len(issues)}, 런기록 {len(runlog)}일, "
               f"무고장 런 {computed['run']['cum']}/{computed['run']['target']}h, "
               f"비정상 {len(abnormal)}건")


def _build_spread(pid: str, config: dict):
    src = _pick_latest_xlsx()
    print(f"[build:{pid}] 입력 파일: {src.name} (stage=spread)")
    sheets = _load_all_sheets(src)
    issues = _parse_generic(_pick_sheet(sheets, ISSUES_SHEET_KEYWORDS), ISSUE_FIELD_ALIASES, "id")
    units  = _parse_generic(_pick_sheet(sheets, UNITQ_SHEET_KEYWORDS), UNIT_FIELD_ALIASES, "unit")
    computed = _compute_spread(issues, units, config)
    _validate_records(computed["records"], "spread")
    out = {"generatedAt": _now_iso(), "source": src.name, "config": config,
           "issues": issues, **computed}
    f = computed["fleet"]
    _write_out(pid, out, f"호기 {f['total']} (퀄 완료 {f['qualified']}), 이슈 {len(issues)}, "
               f"설계성 {len(computed['escalations'])}건")


def _build_ops(pid: str, config: dict):
    src = _pick_latest_xlsx()
    print(f"[build:{pid}] 입력 파일: {src.name} (stage=ops)")
    sheets = _load_all_sheets(src)
    issues  = _parse_generic(_pick_sheet(sheets, ISSUES_SHEET_KEYWORDS), ISSUE_FIELD_ALIASES, "id")
    monthly = _parse_generic(_pick_sheet(sheets, MONTHLY_SHEET_KEYWORDS), MONTHLY_FIELD_ALIASES, "month")
    cip     = _parse_generic(_pick_sheet(sheets, CIP_SHEET_KEYWORDS), CIP_FIELD_ALIASES, "id")
    computed = _compute_ops(issues, monthly, cip, config)
    _validate_records(computed["records"], "ops")
    out = {"generatedAt": _now_iso(), "source": src.name, "config": config,
           "issues": issues, **computed}
    ram = computed["ram"]["current"]
    _write_out(pid, out, f"월간지표 {len(monthly)}개월, 필드 FRACAS {len(issues)}건, "
               f"가동률 {ram.get('avail', '—')}%, CIP {len(cip)}건")


def _build_pilot(pid: str, config: dict):
    src = _pick_latest_xlsx()
    print(f"[build:{pid}] 입력 파일: {src.name} (stage=pilot)")
    daily_rows, errors_rows = _load_workbook_rows(src)
    daily  = _parse_daily(daily_rows)
    errors = _parse_errors(errors_rows)
    codes, actions, _adjud, _dispo = _load_mgmt()   # 판정·처분대장은 양산 시범 평가~인증부터 사용
    computed = _compute_pilot(daily, errors, config, codes, actions)
    records = _records_from_errors(errors, codes, actions)
    _validate_records(records, "pilot")
    out = {"generatedAt": _now_iso(), "source": src.name, "config": config,
           "codes": codes, "daily": daily, "errors": errors, "actions": actions,
           "records": records, "statusDist": _status_dist_of(records), **computed}
    g = computed["growth"]
    _write_out(pid, out, f"daily {len(daily)}, errors {len(errors)}, "
               f"MCBF {g[-1]['mcbf'] if g else '—'}/{computed.get('growthTarget') or '—'}, "
               f"무정지 런 {computed['run']['cum']}/{computed['run']['target']}h")


def _build_mass(pid: str, config: dict):
    src = _pick_latest_xlsx()
    print(f"[build:{pid}] 입력 파일: {src.name}")

    daily_rows, errors_rows = _load_workbook_rows(src)
    daily  = _parse_daily(daily_rows)
    errors = _parse_errors(errors_rows)

    config = _load_config()
    codes, actions, adjud, dispositions = _load_mgmt()
    now = datetime.now(timezone.utc)
    computed = _compute(daily, errors, config, codes, actions, now=now)

    # ── 월별 누적 스냅샷: 처음 ~ 각 달 말까지의 daily/errors로 재계산 ──
    #    프론트에서 월 선택 시 해당 스냅샷으로 교체 렌더한다.
    def _month(x):
        return (x.get("date") or "")[:7]
    months = sorted({_month(x) for x in daily if _month(x)})
    snapshots = {}
    for mo in months:
        dM = [x for x in daily  if _month(x) and _month(x) <= mo]
        eM = [x for x in errors if (not _month(x)) or _month(x) <= mo]
        cM = _compute(dM, eM, config, codes, actions, now=now)
        snapshots[mo] = {**cM, "daily": dM, "errors": eM}

    # 공통 레코드 스토어 — 조치검증 판정이 붙은 computed["actions"] 를 사용 (무발생 진행 포함).
    #   _compute(양산, SEC 원본)는 무변경 — records 는 그 산출물을 읽어 병기만 한다.
    records = _records_from_errors(errors, codes, computed.get("actions") or actions, adjud)
    _validate_records(records, "mass")

    out = {
        "generatedAt": now.strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source":      src.name,
        "config":      config,
        "codes":       codes,
        "daily":       daily,
        "errors":      errors,
        **computed,
        "records":     records,
        "statusDist":  _status_dist_of(records),
        "adjudication": adjud,   # 판정대장 (관련/비관련 합동판정 — docs/CRITERIA.md §4)
        "dispositions": dispositions,   # 처분대장 (인증 준비: 오픈 건 처분 → Known Issues Register)
        "months":      months,
        "snapshots":   snapshots,
    }
    OUT_PATH.write_text(
        json.dumps(out, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    m = computed["metrics"]
    print(f"[build:{pid}] 출력: {OUT_PATH.relative_to(ROOT)}  "
          f"(daily {len(daily)}, errors {len(errors)}, 진행 {m['progress']['pct']}%, "
          f"신뢰수준 {m['confidence']['currentPct']}%, 합격 {computed['acceptance']['passed']}/5)")


def _cause_key(c) -> str:
    """원인분류 → 정준 키 (프론트 c4Key와 동일 규칙 — 홈 분류 구성 바용)."""
    s = str(c or "")
    sl = s.lower()
    if "컨셉" in s:
        return "concept"
    if "설계" in s:
        return "design"
    if "부품" in s:
        return "parts"
    if "제작" in s or "조립" in s:
        return "build"
    if "설치" in s or "시공" in s:
        return "install"
    if "구현" in s or "sw" in sl or "버그" in s:
        return "sw"
    if "운영" in s or "조작" in s:   # '운영·환경'은 oper (확산 layerKey와 동일) — 환경보다 먼저
        return "oper"
    if "환경" in s or "시험" in s or "자재" in s:
        return "env"
    return "etc"


def _portfolio_summary(stage: str, out: dict) -> dict:
    """홈(포트폴리오) 카드용 요약 — 단계별로 헤드라인 수치만 추린다."""
    m = out.get("metrics", {}) or {}
    recs = out.get("errors")
    if recs is None:
        recs = out.get("issues") or []
    s = {"progress": m.get("progress"), "records": len(recs)}
    if stage == "poc":
        # POC의 '동일 모드 다건'은 발굴 단계의 자연스러운 현상 — 재발 KPI에서 제외.
        s["issueStats"] = out.get("issueStats")
        fw = out.get("fourway") or []
        s["concept"] = next((f.get("count") for f in fw if f.get("key") == "concept"), None)
    else:
        s["recur"] = (out.get("recurrence") or {}).get("count")
    if stage == "spread":
        s["fleet"] = out.get("fleet")
        s["issueStats"] = out.get("issueStats")
    if stage == "ops":
        ram = out.get("ram") or {}
        s["ram"] = {"avail": (ram.get("current") or {}).get("avail"), "target": ram.get("availTarget")}
        s["issueStats"] = out.get("issueStats")
    # 폐루프 상태 분포 — 홈 카드의 '에러 진행'(종결률 스택바)용, 전 단계 공통 배관
    s["statusDist"] = out.get("statusDist")
    # 심각도 분포 + 오픈 Critical — 홈의 '심각도 깔때기'(크리티컬 조기 소진 증거)용
    sev_d = {"Critical": 0, "Major": 0, "Minor": 0}
    open_crit = 0
    for r in (out.get("records") or []):
        sev = r.get("severity")
        if sev in sev_d:
            sev_d[sev] += 1
        if sev == "Critical" and _status_bucket(r.get("status")) != "closed":
            open_crit += 1
    s["sevDist"] = sev_d
    s["openCritical"] = open_crit
    # 원인분류 분포 — 홈 깔때기의 '분류 구성 바'(단계마다 병의 종류가 어떻게 바뀌는가)용
    cause_d: dict = {}
    for r in (out.get("records") or []):
        k = _cause_key(r.get("cause"))
        cause_d[k] = cause_d.get(k, 0) + 1
    s["causeDist"] = cause_d
    if stage == "mass":
        s["errorBudget"] = m.get("errorBudget")
        s["mtbf"] = m.get("mtbf")
        s["acceptance"] = {k: (out.get("acceptance") or {}).get(k) for k in ("passed", "total")}
    return s


def write_portfolio():
    """레지스트리의 전 과제를 훑어 data/portfolio.json 생성 (홈 화면 입력).
    dashboard.json 이 없는 과제는 hasData=false 로 등재한다."""
    reg = _load_registry()
    entries = []
    for p in sorted(reg.get("projects", []), key=lambda x: x.get("order", 99)):
        pid = p["id"]
        base = PROJECTS_ROOT / pid
        entry = {"id": pid, "name": p.get("name", pid), "abbr": p.get("abbr", pid[:1].upper()),
                 "hasData": False}
        cfg_path = base / "config.json"
        dash_path = base / "dashboard.json"
        if cfg_path.exists():
            cfg = json.loads(cfg_path.read_text(encoding="utf-8"))
            entry["stage"] = cfg.get("stage", "mass")
            entry["run"] = cfg.get("run")
            entry["gate"] = cfg.get("gate")
            entry["tecop"] = cfg.get("tecop")
            prj = cfg.get("project") or {}
            entry["project"] = {k: prj.get(k) for k in ("name", "department", "team", "startDate", "endDate")}
            # 개발(제작) 단계 계획 — 평가 데이터가 없는 과제의 홈 카드가 이걸로 진척을 표시
            if cfg.get("devPlan"):
                entry["devPlan"] = cfg["devPlan"]
            # 홈 카드의 '개발 진행'용: 세부 단계 위치 + SW 완성도 평균
            lc = cfg.get("lifecycle") or []
            if lc:
                cur_i = next((i for i, st in enumerate(lc) if st.get("status") == "current"), None)
                entry["lifecycle"] = {
                    "total": len(lc),
                    "pos": (cur_i + 1) if cur_i is not None else sum(1 for st in lc if st.get("status") == "done"),
                    "current": next((st.get("stage") for st in lc if st.get("status") == "current"), ""),
                }
            sw = cfg.get("swModules") or []
            if sw:
                entry["swAvg"] = round(sum(_num(m.get("pct")) for m in sw) / len(sw))
        if dash_path.exists():
            out = json.loads(dash_path.read_text(encoding="utf-8"))
            entry["hasData"] = True
            entry["generatedAt"] = out.get("generatedAt")
            entry["summary"] = _portfolio_summary(entry.get("stage", "mass"), out)
        entries.append(entry)
    PORTFOLIO_PATH.write_text(
        json.dumps({"projects": entries}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    n_data = sum(1 for e in entries if e["hasData"])
    print(f"[build] 포트폴리오: {PORTFOLIO_PATH.relative_to(ROOT)} (과제 {len(entries)}, 데이터 보유 {n_data})")


_SEV_RANK = {"Critical": 3, "Major": 2, "Minor": 1}
_CAT_ORDER = ["concept", "design", "parts", "build", "install", "sw", "env", "oper", "etc"]
_CAT_LABEL = {"concept": "컨셉", "design": "설계", "parts": "부품", "build": "제작·조립",
              "install": "설치·시공", "sw": "구현(SW)", "env": "환경·자재", "oper": "운영·조작", "etc": "기타"}


def write_library(exclude: set | None = None, out_path: Path | None = None):
    """부서 고장모드 라이브러리 — 전 과제 records를 고장모드 단위로 통합 (data/library.json).
    그룹 키 = modeCode, 없으면 코드마스터 type 역매핑, 그래도 없으면 모드명(어휘 미등재)."""
    reg = _load_registry()
    projects = [p for p in sorted(reg.get("projects", []), key=lambda x: x.get("order", 99))
                if not (exclude and p["id"] in exclude)]
    # 1패스: 과제 dashboard.json 로드 + 코드마스터 통합 (code 키 dedupe, type→code 역매핑)
    codes_master: dict = {}
    type2code: dict = {}
    loaded = []
    for p in projects:
        path = PROJECTS_ROOT / p["id"] / "dashboard.json"
        if not path.exists():
            continue
        out = json.loads(path.read_text(encoding="utf-8"))
        loaded.append((p, out, (out.get("config") or {}).get("stage", "mass")))
        for c in out.get("codes") or []:
            code = c.get("code")
            if not code:
                continue
            prev = codes_master.get(code)
            if prev and (prev.get("type"), prev.get("severity")) != (c.get("type"), c.get("severity")):
                print(f"[build] ⚠ 코드마스터 충돌: {code} — 선착순 유지 ({p['id']})")
            codes_master.setdefault(code, c)
            if c.get("type"):
                type2code.setdefault(_norm(c.get("type")), code)
    # 2패스: records → 고장모드 그룹
    groups: dict = {}
    for p, out, stage in loaded:
        for r in out.get("records") or []:
            code = r.get("modeCode") or type2code.get(_norm(r.get("mode") or ""), "")
            key = code or _norm(r.get("mode") or "") or "(미분류)"
            master = codes_master.get(code) or {}
            g = groups.setdefault(key, {"key": key, "code": code,
                                        "mode": master.get("type") or r.get("mode") or "",
                                        "desc": master.get("desc") or "", "occurrences": []})
            g["occurrences"].append({
                "project": p["id"], "projectName": p.get("name", p["id"]),
                "abbr": p.get("abbr", ""), "stage": stage,
                "id": r.get("id") or "", "date": r.get("date") or "",
                "severity": r.get("severity") or "", "status": r.get("status") or "",
                "cause": r.get("cause") or "", "causeText": r.get("causeText") or "",
                "detail": r.get("detail") or "", "action": r.get("action") or "",
                "verify": r.get("verify") or "", "recurLink": r.get("recurLink") or "",
                "verdict": r.get("verdict") or "", "closedDate": r.get("closedDate") or "",
                "unit": r.get("unit") or "", "images": r.get("images") or []})
    # 그룹 마감 — 발생 오름차순 정렬, 최고 심각도, 최빈 원인분류, 상태 분포, 발생 과제
    for g in groups.values():
        occ = sorted(g["occurrences"], key=lambda o: (o["date"] or "9999", o["id"]))
        g["occurrences"] = occ
        g["severity"] = max((o["severity"] for o in occ), key=lambda s: _SEV_RANK.get(s, 0), default="")
        freq: dict = {}
        for o in occ:
            if o["cause"]:
                freq[o["cause"]] = freq.get(o["cause"], 0) + 1
        label = max(freq, key=lambda c: freq[c]) if freq else ""
        g["categoryLabel"] = label
        g["category"] = _cause_key(label) if label else "etc"
        g["counts"] = {"total": len(occ), **_status_dist_of(occ)}
        seen, pjs = set(), []
        for o in occ:
            if o["project"] in seen:
                continue
            seen.add(o["project"])
            pjs.append({"id": o["project"], "name": o["projectName"], "abbr": o["abbr"], "stage": o["stage"]})
        g["projects"] = pjs
        g["firstDate"], g["lastDate"] = occ[0]["date"], occ[-1]["date"]
        g["recurCount"] = max(0, len(occ) - 1)
    modes = sorted(groups.values(),
                   key=lambda g: (-_SEV_RANK.get(g["severity"], 0), -g["counts"]["total"], g["key"]))
    # 카테고리 롤업 — 표준 8종은 상시(0건 포함), etc는 있을 때만
    cats = []
    for ck in _CAT_ORDER:
        ms = [m for m in modes if m["category"] == ck]
        total = sum(m["counts"]["total"] for m in ms)
        opened = sum(m["counts"]["new"] + m["counts"]["acting"] + m["counts"]["verifying"] for m in ms)
        if ck == "etc" and not ms:
            continue
        cats.append({"key": ck, "label": _CAT_LABEL[ck], "modes": len(ms), "records": total, "open": opened})
    total_rec = sum(m["counts"]["total"] for m in modes)
    total_open = sum(m["counts"]["new"] + m["counts"]["acting"] + m["counts"]["verifying"] for m in modes)
    lib = {"generatedAt": _now_iso(),
           "totals": {"projects": len(loaded), "modes": len(modes), "records": total_rec, "open": total_open},
           "codes": [codes_master[k] for k in sorted(codes_master)],
           "categories": cats, "modes": modes}
    path = out_path or LIBRARY_PATH
    path.write_text(json.dumps(lib, ensure_ascii=False, indent=2), encoding="utf-8")
    rel = path.relative_to(ROOT) if path.is_relative_to(ROOT) else path
    print(f"[build] 라이브러리: {rel} (모드 {len(modes)}, 레코드 {total_rec}, 과제 {len(loaded)})")


def main():
    import argparse
    ap = argparse.ArgumentParser(description="과제별 엑셀 → dashboard.json (+ 전사 portfolio.json)")
    ap.add_argument("--project", help="과제 id (data/projects/<id>). 생략 시 데이터가 있는 전 과제 빌드")
    ap.add_argument("--validate-stage", choices=["poc", "pilot", "mass", "ops"],
                    help="이관 리허설: 빌드된 공통 레코드(records)를 지정 단계의 필수 필드 잣대로 재검증 "
                         "— 단계 이관 전 '대장이 다음 단계 잣대를 견디는가' 사전 점검")
    args = ap.parse_args()

    reg = _load_registry()
    if args.project:
        pids = [args.project]
    else:
        # config.json 이 있는 과제만 빌드 대상 (레지스트리 등재만 된 과제는 스킵)
        pids = [p["id"] for p in reg.get("projects", [])
                if (PROJECTS_ROOT / p["id"] / "config.json").exists()]
        if not pids:
            raise SystemExit("빌드할 과제가 없습니다 (data/projects/<id>/config.json 필요)")
    for pid in pids:
        try:
            build_project(pid)
        except SystemExit as ex:
            # 개발(제작) 단계 과제: config만 있고 평가 엑셀이 아직 없음 — 정상 (홈 카드는 devPlan으로 표시)
            print(f"[build:{pid}] 스킵 — {ex}")
    write_portfolio()
    write_library()

    if args.validate_stage:
        for pid in pids:
            path = PROJECTS_ROOT / pid / "dashboard.json"
            if not path.exists():
                continue
            out = json.loads(path.read_text(encoding="utf-8"))
            recs = out.get("records") or []
            cur = (out.get("config") or {}).get("stage", "mass")
            print(f"\n[리허설:{pid}] 현 단계 {cur} 대장(레코드 {len(recs)}건)을 "
                  f"'{args.validate_stage}' 잣대로 검증:")
            _validate_records(recs, args.validate_stage,
                              tag=f"이관 리허설 {cur}→{args.validate_stage}")


if __name__ == "__main__":
    main()
