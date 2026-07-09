"""
generate_demo_data.py
---------------------
웹 기능 확인용 "확장 데모 데이터셋"을 내부 정합성이 보장되도록 생성한다.

생성물:
  data/raw/260601_SEC양산평가.xlsx   ← 업체 제출 양식(일일평가/에러로그) 25일치
  data/SEC_REPORT.xlsx                ← PM 관리(코드마스터/조치검증)
  data/errors/*.jpg|png               ← 에러 모달 사진 플레이스홀더

정합성 규칙(이 스크립트가 보장):
  - 일일평가 '일일에러' 합 == 에러로그 행 수
  - 에러 '회차'(누적 Cycle) == 시뮬레이션상의 실제 발생 위치 (누적 총량 이하)
  - 일일평가 '연속성공' == 마지막 에러 이후 연속 무에러 Cycle 수
  - 조치 '검증시작Cycle' == 누적 총량 이하
  - 에러/조치의 코드는 모두 코드마스터에 존재

생성 후:  python3 scripts/build_dashboard_json.py  로 dashboard.json 재생성.
"""

from __future__ import annotations

from datetime import time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
RAW_PATH = ROOT / "data" / "raw" / "260601_SEC양산평가.xlsx"
MGMT_PATH = ROOT / "data" / "SEC_REPORT.xlsx"
ERRORS_DIR = ROOT / "data" / "errors"

# ── 시나리오 정의 ────────────────────────────────────────────────
# 근무일(월~금) 25일, 하루 Cycle 수. 합계 = 358 (목표 360 대비 99%).
DATES = [
    "2026-06-01", "2026-06-02", "2026-06-03", "2026-06-04", "2026-06-05",
    "2026-06-08", "2026-06-09", "2026-06-10", "2026-06-11", "2026-06-12",
    "2026-06-15", "2026-06-16", "2026-06-17", "2026-06-18", "2026-06-19",
    "2026-06-22", "2026-06-23", "2026-06-24", "2026-06-25", "2026-06-26",
    "2026-06-29", "2026-06-30", "2026-07-01", "2026-07-02", "2026-07-03",
]
TOTALS = [8, 10, 12, 12, 14, 14, 15, 15, 16, 16, 16, 16, 16, 16, 16,
          14, 15, 15, 16, 16, 14, 14, 14, 14, 14]
ACTIVITIES = [
    "JOB 생성 · 픽업–적재 사이클 셋업", "사이클 반복 안정성 검증", "연속 적재 정밀도 점검",
    "비전 파라미터 튜닝", "노출 보정 후 재검증", "그리퍼 압력 캘리브레이션",
    "표면 검사 단계 검증", "경로 플래너 검증", "역광 조건 재현 시험", "차광 설치 후 재검증",
    "연속 운전 안정성", "후속 로트 그립 검증", "센서 노이즈 점검", "그리퍼 패드 점검",
    "안정화 연속 운전", "무고장 연속 운전", "무고장 연속 운전", "무고장 연속 운전",
    "무고장 연속 운전", "무고장 연속 운전", "장기 신뢰성 입증", "장기 신뢰성 입증",
    "장기 신뢰성 입증", "장기 신뢰성 입증", "장기 신뢰성 입증",
]
PERSONNEL = "홍길동, 김철수"

# 에러 이벤트: (발생 누적 Cycle, 코드, 유형, 상세, 원인, 조치, 결과, 삼성, 업체, 상세설명, 사진)
ERROR_EVENTS = [
    (15, "ERR-001", "비전 인식 오류", "픽업 대상 부품 비전 좌표 인식 실패, 로봇 정지",
     "조도 변화로 카메라 노출값 부적합", "조명 LUX 재조정 + 비전 threshold 보정", "정상복귀",
     "이상혁", "박영희",
     "현장 조도 320→210 LUX 급감 구간에서 반복 발생. 노출 EV+0.7 보정 후 재현되지 않음. "
     "비전 threshold 0.82→0.74 하향으로 픽업 성공률 회복.",
     "ERR-001_1.jpg, ERR-001_2.jpg"),
    (64, "ERR-002", "그리퍼 그립 실패", "부품 표면 마찰계수 편차로 그리핑 실패, 자동 정지",
     "부품 표면 코팅 로트 편차", "그리퍼 압력 +5% 조정", "정상복귀",
     "이상혁", "홍길동", "", ""),
    (92, "ERR-003", "경로 계획 오류", "경유점 보간 오류로 경로 이탈, 자동 정지",
     "플래너 펌웨어 보간 버그", "경로 플래너 펌웨어 업데이트", "정상복귀",
     "이상혁", "김철수", "", ""),
    (120, "ERR-001", "비전 인식 오류", "역광 조건에서 좌표 인식 지연 후 타임아웃",
     "오전 특정 시간대 역광 유입", "차광 커튼 설치 + 노출 재보정", "정상복귀",
     "이상혁", "박영희", "오전 역광 유입 구간에서 ERR-001 재발. 차광 설치 후 해결.", ""),
    (150, "ERR-002", "그리퍼 그립 실패", "재그립 시도 중 슬립 재발",
     "코팅 편차 후속 로트", "표면 사전검사 단계 추가", "정상복귀",
     "이상혁", "홍길동", "", ""),
    (175, "ERR-005", "센서 노이즈", "근접 센서 노이즈로 오감지 경고 후 자가복구",
     "인접 설비 EMI 유입", "EMI 필터 추가", "정상복귀",
     "이상혁", "김철수", "자가 복구되어 라인 정지 없음. 경고만 발생.", ""),
    (190, "ERR-002", "그리퍼 그립 실패", "조치 적용 후에도 그립 실패 재발생",
     "압력 조정만으로 미해결(패드 마모)", "그리퍼 패드 교체 + 재검증 진행", "정상복귀",
     "이상혁", "홍길동",
     "조치 A-003 적용 후에도 재발 — 근본원인이 패드 마모임을 추가 확인. 패드 교체로 재검증 진행.",
     "ERR-002_grip.png"),
]

# 조치검증: (조치ID, 대상코드, 조치내용, 담당, 목표일, 상태, 검증시작Cycle)
ACTIONS = [
    ("A-001", "ERR-001", "카메라 노출/threshold 재튜닝 + 차광 설치", "양희두", "2026-06-13", "완료", 125),
    ("A-002", "ERR-003", "경로 플래너 펌웨어 업데이트", "김현일", "2026-06-16", "완료", 100),
    ("A-003", "ERR-002", "그리퍼 압력 조정 + 표면 사전검사 추가", "양희두", "2026-06-22", "진행중", 160),
    ("A-004", "ERR-005", "EMI 필터 추가", "김현일", "2026-06-26", "완료", 300),
]

CODES = [
    ["ERR-001", "Vision Timeout", "Critical", "비전 좌표 인식 실패, 로봇 정지"],
    ["ERR-002", "Gripper Slip", "Major", "그리퍼 그립 실패"],
    ["ERR-003", "Path Error", "Major", "경로 계획 오류"],
    ["ERR-005", "Sensor Noise", "Minor", "센서 노이즈/EMI"],
]

# ── 시뮬레이션: 일일 행 + 에러 행을 정합성 보장하며 생성 ────────────
def simulate():
    events = sorted(ERROR_EVENTS, key=lambda e: e[0])
    daily, errlog = [], []
    cum = 0
    no = 0
    for date, total, act in zip(DATES, TOTALS, ACTIVITIES):
        prev = cum
        cum += total
        day_events = [e for e in events if prev < e[0] <= cum]
        for e in day_events:
            no += 1
            cyc, code, typ, det, cause, action, res, sec, ven, more, imgs = e
            errlog.append([no, date, e[0], code, typ, det, cause, action, res, sec, ven, more, imgs])
        # ── 업체 실제 양식: 하루를 에러 기준으로 여러 행(세그먼트)으로 분할 ──
        # 각 세그먼트 = (에러 전 연속 성공) + (에러 1). '연속성공'은 그 세그먼트의 당일 연속 성공.
        start = prev
        for e in day_events:
            p = e[0]
            succ = p - start - 1                    # 에러 직전까지 연속 성공
            daily.append([date, PERSONNEL, act, p - start, 1, succ,
                          f"{e[1]} 발생 → 연속 리셋"])
            start = p
        tail = cum - start                          # 마지막 에러 이후(또는 무에러일) 성공 구간
        if tail > 0 or not day_events:
            daily.append([date, PERSONNEL, act, tail, 0, tail, ""])
    return daily, errlog, cum


# ── 엑셀 스타일 ─────────────────────────────────────────────────
NAVY = "1E4A7A"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="맑은 고딕", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_THIN = Side(border_style="thin", color="C9DAEE")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _header(ws, cols, widths):
    for i, (name, w) in enumerate(zip(cols, widths), start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def _row(ws, r, values, *, time_cols=()):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = BODY_FONT; c.alignment = LEFT; c.border = BORDER
        if i in time_cols:
            c.number_format = "h:mm"; c.alignment = CENTER


def build_raw(daily, errlog):
    wb = Workbook(); wb.remove(wb.active)

    ws = wb.create_sheet("일일평가")
    _header(ws, ["평가일", "입실인원", "주평가내용", "일일평가", "일일에러", "연속성공", "비고"],
            [13, 18, 34, 11, 11, 11, 30])
    for i, d in enumerate(daily, start=2):
        _row(ws, i, d)

    ws = wb.create_sheet("에러로그")
    _header(ws, ["No", "발생일", "시각", "회차", "코드", "유형", "상세", "원인", "조치", "결과",
                 "삼성 담당자", "업체 담당자", "상세설명", "사진(파일명)"],
            [6, 13, 9, 9, 11, 16, 38, 30, 38, 12, 13, 13, 52, 26])
    times = [time(9, 12), time(10, 41), time(13, 5), time(9, 38), time(14, 22), time(11, 47), time(15, 9)]
    for i, e in enumerate(errlog, start=2):
        # 회차 다음(컬럼 3)에 시각 삽입: errlog 행은 [no,date,cycle,...] 이므로 시각을 끼워 재배열
        no, date, cycle, code, typ, det, cause, action, res, sec, ven, more, imgs = e
        row = [no, date, times[(i - 2) % len(times)], cycle, code, typ, det, cause, action, res, sec, ven, more, imgs]
        _row(ws, i, row, time_cols=(3,))
        ws.row_dimensions[i].height = 40

    RAW_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(RAW_PATH)


def build_mgmt():
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("코드마스터")
    _header(ws, ["코드", "유형", "등급", "설명"], [13, 20, 12, 46])
    for i, c in enumerate(CODES, start=2):
        _row(ws, i, c)
    ws = wb.create_sheet("조치검증")
    _header(ws, ["조치ID", "대상코드", "조치내용", "담당", "목표일", "상태", "검증시작Cycle"],
            [10, 12, 38, 12, 13, 12, 14])
    for i, a in enumerate(ACTIONS, start=2):
        _row(ws, i, list(a))
    wb.save(MGMT_PATH)


# ── 에러 사진 플레이스홀더 ──────────────────────────────────────
def build_photos():
    try:
        from PIL import Image, ImageDraw, ImageFont
    except ImportError:
        print("[demo] PIL 미설치 — 사진 플레이스홀더 생략")
        return

    def font(sz):
        for n in ("DejaVuSans-Bold.ttf", "DejaVuSans.ttf"):
            try:
                return ImageFont.truetype(n, sz)
            except OSError:
                continue
        return ImageFont.load_default()

    def make(path, label, caption, bg):
        W, H = 800, 600
        img = Image.new("RGB", (W, H), bg)
        d = ImageDraw.Draw(img)
        d.rectangle([16, 16, W - 16, H - 16], outline=(255, 255, 255), width=4)
        d.line([16, 16, W - 16, H - 16], fill=(255, 255, 255), width=1)
        d.line([16, H - 16, W - 16, 16], fill=(255, 255, 255), width=1)
        for txt, f, y, fill in [("DEMO", font(54), 150, (255, 255, 255)),
                                ("error site photo", font(30), 240, (220, 230, 245)),
                                (label, font(30), 360, (255, 255, 255)),
                                (caption, font(22), 430, (210, 220, 235))]:
            l, t, r, b = d.textbbox((0, 0), txt, font=f)
            d.text(((W - (r - l)) / 2, y), txt, font=f, fill=fill)
        path.parent.mkdir(parents=True, exist_ok=True)
        if path.suffix.lower() in (".jpg", ".jpeg"):
            img.save(path, "JPEG", quality=85)
        else:
            img.save(path, "PNG")

    make(ERRORS_DIR / "ERR-001_1.jpg", "ERR-001_1.jpg", "vision misdetect - before fix", (30, 74, 122))
    make(ERRORS_DIR / "ERR-001_2.jpg", "ERR-001_2.jpg", "after exposure / threshold tuning", (47, 93, 63))
    make(ERRORS_DIR / "ERR-002_grip.png", "ERR-002_grip.png", "gripper slip - pad wear", (139, 46, 31))


def main():
    daily, errlog, cum = simulate()
    build_raw(daily, errlog)
    build_mgmt()
    build_photos()
    print(f"[demo] raw  : {RAW_PATH.relative_to(ROOT)}  (일일 {len(daily)}행, 에러 {len(errlog)}행, 누적 {cum} Cycle)")
    print(f"[demo] mgmt : {MGMT_PATH.relative_to(ROOT)}  (코드 {len(CODES)}, 조치 {len(ACTIONS)})")
    print(f"[demo] 다음:  python3 scripts/build_dashboard_json.py")


if __name__ == "__main__":
    main()
