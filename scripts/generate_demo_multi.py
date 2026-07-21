"""
generate_demo_multi.py
----------------------
드럼(POC)·분류(Pilot) 과제의 가상 데모 데이터셋을 생성한다.
  - data/projects/drum/ : config.json + raw/드럼POC_샘플.xlsx (이슈로그·런기록·비정상평가)
  - data/projects/sort/ : config.json + raw/분류Pilot_샘플.xlsx (일일평가·에러로그) + REPORT.xlsx

정합성 규칙: 일일 에러 합 = 에러로그 행수, 4분류 합 = 이슈 총수, 주차 MCBF는 누적 cycles/누적 errors.
실행: python3 scripts/generate_demo_multi.py  →  python3 scripts/build_dashboard_json.py
"""
from __future__ import annotations

import json
import shutil
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
PROJECTS = ROOT / "data" / "projects"

HDR_FILL = PatternFill("solid", fgColor="1E4A7A")
HDR_FONT = Font(color="FFFFFF", bold=True)


# ── 데모 날짜 시프트 ─────────────────────────────────────────────
# 원 설계 기준일(PIN = POC 게이트 리뷰 2026-07-15)이 항상 "내일"이 되도록 전 여정 날짜를
# 균등 이동 — 언제 재생성해도 데모가 살아있는 시점(D-1)이 된다. 스토리·간격은 그대로.
import re as _re

PIN_DATE = date(2026, 7, 15)
DATE_OFFSET = (date.today() + timedelta(days=1)) - PIN_DATE

def _shift_iso(m):
    try:
        d = date(int(m.group(1)), int(m.group(2)), int(m.group(3))) + DATE_OFFSET
        return d.isoformat()
    except ValueError:
        return m.group(0)

def _shift_md(m):
    mo, dy = int(m.group(1)), int(m.group(2))
    if not (1 <= mo <= 12 and 1 <= dy <= 31):
        return m.group(0)
    try:
        d = date(2026, mo, dy) + DATE_OFFSET
        return f"{d.month:02d}-{d.day:02d}"
    except ValueError:
        return m.group(0)

def _shift_str(s: str) -> str:
    s = _re.sub(r"(20\d{2})-(\d{2})-(\d{2})", _shift_iso, s)
    s = _re.sub(r"(?<![\w/.-])(\d{2})-(\d{2})(?![\d-])", _shift_md, s)   # 비고 속 (MM-DD)
    return s

def _shift_obj(o):
    if isinstance(o, str):
        return _shift_str(o)
    if isinstance(o, dict):
        return {k: _shift_obj(v) for k, v in o.items()}
    if isinstance(o, (list, tuple)):
        t = [_shift_obj(v) for v in o]
        return t if isinstance(o, list) else tuple(t)
    return o


def _sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in ws[1]:
        c.fill, c.font = HDR_FILL, HDR_FONT
    for r in rows:
        ws.append(_shift_obj(r))
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(12, len(str(h)) * 2 + 4)
    return ws


def _write_config(pid: str, cfg: dict):
    p = PROJECTS / pid / "config.json"
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(_shift_obj(cfg), ensure_ascii=False, indent=2), encoding="utf-8")


# ══════════════════════════ 드럼 자동화 (POC) ══════════════════════════
def gen_drum():
    pid = "drum"
    (PROJECTS / pid / "raw").mkdir(parents=True, exist_ok=True)
    (PROJECTS / pid / "errors").mkdir(exist_ok=True)

    # 이슈로그 18건 — 4분류: 컨셉 0 / 설계 6 / 구현 9 / 시험환경 3 · 종결 11 / 검증중 3 / 조치중 4
    # POC 필수 5필드 + 선택(발생일·종결일·무발생검증·상세) — 종결일이 있어야 수렴 추이(발견 vs 종결)가 그려진다
    issues = [
        # (mode, severity, cause4, status, 발생일, 종결일, 무발생검증, 상세)
        ("비전 오인식", "Major", "구현(SW)", "종결", "2026-06-08", "2026-06-12", "",
         "저조도에서 드럼 마킹 오인식 → 노출 파라미터 보정"),
        ("체결 토크 이탈", "Major", "구현(SW)", "종결", "2026-06-09", "2026-06-13", "",
         "토크 상한 파라미터 오설정 → 레시피 수정"),
        ("전장 간섭", "Minor", "시험환경", "종결", "2026-06-09", "2026-06-11", "",
         "임시 배선 간섭 (시험환경) — 정리 후 미재현"),
        ("비전 오인식", "Major", "구현(SW)", "종결", "2026-06-10", "2026-06-15", "",
         "동일 모드 재발 — 역광 조건 한정 재현 → 조명 후드 1차 개선"),
        ("그리퍼 파지 실패", "Minor", "구현(SW)", "종결", "2026-06-11", "2026-06-16", "",
         "파지 좌표 오프셋 → 티칭 보정"),
        ("비전 오인식", "Minor", "시험환경", "종결", "2026-06-12", "2026-06-14", "",
         "랩 조명 교체 공사 영향 (시험환경) — 환경 복구 후 미재현"),
        ("통신 지연", "Minor", "설계", "종결", "2026-06-13", "2026-06-19", "",
         "PLC 핸드셰이크 타임아웃 여유 부족 → 설계값 조정"),
        ("체결 토크 이탈", "Major", "구현(SW)", "종결", "2026-06-14", "2026-06-18", "",
         "토크 프로파일 과도구간 이탈 → 램프업 로직 수정"),
        ("로그 유실", "Minor", "구현(SW)", "종결", "2026-06-15", "2026-06-17", "",
         "버퍼 오버플로 → 플러시 주기 수정"),
        ("자재 급송 지연", "Minor", "설계", "종결", "2026-06-16", "2026-06-20", "",
         "급송 슈트 각도 미세 조정 (설계)"),
        ("그리퍼 파지 실패", "Minor", "구현(SW)", "종결", "2026-06-17", "2026-06-22", "",
         "동일 모드 재발 — 접근 속도 프로파일 수정"),
        ("비전 오인식", "Major", "시험환경", "검증중", "2026-06-20", "", "38/50Cy",
         "외란광 유입 각도 한정 재현 → 후드 연장 적용, 무발생 감시 중"),
        ("체결 토크 이탈", "Critical", "설계", "검증중", "2026-06-22", "", "41/50Cy",
         "반력 편차로 상한 근접 — 토크 암 강성 보강(설계) 후 검증 런"),
        ("소음 초과", "Minor", "구현(SW)", "검증중", "2026-06-24", "", "45/50Cy",
         "고속 구간 공진음 → 속도 프로파일 보정, 무발생 감시 중"),
        ("체결 토크 이탈", "Major", "설계", "조치중", "2026-06-30", "", "",
         "체결 자세 편차 기인 추정 — 지그 재설계 협의 중"),
        ("비전 오인식", "Major", "설계", "조치중", "2026-07-01", "", "",
         "고반사 드럼 한정 오인식 → 편광 필터 설계 검토 중"),
        ("그리퍼 파지 실패", "Major", "설계", "조치중", "2026-07-03", "", "",
         "드럼 림 변형품 파지 불가 → 핑거 형상 변경 설계 진행"),
        ("통신 지연", "Minor", "구현(SW)", "조치중", "2026-07-04", "", "",
         "로그 폭주 시 지연 재현 → 로그 레벨 조정 적용 중"),
    ]
    issue_rows = [
        (f"ISS-{i + 1:03d}", m, sev, c4, st, d1, d2, vf, detail, "")
        for i, (m, sev, c4, st, d1, d2, vf, detail) in enumerate(issues)
    ]

    # 런기록 — 1차 시도 31h 시점 에러 리셋 → 2차 52/72h 진행
    runs = [
        ("2026-06-29", 10, 0, "1차 시도 개시"), ("2026-06-30", 11, 0, ""),
        ("2026-07-01", 10, 1, "31h 시점 비전 오인식 → 조명 후드 개선 · 리셋"),
        ("2026-07-02", 12, 0, "2차 시도 개시"), ("2026-07-03", 11, 0, ""),
        ("2026-07-04", 10, 0, ""), ("2026-07-06", 9, 0, ""), ("2026-07-07", 10, 0, ""),
    ]
    abn = [
        ("비상정지 후 재기동", "45s", "PASS", ""), ("순간 정전 복구", "2.1m", "PASS", ""),
        ("자재 걸림 제거 후 재개", "1.4m", "PASS", ""),
        ("통신 두절 → 자동 재접속", "—", "FAIL", "재접속 로직 설계 변경 후 재시험"),
        ("도어 오픈 인터록", "30s", "PASS", ""), ("외란 조명 변화", "—", "PASS", ""),
        ("이종 자재 투입 감지", "50s", "PASS", ""), ("과부하 정지 복구", "—", "대기", "7/10 예정"),
    ]

    wb = Workbook(); wb.remove(wb.active)
    _sheet(wb, "안내", ["POC 이슈로그 — 필수 5필드(이슈ID·고장모드·심각도·원인분류·상태), 나머지 선택. "
                       "종결일을 적으면 수렴 추이(발견 vs 종결), 무발생검증(n/목표Cy)을 적으면 폐루프 진행이 표시된다. docs/RECORD_SCHEMA.md 참조"], [])
    _sheet(wb, "이슈로그", ["이슈ID", "고장모드", "심각도", "원인분류", "상태", "발생일", "종결일", "무발생검증", "상세", "사진(파일명)"], issue_rows)
    _sheet(wb, "런기록", ["일자", "런시간(h)", "에러수", "비고"], runs)
    _sheet(wb, "비정상평가", ["시나리오", "복구시간", "판정", "비고"], abn)
    wb.save(PROJECTS / pid / "raw" / "드럼POC_샘플.xlsx")

    _write_config(pid, {
        "stage": "poc",
        "run": {"target": 72, "unit": "h", "criterion": "무에러", "env": "사외 랩"},
        "tecop": [
            {"k": "T", "status": "warn", "note": "72h 무에러 2차 진행 중"},
            {"k": "E", "status": "ok", "note": "타당성 분석 — 개략 사업성 확인"},
            {"k": "C", "status": "ok", "note": "해당 없음 (사내)"},
            {"k": "O", "status": "warn", "note": "수혜부서 후보 협의 전 — Pilot 지표 정의 참여 필요"},
            {"k": "P", "status": "ok", "note": "안전인증 컨셉 합의 완료"},
        ],
        "gate": {"reviewDate": "2026-07-15", "label": "게이트 리뷰(Pilot 이관)", "criteria": [
            {"label": "① 기성능 스펙", "value": "3/3", "status": "pass"},
            {"label": "② 72h 무에러", "value": "auto:run", "status": "prog"},
            {"label": "③ 비정상 시나리오", "value": "auto:abnormal", "status": "prog"},
            {"label": "④ 상위 심각도 미해결", "value": "0건", "status": "pass"},
            {"label": "⑤ FMEA 상위 리스크", "value": "조치계획 수립", "status": "pass"},
        ]},
        "project": {"name": "드럼 자동화 (POC)", "department": "인프라 기술팀", "team": "김OO, 박OO",
                    "startDate": "2026-06-05", "endDate": "2026-07-15"},
        # 기술 개발(SW 완성도) — 케미컬과 동일 키 (한눈에 보기 하단 트랙)
        "swModules": [
            {"name": "비전 인식", "pct": 90, "group": "로봇"},
            {"name": "체결 시퀀스", "pct": 80, "group": "로봇"},
            {"name": "그리퍼 제어", "pct": 75, "group": "로봇"},
            {"name": "PLC I/F", "pct": 60, "group": "상위시스템"},
            {"name": "로그/리포트", "pct": 55, "group": "상위시스템"},
            {"name": "안전 인터록", "pct": 100, "group": "환경"},
        ],
        "lifecycle": [
            {"stage": "P1 타당성·평가항목 정의", "status": "done", "note": "FMEA 초판 · 판정기준서 v1 · 예상 ROI 산출"},
            {"stage": "P2 안전인증 컨셉미팅", "status": "done", "note": "위험원 12건 식별 → 설계 반영 합의 (06-12)"},
            {"stage": "P3 기성능 평가/검증", "status": "done", "note": "택트·반복정밀도·체결성공률 3/3 충족"},
            {"stage": "P4 사외 72h 무에러 + 비정상 평가", "status": "current", "note": "2차 시도 진행 · 비정상 6/8 PASS"},
        ],
        "ui": {
            "app": {"title": "드럼 자동화 — POC", "brandLogo": "드", "brandName": "드럼 자동화<br>POC",
                    "evalDateLabel": "평가일", "printBtn": "PDF 리포트",
                    "footBrand": "FRACAS-lite · 전수 4분류", "updatedPrefix": "업데이트 "},
            "nav": {"overview": "한눈에 보기", "all": "평가 상세 내역"},
            "common": {"stDone": "완료", "stCurrent": "진행 중", "stTodo": "예정", "noteEmpty": "— 메모"},
            "overview": {"stageTitle": "POC 세부 단계", "stageCurrentPrefix": "현재: ",
                         "stageSub": "P1→P4 · 게이트 통과 시 Pilot 이관",
                         # 부서 협의(케미컬과 동일 키 — 재빌드 불필요, config만 수정)
                         "discussItems": [
                             {"topic": "통신 재접속 로직 변경 후 재시험 일정 (비정상 FAIL 1건)", "tag": "긴급", "group": "안전"},
                             {"topic": "과부하 정지 복구 시나리오 시험 (7/10)", "tag": "진행", "group": "안전"},
                             {"topic": "편광 필터 설계 변경 — 고반사 드럼 오인식 (ISS-016)", "tag": "협의", "group": "운영"},
                             {"topic": "드럼 림 변형품 파지 — 핑거 형상 변경 (ISS-017)", "tag": "검토", "group": "운영"},
                             {"topic": "Pilot 가동 지표 정의 — 수혜부서 참여 요청", "tag": "협의", "group": "기타"},
                         ],
                         # 업무 목표 (사이드바 카드)
                         "goalsMonth": "72h 무에러 완주 + 비정상 8/8 통과\n게이트 리뷰(7/15) 산출물 준비",
                         "goalsWeek": "2차 시도 잔여 20h 완주\nISS-013 토크 암 보강 무발생 검증 마감"},
            "modal": {"title": "상세"},
        },
    })
    print(f"[demo] drum(POC): 이슈 {len(issues)} · 런기록 {len(runs)}일 · 비정상 {len(abn)}건")


# ══════════════════════════ 분류 자동화 (Pilot) ══════════════════════════
def gen_sort():
    pid = "sort"
    (PROJECTS / pid / "raw").mkdir(parents=True, exist_ok=True)
    (PROJECTS / pid / "errors").mkdir(exist_ok=True)

    # 6주 일일평가 (월~토, 주 6일) — 주차 누적 MCBF 480→1240 성장
    # 주별 (cycles, error건수): w1 (2400,5) w2 (2480,2) w3 (1960,1) w4 (1440,0) w5 (1440,1) w6 (1440,0)
    weeks = [(2400, 5), (2480, 2), (1960, 1), (1440, 0), (1440, 1), (1440, 0)]
    start = date(2026, 5, 25)  # 월요일
    daily_rows, err_slots = [], []
    for wi, (cyc, errn) in enumerate(weeks):
        days = [start + timedelta(days=wi * 7 + d) for d in range(6)]  # 월~토
        per = [cyc // 6] * 6
        per[-1] += cyc - sum(per)
        # 에러를 주 초반 날짜에 배치
        errs_per_day = [0] * 6
        for k in range(errn):
            errs_per_day[k % 3] += 1
        for d, c, e in zip(days, per, errs_per_day):
            daily_rows.append((d.isoformat(), 2, "분류·이적재 반복 운전", c, e, 0, 11, ""))
            for _ in range(e):
                err_slots.append(d)

    # 에러로그 9건 — 재발은 '파지 실패' 1개 모드만 (4회).
    # 원인분류 = 근본원인 축 (POC 4분류의 상위호환 세분화, docs/RECORD_SCHEMA.md §3)
    err_defs = [
        ("SRT-01", "파지 실패", "박스 모서리 파지 미끄러짐", "그리퍼 압력 프로파일 미흡", "설계", "파라미터 보정", "v0.9.1", "Rev B"),
        ("SRT-02", "정위치 이탈", "AMR 정지 위치 12mm 이탈", "오도메트리 드리프트", "SW", "보정 패치", "v0.9.1", "Rev B"),
        ("SRT-03", "컨베이어 I/F 타임아웃", "핸드셰이크 응답 지연", "타임아웃 설정 과소", "SW", "타임아웃 로직 변경", "v0.9.1", "Rev B"),
        ("SRT-04", "충전 도킹 실패", "도킹 마커 오염", "마커 청소 주기 부재", "운영·조작", "청소 SOP 반영", "v0.9.1", "Rev B"),
        ("SRT-01", "파지 실패", "동일 모드 재발 — 코팅 편차", "그리퍼 코팅 로트 편차", "부품", "그리퍼 Rev C 교체", "v0.9.2", "Rev B"),
        ("SRT-05", "센서 오검지", "근접 센서 노이즈", "EMI 영향", "시험환경·자재", "필터 추가", "v0.9.2", "Rev B"),
        ("SRT-06", "티칭 이탈", "적재 좌표 오차 누적", "티칭 기준점 관리 미흡", "운영·조작", "기준점 재티칭 SOP", "v0.9.2", "Rev C"),
        ("SRT-01", "파지 실패", "재발 검증 중 재현", "Rev C 적용 전 잔존", "부품", "Rev C 전수 적용", "v0.9.3", "Rev C"),
        ("SRT-01", "파지 실패", "고속 모드 한정 재현", "속도 프로파일 한계", "설계", "v0.9.4 속도 보정", "v0.9.3", "Rev C"),
    ]
    error_rows = [
        (i + 1, err_slots[i].isoformat(), f"{9 + i % 8}:{10 + i * 5 % 50:02d}", 400 * (i + 1),
         code, mode, det, cause, cls, act, "정상복귀", "이OO", "업체 김OO", sw, hw, "", "")
        for i, (code, mode, det, cause, cls, act, sw, hw) in enumerate(err_defs)
    ]

    wb = Workbook(); wb.remove(wb.active)
    _sheet(wb, "안내", ["Pilot 보고 양식 — 일일평가에 가동시간(h), 에러로그에 SW/HW버전 필수. docs/RECORD_SCHEMA.md 참조"], [])
    _sheet(wb, "일일평가", ["평가일", "입실인원", "주평가내용", "일일평가", "일일에러", "연속성공", "가동시간(h)", "비고"], daily_rows)
    _sheet(wb, "에러로그", ["No", "발생일", "시각", "회차", "코드", "유형", "상세", "원인", "원인분류", "조치", "결과",
                          "삼성 담당자", "업체 담당자", "SW버전", "HW버전", "상세설명", "사진(파일명)"], error_rows)
    wb.save(PROJECTS / pid / "raw" / "분류Pilot_샘플.xlsx")

    # 관리 엑셀 (코드마스터 + 조치검증)
    wb2 = Workbook(); wb2.remove(wb2.active)
    _sheet(wb2, "코드마스터", ["코드", "유형", "등급", "설명"], [
        ("SRT-01", "파지 실패", "Major", "그리퍼 파지 실패/미끄러짐"),
        ("SRT-02", "정위치 이탈", "Major", "AMR 정위치 정지 실패"),
        ("SRT-03", "컨베이어 I/F 타임아웃", "Minor", "설비 인터페이스 응답 지연"),
        ("SRT-04", "충전 도킹 실패", "Minor", "자동충전 도킹 실패"),
        ("SRT-05", "센서 오검지", "Minor", "근접/영역 센서 오검지"),
        ("SRT-06", "티칭 이탈", "Major", "적재 좌표 티칭 이탈"),
    ])
    _sheet(wb2, "조치검증", ["조치ID", "대상코드", "조치내용", "담당", "목표일", "상태", "검증시작"], [
        ("A-001", "SRT-01", "그리퍼 Rev C 전수 적용 + v0.9.4 속도 보정", "김OO", "2026-07-08", "검증중", "2026-06-27"),
        ("A-002", "SRT-02", "오도메트리 보정 패치 (v0.9.2)", "이OO", "2026-06-10", "검증완료", "2026-06-03"),
        ("A-003", "SRT-03", "핸드셰이크 타임아웃 로직 변경", "이OO", "2026-06-12", "검증완료", "2026-06-05"),
        ("A-004", "SRT-04", "도킹 마커 청소 주기 SOP", "박OO", "2026-06-08", "검증완료", "2026-06-02"),
        ("A-005", "SRT-05", "EMI 필터 추가", "박OO", "2026-06-15", "검증완료", "2026-06-10"),
        ("A-006", "SRT-06", "기준점 재티칭 SOP + 주기 점검", "김OO", "2026-06-20", "검증완료", "2026-06-14"),
    ])
    wb2.save(PROJECTS / pid / "REPORT.xlsx")

    _write_config(pid, {
        "stage": "pilot",
        "run": {"target": 300, "unit": "h", "criterion": "무정지", "env": "사내 (공정 연결 없이)",
                "growthTarget": 1500},
        "tecop": [
            {"k": "T", "status": "ok", "note": "MCBF 성장곡선 순항"},
            {"k": "E", "status": "warn", "note": "양산 원가 절감안 검토 중 (그리퍼 단가)"},
            {"k": "C", "status": "ok", "note": "공급계약 초안 합의"},
            {"k": "O", "status": "ok", "note": "수혜부서 가동 지표 정의 참여 완료"},
            {"k": "P", "status": "ok", "note": "안전인증 심사 일정 정상"},
        ],
        "gate": {"reviewDate": "2026-07-29", "label": "게이트 리뷰(양산시범 이관)", "criteria": [
            {"label": "① MCBF 성장 목표", "value": "auto:growth", "status": "prog"},
            {"label": "② 만성(재발) 고장", "value": "1건 — 마감 필요", "status": "fail"},
            {"label": "③ 시정조치 검증마감", "value": "auto:actions", "status": "prog"},
            {"label": "④ 안전인증서", "value": "심사 중 (~07-20)", "status": "prog"},
            {"label": "⑤ 무정지 300h", "value": "설계 동결(07-25) 후", "status": "wait"},
        ]},
        "project": {"name": "분류 자동화 (Pilot)", "department": "인프라 기술팀", "team": "이OO, 김OO",
                    "startDate": "2026-05-25", "endDate": "2026-08-31"},
        "swModules": [
            {"name": "파지 제어", "pct": 85, "group": "로봇"},
            {"name": "주행/오도메트리", "pct": 90, "group": "로봇"},
            {"name": "분류 로직", "pct": 95, "group": "로봇"},
            {"name": "컨베이어 I/F", "pct": 70, "group": "상위시스템"},
            {"name": "관제 연동", "pct": 40, "group": "상위시스템"},
            {"name": "충전 도킹", "pct": 80, "group": "환경"},
        ],
        "lifecycle": [
            {"stage": "L1 가동 지표 정의", "status": "done", "note": "MCBF/MTTR/가동률 산식 확정 — 양산시범과 동일 산식"},
            {"stage": "L2 신뢰성 평가/검증 (TAAF)", "status": "current", "note": "반복운전→FRACAS→재운전 · 성장 추적"},
            {"stage": "L3 신뢰성시험사양서 · 임시 사용 안전 승인", "status": "current", "note": "v0.9 협의 — 300h 판정·리셋 규칙 명문화"},
            {"stage": "L4 무정지 300h", "status": "todo", "note": "설계 동결(07-25 예정) 후 본 런 · 현재 예열 런"},
        ],
        "ui": {
            "app": {"title": "분류 자동화 — Pilot", "brandLogo": "분", "brandName": "분류 자동화<br>Pilot",
                    "evalDateLabel": "평가일", "printBtn": "PDF 리포트",
                    "footBrand": "FRACAS + MCBF 성장 추적", "updatedPrefix": "업데이트 "},
            "nav": {"overview": "한눈에 보기", "all": "평가 상세 내역"},
            "common": {"stDone": "완료", "stCurrent": "진행 중", "stTodo": "예정", "noteEmpty": "— 메모"},
            "overview": {"stageTitle": "Pilot 세부 단계", "stageCurrentPrefix": "현재: ",
                         "stageSub": "L1→L4 · 게이트 통과 시 양산 시범 평가 이관",
                         "discussItems": [
                             {"topic": "안전인증 심사 서류 보완 (~7/20)", "tag": "진행", "group": "안전"},
                             {"topic": "임시 사용 안전 승인 갱신 — 야간 무인 런 포함 여부", "tag": "협의", "group": "안전"},
                             {"topic": "그리퍼 Rev C 전수 적용 — 만성 재발(파지) 마감", "tag": "긴급", "group": "운영"},
                             {"topic": "설계 동결(7/25) 전 변경 등급 × 재검증 매트릭스 합의", "tag": "협의", "group": "운영"},
                             {"topic": "양산 시범 평가 계약 파라미터 초안 (수혜부서)", "tag": "검토", "group": "기타"},
                         ],
                         "goalsMonth": "설계 동결(7/25) → 무정지 300h 본 런 착수\n만성 재발(파지 실패) 마감",
                         "goalsWeek": "그리퍼 Rev C 무발생 검증 완료\n안전인증 심사 서류 제출"},
            "modal": {"title": "상세"},
        },
    })
    n_err_daily = sum(r[4] for r in daily_rows)
    assert n_err_daily == len(error_rows), f"정합성 위반: 일일에러 합 {n_err_daily} ≠ 에러로그 {len(error_rows)}"
    print(f"[demo] sort(Pilot): daily {len(daily_rows)}일 · 에러 {len(error_rows)}건 · 조치 6건 (정합 확인)")


# ══════════════════════════ 크린룸 반송 (확산) ══════════════════════════
def gen_clean():
    pid = "clean"
    (PROJECTS / pid / "raw").mkdir(parents=True, exist_ok=True)
    (PROJECTS / pid / "errors").mkdir(exist_ok=True)

    # 호기퀄 — 설치 → SAT → 축약 무고장 런(48h) · 8호기/4라인
    units = [
        ("1호기", "L1", "2026-05-12", "PASS", 48, 48, "퀄 완료", ""),
        ("2호기", "L1", "2026-05-14", "PASS", 48, 48, "퀄 완료", ""),
        ("3호기", "L2", "2026-05-26", "PASS", 48, 48, "퀄 완료", ""),
        ("4호기", "L2", "2026-05-28", "PASS", 48, 48, "퀄 완료", ""),
        ("5호기", "L3", "2026-06-09", "PASS", 48, 48, "퀄 완료", ""),
        ("6호기", "L3", "2026-06-11", "PASS", 31, 48, "런 진행", "공진 이슈(CIS-001) 보강 후 재개"),
        ("7호기", "L4", "2026-06-23", "진행", 0, 48, "SAT 진행", ""),
        ("8호기", "L4", "2026-07-02", "—", 0, 48, "설치 중", ""),
    ]
    # 이슈로그 — 원인계층(설계/제작·조립/설치·시공/운영·환경) · 설계성 = 전 함대 리스크
    issues = [
        # (mode, sev, 계층, status, 발생, 종결, 무발생, 호기, 상세)
        ("반송 암 공진 진동", "Major", "설계", "조치중", "2026-07-02", "", "", "6호기",
         "고속 구간 공진 — 설계성 판정, 전 함대 브래킷 보강 설계 전개 중 (에스컬레이션)"),
        ("케이블 체결 불량", "Minor", "제작·조립", "종결", "2026-05-16", "2026-05-18", "", "2호기",
         "하네스 체결 토크 미달 — 조립 체크리스트 개정"),
        ("베어링 프리로드 편차", "Major", "제작·조립", "종결", "2026-05-30", "2026-06-05", "", "4호기",
         "제작 로트 편차 — 수입검사 항목 추가"),
        ("커넥터 미압착", "Minor", "제작·조립", "종결", "2026-06-12", "2026-06-14", "", "5호기",
         "압착 불량 1건 — 전수 재확인"),
        ("레일 수평도 초과", "Major", "설치·시공", "종결", "2026-05-13", "2026-05-15", "", "1호기",
         "수평도 0.8mm 초과 — 재시공 후 SAT 재수행"),
        ("통신 음영 끊김", "Major", "설치·시공", "종결", "2026-05-27", "2026-06-02", "", "3호기",
         "AP 음영 구간 — AP 1식 증설"),
        ("간섭물 미철거", "Minor", "설치·시공", "종결", "2026-06-10", "2026-06-11", "", "5호기",
         "가설 브래킷 잔존 — 시공 인수 체크리스트 보완"),
        ("전원 노이즈 트립", "Major", "설치·시공", "검증중", "2026-06-24", "", "12/20일", "7호기",
         "접지 불량 추정 — 접지 보강 후 무발생 감시"),
        ("티칭 기준점 상이", "Minor", "설치·시공", "종결", "2026-06-25", "2026-06-27", "", "7호기",
         "호기 간 기준점 관리 표준 반영"),
        ("이물 유입 오검지", "Minor", "운영·환경", "종결", "2026-06-16", "2026-06-18", "", "3호기",
         "크린룸 파티클 이벤트 — 센서 임계 재설정"),
        ("수동 개입 절차 미준수", "Minor", "운영·환경", "조치중", "2026-06-30", "", "", "4호기",
         "인터록 해제 절차 위반 — 운영 SOP 재교육"),
        ("온습도 경보 정지", "Minor", "운영·환경", "조치중", "2026-07-06", "", "", "6호기",
         "공조 이벤트 연동 정지 — 재기동 조건 협의"),
    ]
    issue_rows = [(f"CIS-{i+1:03d}", m, sev, c, st, d1, d2, vf, u, det, "")
                  for i, (m, sev, c, st, d1, d2, vf, u, det) in enumerate(issues)]

    wb = Workbook(); wb.remove(wb.active)
    _sheet(wb, "안내", ["확산 보고 — 이슈로그(원인계층·호기 필수) + 호기퀄(설치→SAT→축약 런). "
                       "설계성 고장 = 전 함대 리스크, 즉시 에스컬레이션. docs/PROCESS.md §2.5"], [])
    _sheet(wb, "이슈로그", ["이슈ID", "고장모드", "심각도", "원인분류", "상태", "발생일", "종결일",
                          "무발생검증", "호기", "상세", "사진(파일명)"], issue_rows)
    _sheet(wb, "호기퀄", ["호기", "라인", "설치일", "SAT", "축약런(h)", "목표(h)", "상태", "비고"], units)
    wb.save(PROJECTS / pid / "raw" / "크린룸확산_샘플.xlsx")

    _write_config(pid, {
        "stage": "spread",
        # unit=호기: 홈 카드 진행 표기(퀄 완료 호기 n/m)용 — 축약 런 자체 목표는 target(h)
        "run": {"target": 48, "unit": "호기", "criterion": "축약 런", "env": "각 적용 라인 (호기별 SAT)"},
        "tecop": [
            {"k": "T", "status": "warn", "note": "공진 이슈 — 전 함대 브래킷 보강 전개 중"},
            {"k": "E", "status": "ok", "note": "확산 ROI — 투자심의 승인 조건 유지"},
            {"k": "C", "status": "ok", "note": "호기 추가 발주 계약 정상"},
            {"k": "O", "status": "ok", "note": "라인별 운영 인수 교육 3/4 완료"},
            {"k": "P", "status": "ok", "note": "호기별 가동인증 일정 정상"},
        ],
        "gate": {"reviewDate": "2026-07-31", "label": "확산 완료 리뷰", "criteria": [
            {"label": "① 호기별 퀄 (SAT+런)", "value": "auto:fleet", "status": "prog"},
            {"label": "② 설계성 고장 0", "value": "1건 — 보강 전개 중", "status": "fail"},
            {"label": "③ 가동지표 검증", "value": "5/5 호기 (완료분)", "status": "prog"},
            {"label": "④ 기준 구성 동결", "value": "승인 편차 1건", "status": "pass"},
            {"label": "⑤ 횡전개 문서", "value": "설치 표준 v1.1", "status": "pass"},
        ]},
        "project": {"name": "크린룸 반송 자동화 (확산)", "department": "인프라 기술팀", "team": "박OO, 정OO",
                    "startDate": "2026-05-10", "endDate": "2026-08-31"},
        "swModules": [
            {"name": "반송 제어 (동결 v2.3)", "pct": 100, "group": "로봇"},
            {"name": "호기 파라미터 셋", "pct": 85, "group": "로봇"},
            {"name": "관제 연동", "pct": 90, "group": "상위시스템"},
            {"name": "설치 자동 점검 스크립트", "pct": 70, "group": "상위시스템"},
            {"name": "라인별 안전 인터록", "pct": 100, "group": "환경"},
        ],
        "lifecycle": [
            {"stage": "S1 기준 구성 동결", "status": "done", "note": "승인 편차만 예외 — 변경관리 절차"},
            {"stage": "S2 호기 설치·SAT", "status": "current", "note": "8호기 중 SAT 6 통과 · 1 진행 · 1 설치중"},
            {"stage": "S3 축약 무고장 런 (48h)", "status": "current", "note": "5호기 통과 · 6호기 재개 대기"},
            {"stage": "S4 가동지표 검증·횡전개", "status": "todo", "note": "완료 호기부터 순차 — 설치 표준 개정"},
        ],
        "ui": {
            "app": {"title": "크린룸 반송 — 확산", "brandLogo": "크", "brandName": "크린룸 반송<br>확산",
                    "evalDateLabel": "평가일", "printBtn": "PDF 리포트",
                    "footBrand": "원인계층 + 호기별 층화", "updatedPrefix": "업데이트 "},
            "nav": {"overview": "한눈에 보기", "all": "평가 상세 내역"},
            "common": {"stDone": "완료", "stCurrent": "진행 중", "stTodo": "예정", "noteEmpty": "— 메모"},
            "overview": {"stageTitle": "확산 세부 단계", "stageCurrentPrefix": "현재: ",
                         "stageSub": "S1→S4 · 호기별 양산 퀄 (설치→SAT→축약 런)",
                         "discussItems": [
                             {"topic": "브래킷 보강 — 전 함대 리스크 전개 계획 승인 (CIS-001)", "tag": "긴급", "group": "안전"},
                             {"topic": "접지 보강 표준 — 신규 라인 시공 사양 반영", "tag": "협의", "group": "운영"},
                             {"topic": "8호기 설치 일정 — 라인 정지 슬롯 협의", "tag": "진행", "group": "운영"},
                             {"topic": "운영 인수 교육 4차 (L4)", "tag": "검토", "group": "기타"},
                         ],
                         "goalsMonth": "6호기 재개 → 8/8 퀄 완료\n설계성(공진) 보강 전 함대 전개",
                         "goalsWeek": "7호기 SAT 완료\n전원 노이즈 무발생 감시 마감"},
            "modal": {"title": "상세"},
        },
    })
    print(f"[demo] clean(확산): 호기 {len(units)} · 이슈 {len(issues)}건")


# ══════════════════════════ 물류 AMR (운영/관제) ══════════════════════════
def gen_agv():
    pid = "agv"
    (PROJECTS / pid / "raw").mkdir(parents=True, exist_ok=True)
    (PROJECTS / pid / "errors").mkdir(exist_ok=True)

    monthly = [
        ("2026-01", 96.2, 210, 18, 26.4, 420, 6),
        ("2026-02", 96.9, 240, 15, 21.8, 388, 4),
        ("2026-03", 97.4, 265, 13, 18.2, 356, 3),
        ("2026-04", 97.8, 290, 12, 15.1, 330, 3),
        ("2026-05", 98.1, 315, 10, 12.6, 302, 2),
        ("2026-06", 98.4, 340, 9, 9.8, 280, 2),
    ]
    # 필드 FRACAS — 알람에서 승격된 건만 (승격 기준: 정지 유발 또는 반복 3회↑)
    issues = [
        ("경로 계획 교착", "Major", "설계", "종결", "2026-01-22", "2026-02-05", "", "AGV-07", 360,
         "교차로 데드락 — SW v2.1 패치 전 함대 배포"),
        ("승강 모터 과열", "Major", "제작·조립", "종결", "2026-02-14", "2026-02-20", "", "AGV-03", 480,
         "냉각 유로 막힘 — 방열 구조 개선"),
        ("충전 도킹 실패", "Minor", "운영·환경", "종결", "2026-03-08", "2026-03-12", "", "AGV-02", 120,
         "도킹 마커 오염 — 청소 주기 단축 + 가이드 형상 개선(CIP-01)"),
        ("리프트 센서 오검지", "Minor", "설치·시공", "종결", "2026-03-27", "2026-04-02", "", "AGV-05", 60,
         "센서 브래킷 유격 — 시공 표준 개정"),
        ("통신 음영 정지", "Major", "설치·시공", "종결", "2026-04-15", "2026-04-25", "", "AGV-09", 240,
         "AP 음영 맵 반영 — AP 증설 + 경로 회피(CIP-03)"),
        ("팔레트 인식 실패", "Minor", "운영·환경", "종결", "2026-05-04", "2026-05-06", "", "AGV-01", 75,
         "랩핑 반사 — 노출 프로파일 추가"),
        ("충전 도킹 실패", "Minor", "운영·환경", "종결", "2026-05-19", "2026-05-22", "", "AGV-11", 90,
         "동일 모드 재발 — 가이드 개선판 적용 전 호기"),
        ("승강 모터 과열", "Major", "제작·조립", "검증중", "2026-06-02", "", "18/30일", "AGV-12", 150,
         "동일 모드 재발 — 냉각 팬 추가(CIP-02) 무발생 감시"),
        ("타이어 마모 슬립", "Minor", "운영·환경", "검증중", "2026-06-10", "", "600/1,000Cy", "AGV-04", 45,
         "고부하 구간 마모 가속 — 소재 변경(CIP-04) 검증"),
        ("비상정지 오탐", "Minor", "설치·시공", "조치중", "2026-06-24", "", "", "AGV-06", 30,
         "라이트커튼 정렬 편차 — 재정렬 및 원인 분석 중"),
    ]
    issue_rows = [(f"FLD-{i+1:03d}", m, sev, c, st, d1, d2, vf, u, dt, det, "")
                  for i, (m, sev, c, st, d1, d2, vf, u, dt, det) in enumerate(issues)]
    cip = [
        ("CIP-01", "충전 도킹 가이드 형상 개선", "충전 도킹 실패", "완료", "도킹 실패 월 5→1건"),
        ("CIP-02", "승강부 냉각 팬 추가", "승강 모터 과열", "검증중", "과열 정지 재발 차단"),
        ("CIP-03", "AP 음영 맵 기반 경로 회피", "통신 음영 정지", "완료", "음영 정지 0건 (2개월)"),
        ("CIP-04", "타이어 소재 변경 (고부하 구간)", "타이어 마모 슬립", "진행", "교체 주기 2배 연장"),
        ("CIP-05", "PM → CBM 전환 (모터 전류 추세)", "승강 모터 과열", "진행", "계획 정지 30% 축소"),
    ]

    wb = Workbook(); wb.remove(wb.active)
    _sheet(wb, "안내", ["운영/관제 보고 — 월간지표(RAM) + 이슈로그(필드 FRACAS: 알람 승격 건, 다운타임·호기 필수) + CIP. "
                       "질문: '어떤 고장부터 없애는 게 경제적인가' — 다운타임 Pareto. docs/PROCESS.md §2.5·§3"], [])
    _sheet(wb, "이슈로그", ["이슈ID", "고장모드", "심각도", "원인분류", "상태", "발생일", "종결일",
                          "무발생검증", "호기", "다운타임(분)", "상세", "사진(파일명)"], issue_rows)
    _sheet(wb, "월간지표", ["월", "가동률(%)", "MTBF(h)", "MTTR(분)", "다운타임(h)", "알람수", "승격수"], monthly)
    _sheet(wb, "CIP", ["CIP ID", "과제", "대상모드", "상태", "기대효과"], cip)
    wb.save(PROJECTS / pid / "raw" / "물류AMR운영_샘플.xlsx")

    _write_config(pid, {
        "stage": "ops",
        "acceptance": {"availTargetPct": 98.0},
        # unit=%: 홈 카드 진행 표기(가동률/목표)용
        "run": {"target": 98, "unit": "%", "criterion": "가동률", "env": "양산 라인 (12호기 함대)"},
        "tecop": [
            {"k": "T", "status": "ok", "note": "가동률 목표 달성 — CBM 전환 진행"},
            {"k": "E", "status": "ok", "note": "다운타임 비용 월 -18% (CIP 효과)"},
            {"k": "C", "status": "ok", "note": "유지보수 계약 갱신 완료"},
            {"k": "O", "status": "warn", "note": "야간 비상 대응 훈련 1회 미실시"},
            {"k": "P", "status": "ok", "note": "정기 안전점검 일정 정상"},
        ],
        "gate": {"reviewDate": "2026-07-31", "label": "월간 RAM 리뷰", "criteria": [
            {"label": "① 가동률 ≥98%", "value": "auto:avail", "status": "pass"},
            {"label": "② MTBF 성장", "value": "340h (+62% YTD)", "status": "pass"},
            {"label": "③ MTTR ≤10분", "value": "9분", "status": "pass"},
            {"label": "④ 만성(재발) 마감", "value": "2모드 — CIP 검증 중", "status": "prog"},
            {"label": "⑤ CIP 진행", "value": "완료 2 · 진행 3", "status": "prog"},
        ]},
        "project": {"name": "물류 AMR (운영/관제)", "department": "인프라 기술팀", "team": "정OO, 한OO",
                    "startDate": "2026-01-01", "endDate": "—"},
        "swModules": [
            {"name": "알람 자동 수집", "pct": 100, "group": "상위시스템"},
            {"name": "CBM 모델 (모터 전류)", "pct": 60, "group": "상위시스템"},
            {"name": "월간 RAM 자동 리포트", "pct": 80, "group": "상위시스템"},
            {"name": "함대 SW 배포 (OTA)", "pct": 90, "group": "로봇"},
            {"name": "비상 대응 매트릭스", "pct": 100, "group": "환경"},
        ],
        "lifecycle": [
            {"stage": "O1 통합관제·알람 수집", "status": "done", "note": "알람 자동수집 — 승격 기준 확정"},
            {"stage": "O2 비상 대응 체계", "status": "done", "note": "에스컬레이션 매트릭스 · 훈련 3/4"},
            {"stage": "O3 PM → CBM 전환", "status": "current", "note": "모터 전류 추세 모델 시범 (CIP-05)"},
            {"stage": "O4 CIP · FMEA 환류", "status": "current", "note": "필드 고장모드 → 차기 과제 FMEA 반영"},
        ],
        "ui": {
            "app": {"title": "물류 AMR — 운영/관제", "brandLogo": "물", "brandName": "물류 AMR<br>운영/관제",
                    "evalDateLabel": "평가일", "printBtn": "PDF 리포트",
                    "footBrand": "필드 FRACAS + 다운타임 Pareto + CIP", "updatedPrefix": "업데이트 "},
            "nav": {"overview": "한눈에 보기", "all": "평가 상세 내역"},
            "common": {"stDone": "완료", "stCurrent": "진행 중", "stTodo": "예정", "noteEmpty": "— 메모"},
            "overview": {"stageTitle": "운영/관제 체계", "stageCurrentPrefix": "현재: ",
                         "stageSub": "O1→O4 · 필드 고장모드는 차기 과제 FMEA로 환류",
                         "discussItems": [
                             {"topic": "야간 비상 대응 훈련 일정 (미실시 1회)", "tag": "긴급", "group": "안전"},
                             {"topic": "승강 모터 냉각 팬 — 전 함대 확대 적용 여부", "tag": "협의", "group": "운영"},
                             {"topic": "CBM 전환 범위 — 구동부 추가 검토", "tag": "검토", "group": "운영"},
                             {"topic": "차기 과제 FMEA 환류 회의 (7월)", "tag": "진행", "group": "기타"},
                         ],
                         "goalsMonth": "가동률 98.5% 유지\n만성 2모드 CIP 검증 마감",
                         "goalsWeek": "AGV-12 무발생 감시 18/30일\n비상 대응 훈련 4차 실시"},
            "modal": {"title": "상세"},
        },
    })
    print(f"[demo] agv(운영): 월간지표 {len(monthly)}개월 · 필드 FRACAS {len(issues)}건 · CIP {len(cip)}건")


# ══════════════════════════ 팔레타이저 (양산 시범 평가 — 순수 가상) ══════════════════════════
def gen_pack():
    """양산 시범 평가(mass) 템플릿의 순수 가상 데모 — chem은 SEC 샘플 사본이라 시연이 조심스러움.
    config는 chem을 복제해 과제 고유 문구만 치환한다 (계약 파라미터·ui 구조 = 표준 그대로)."""
    pid = "pack"
    (PROJECTS / pid / "raw").mkdir(parents=True, exist_ok=True)
    (PROJECTS / pid / "errors").mkdir(exist_ok=True)

    # 일일평가 24일 (주 6일 × 4주 · 일 15Cy = 계약 360Cy) — 에러일 3일(5건), 연속성공은 에러일 리셋
    days, d = [], date(2026, 6, 8)
    while len(days) < 24:
        if d.weekday() < 6:
            days.append(d)
        d += timedelta(days=1)
    err_at = {5: 2, 11: 1, 17: 2}   # 날짜 인덱스 → 에러 건수 (합 5)
    daily_rows, err_cycles = [], []
    streak = cum = 0
    for i, dt in enumerate(days):
        e = err_at.get(i, 0)
        streak = 0 if e else streak + 15
        for k in range(e):
            err_cycles.append((dt, cum + 7 + k * 4))
        cum += 15
        daily_rows.append((dt.isoformat(), 2, "팔레트 적재·반송 반복 운전", 15, e, streak,
                           "에러 발생 — 연속 리셋" if e else ""))

    # 에러로그 5건 — PAL-001 재발 1회 포함 (원인분류 = 근본원인 축)
    err_defs = [
        ("PAL-001", "파렛 파지 낙하", "흡착 그리퍼 진공 저하로 고중량 박스 낙하", "진공 패드 마모", "부품",
         "패드 교체 주기 단축 + 진공압 모니터링", "v1.0.2"),
        ("PAL-002", "비전 오인식", "혼적 파렛 상단 박스 좌표 오인식", "저조도 구간 노출 부적합", "SW",
         "노출 프로파일 보정", "v1.0.2"),
        ("PAL-001", "파렛 파지 낙하", "동일 모드 재발 — 고중량 박스 한정 재현", "흡착력 설계 여유 부족", "설계",
         "그리퍼 흡착부 증설 설계 진행", "v1.0.3"),
        ("PAL-003", "적재 경로 간섭", "3단 적재 시 기둥 간섭 근접 정지", "경로 마진 설정 과소", "SW",
         "경로 플래너 마진 재설정", "v1.0.3"),
        ("PAL-004", "컨베이어 I/F 타임아웃", "출하 컨베이어 핸드셰이크 지연", "PLC 응답 타임아웃 과소", "SW",
         "타임아웃 로직 변경", "v1.0.3"),
    ]
    error_rows = [
        (i + 1, dt.isoformat(), f"{9 + i * 2}:{15 + i * 7 % 40:02d}", cyc,
         code, mode, det, cause, cls, act, "정상복귀", "양OO", "업체 정OO", sw, "Rev A", "", "")
        for i, ((dt, cyc), (code, mode, det, cause, cls, act, sw)) in enumerate(zip(err_cycles, err_defs))
    ]

    wb = Workbook(); wb.remove(wb.active)
    _sheet(wb, "안내", ["양산 시범 평가 보고 — 일일평가(사이클·에러·연속성공) + 에러로그. 판정·처분은 REPORT.xlsx(PM). docs/RECORD_SCHEMA.md"], [])
    _sheet(wb, "일일평가", ["평가일", "입실인원", "주평가내용", "일일평가", "일일에러", "연속성공", "비고"], daily_rows)
    _sheet(wb, "에러로그", ["No", "발생일", "시각", "회차", "코드", "유형", "상세", "원인", "원인분류", "조치", "결과",
                          "삼성 담당자", "업체 담당자", "SW버전", "HW버전", "상세설명", "사진(파일명)"], error_rows)
    wb.save(PROJECTS / pid / "raw" / "팔레타이저_샘플.xlsx")

    # 관리 엑셀 (PM): 코드마스터 · 조치검증 · 판정대장 · 처분대장
    wb2 = Workbook(); wb2.remove(wb2.active)
    _sheet(wb2, "코드마스터", ["코드", "유형", "등급", "설명"], [
        ("PAL-001", "파렛 파지 낙하", "Critical", "흡착 파지 실패로 박스 낙하 — 안전 관련"),
        ("PAL-002", "비전 오인식", "Major", "팔레트/박스 좌표 인식 오류"),
        ("PAL-003", "적재 경로 간섭", "Major", "적재 경로 상 구조물 간섭 근접"),
        ("PAL-004", "컨베이어 I/F 타임아웃", "Minor", "출하 설비 인터페이스 응답 지연"),
    ])
    _sheet(wb2, "조치검증", ["조치ID", "대상코드", "조치내용", "담당", "목표일", "상태", "검증시작(날짜/Cycle)"], [
        ("A-001", "PAL-001", "그리퍼 흡착부 증설 + 패드 주기 단축", "정OO", "2026-07-20", "진행중", 262),
        ("A-002", "PAL-002", "노출 프로파일 보정 (v1.0.2)", "양OO", "2026-06-20", "완료", "2026-06-15"),
        ("A-003", "PAL-003", "경로 플래너 마진 재설정", "양OO", "2026-07-05", "완료", "2026-06-30"),
        ("A-004", "PAL-004", "핸드셰이크 타임아웃 로직 변경", "정OO", "2026-07-05", "검증완료", "2026-06-30"),
    ])
    _sheet(wb2, "판정대장", ["사건ID", "대상에러No", "판정", "귀책분류", "증거", "합의상태", "판정일"], [
        ("JD-01", 1, "관련", "설비 (부품)", "진공압 로그 + 재현시험", "합의완료", "2026-06-16"),
        ("JD-02", 2, "관련", "설비 (SW)", "영상 + 조도 로그", "합의완료", "2026-06-17"),
        ("JD-03", 3, "관련", "설비 (설계)", "재현시험 (고중량 한정)", "합의완료", "2026-06-24"),
        ("JD-04", 4, "관련", "설비 (SW)", "경로 로그", "합의완료", "2026-07-01"),
        ("JD-05", 5, "판정중", "", "PLC 로그 분석 중 — 자재/설비 판별", "합동리뷰 07-16", ""),
    ])
    _sheet(wb2, "처분대장", ["처분ID", "대상ID", "처분", "사유", "기한", "오너", "합의"], [
        ("DSP-01", "PAL-003", "종결예정", "마진 재설정 후 무발생 검증 중 — 심의 전 종결 예상", "2026-08-10", "양OO", "합의완료"),
        ("DSP-02", "PAL-004", "waiver", "경미·타임아웃 재설정으로 해소 — 위험수용, 월간 모니터링", "—", "정OO", "합의완료"),
    ])
    wb2.save(PROJECTS / pid / "REPORT.xlsx")

    # config: chem 표준 config 복제 후 과제 고유 문구만 치환 (계약 파라미터·화면 구조는 표준 그대로)
    cfg = json.loads((PROJECTS / "chem" / "config.json").read_text(encoding="utf-8"))
    for k in ("_stage_readme", "_swModules_readme", "_ui_readme", "_edit_guide"):
        cfg.pop(k, None)
    cfg["project"] = {"name": "팔레타이저 (양산 시범 평가)", "department": "인프라 기술팀",
                      "team": "정OO, 양OO", "startDate": "2026-06-08", "endDate": "2026-08-31"}
    cfg["gate"] = {"reviewDate": "2026-08-28", "label": "가동인증"}
    cfg["tecop"] = [
        {"k": "T", "status": "warn", "note": "PAL-001 재발 — 흡착부 증설 검증 전"},
        {"k": "E", "status": "ok", "note": "투자심의 입력용 실증치 축적 중"},
        {"k": "C", "status": "ok", "note": "유지보수 계약 조건 합의"},
        {"k": "O", "status": "ok", "note": "수혜부서 합동판정 참여 정상"},
        {"k": "P", "status": "ok", "note": "설치 상태 위험성 평가 완료"},
    ]
    cfg["swModules"] = [
        {"name": "비전 인식", "pct": 95, "group": "로봇"},
        {"name": "적재 패턴 플래너", "pct": 90, "group": "로봇"},
        {"name": "흡착 그리퍼 제어", "pct": 80, "group": "로봇"},
        {"name": "WMS 연동", "pct": 70, "group": "상위시스템"},
        {"name": "리포트 연동", "pct": 60, "group": "상위시스템"},
        {"name": "안전 인터록", "pct": 100, "group": "환경"},
    ]
    cfg["ui"]["app"].update({"title": "팔레타이저 — 양산 시범 평가", "brandLogo": "팔",
                             "brandName": "팔레타이저<br>양산 시범 평가"})
    ov = cfg["ui"].setdefault("overview", {})
    ov.update({
        "goalsMonth": "연속 360Cy 완주 재도전 — PAL-001 마감\n판정중 1건(JD-05) 합동리뷰",
        "goalsWeek": "흡착부 증설 검증 런 개시\n출하 I/F 재시험",
        "discussItems": [
            {"topic": "PAL-001 흡착부 증설 — 검증 런 일정", "tag": "긴급", "group": "안전"},
            {"topic": "JD-05 합동리뷰 (7/16) — 자재/설비 판별", "tag": "협의", "group": "운영"},
            {"topic": "가동인증 산출물 체크리스트 사전 점검", "tag": "검토", "group": "기타"},
        ],
        "lineCaption": "현재 평가 <b>적재 셀 2 · {cum}/{target}</b> · 셀 1 통과 · 출하 I/F 연동 대기",
    })
    _write_config(pid, cfg)
    # 라인 레이아웃 이미지: chem 샘플 재사용 (없어도 동작 — onerror 처리)
    src_img = PROJECTS / "chem" / "assets" / "line_layout.png"
    if src_img.exists():
        (PROJECTS / pid / "assets").mkdir(exist_ok=True)
        shutil.copy(src_img, PROJECTS / pid / "assets" / "line_layout.png")
    n_err = sum(r[4] for r in daily_rows)
    assert n_err == len(error_rows), f"정합성 위반: 일일에러 합 {n_err} ≠ 에러로그 {len(error_rows)}"
    print(f"[demo] pack(양산평가): daily {len(daily_rows)}일 · 에러 {len(error_rows)}건 · 판정 5 · 처분 2 (정합 확인)")


# ══════════════════════════ 용접 협동로봇 (개발 중 — 평가 착수 전) ══════════════════════════
def gen_weld():
    """config만 생성 (평가 엑셀 없음) — 순수 개발(제작) 기간의 과제.
    홈 카드는 devPlan(마일스톤·일정·평가 착수 예정)으로 '무엇을 언제까지, 지금 어디까지'를 표시한다."""
    pid = "weld"
    _write_config(pid, {
        "stage": "poc",
        "run": {"target": 72, "unit": "h", "criterion": "무에러", "env": "사외 랩"},
        "project": {"name": "용접 협동로봇 (개발 중)", "department": "인프라 기술팀", "team": "한OO, 최OO",
                    "startDate": "2026-06-01", "endDate": "2026-11-30"},
        # 개발(제작) 계획 — 평가 데이터가 생기기 전 카드의 단일 출처
        "devPlan": {
            "label": "설계·제작 (평가 착수 전)",
            "start": "2026-06-01", "end": "2026-09-01",
            "evalStart": "2026-09-08", "evalLabel": "사외 72h 무에러 + 비정상 평가",
            "items": [
                {"name": "요구사양·컨셉 설계", "due": "2026-06-20", "pct": 100},
                {"name": "FMEA 초판·판정기준서 v1", "due": "2026-07-05", "pct": 100},
                {"name": "본체·토치부 상세 설계", "due": "2026-07-25", "pct": 60},
                {"name": "지그·시제 제작", "due": "2026-08-10", "pct": 25},
                {"name": "제어 SW 프레임", "due": "2026-08-20", "pct": 30},
                {"name": "사외 랩 셋업·시운전", "due": "2026-09-01", "pct": 0},
            ],
        },
        "tecop": [
            {"k": "T", "status": "warn", "note": "토치 냉각 설계 검증 전 — 시제에서 확인"},
            {"k": "E", "status": "ok", "note": "타당성 분석 완료 — 예상 ROI 산출"},
            {"k": "C", "status": "ok", "note": "개발 계약 체결 완료"},
            {"k": "O", "status": "ok", "note": "수혜부서 요구사양 합의"},
            {"k": "P", "status": "warn", "note": "용접 흄 안전 컨셉 미팅 예정 (7/22)"},
        ],
        "gate": {"reviewDate": "2026-09-08", "label": "평가 착수 리뷰(DR)", "criteria": [
            {"label": "① 상세 설계 승인", "value": "60%", "status": "prog"},
            {"label": "② FMEA·판정기준서", "value": "v1 확정", "status": "pass"},
            {"label": "③ 시제 제작", "value": "25%", "status": "prog"},
            {"label": "④ 안전 컨셉 합의", "value": "미팅 7/22", "status": "wait"},
            {"label": "⑤ 랩 셋업", "value": "9/1 예정", "status": "wait"},
        ]},
        "lifecycle": [
            {"stage": "P1 타당성·평가항목 정의", "status": "done", "note": "FMEA 초판 · 판정기준서 v1"},
            {"stage": "P2 설계·제작", "status": "current", "note": "상세 설계 60% · 시제 25%"},
            {"stage": "P3 기성능 평가/검증", "status": "todo", "note": "시제 완성 후"},
            {"stage": "P4 사외 72h 무에러 + 비정상 평가", "status": "todo", "note": "9/8 착수 예정"},
        ],
        "ui": {
            "app": {"title": "용접 협동로봇 — 개발 중", "brandLogo": "용", "brandName": "용접 협동로봇<br>개발 중",
                    "evalDateLabel": "평가일", "printBtn": "PDF 리포트",
                    "footBrand": "설계·제작 — 평가 착수 전", "updatedPrefix": "업데이트 "},
            "nav": {"overview": "한눈에 보기", "all": "평가 상세 내역"},
        },
    })
    print("[demo] weld(개발 중): config만 생성 — 평가 착수 전 (devPlan 카드)")


# ═════════════════ 드럼 자동화 — 한 과제의 4단계 여정 (POC→Pilot→양산평가→양산적용) ═════════════════
# 목적 시연: 공통 어휘(DRM-01~08)가 전 단계를 관통하고, Critical(DRM-01 체결 토크 이탈)을
# POC에서 우선 소진 → Pilot부터 Critical 0 → 뒤 단계가 싸진다 (심각도 깔때기).
DRM_CODES = [
    ("DRM-01", "체결 토크 이탈", "Critical", "토크 상한 초과/이탈 체결 — 파손 위험"),
    ("DRM-02", "비전 오인식", "Major", "드럼 마킹/좌표 인식 오류"),
    ("DRM-03", "그리퍼 파지 실패", "Major", "파지 미끄러짐·낙하"),
    ("DRM-04", "통신 지연", "Minor", "PLC/상위 통신 지연·두절"),
    ("DRM-05", "전장 간섭", "Minor", "배선·EMI 간섭"),
    ("DRM-06", "자재 급송 지연", "Minor", "드럼 급송 슈트 지연"),
    ("DRM-07", "소음 초과", "Minor", "구동 구간 소음/공진"),
    ("DRM-08", "로그 유실", "Minor", "제어 로그 누락"),
]


def _codes_sheet(wb):
    _sheet(wb, "코드마스터", ["코드", "유형", "등급", "설명"], DRM_CODES)


def _poc_asset_svgs(pid: str):
    """설계·셋업 도식 SVG 플레이스홀더 — 실사진 교체 전 데모용 (assets/)."""
    a = PROJECTS / pid / "assets"
    a.mkdir(parents=True, exist_ok=True)
    def svg(name, title, sub, body):
        (a / name).write_text(
            f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 640 400">'
            f'<rect width="640" height="400" fill="#F2F6FA"/><rect width="640" height="400" fill="none" stroke="#C9DCEC" stroke-width="2"/>'
            f'{body}'
            f'<text x="24" y="42" font-size="22" font-weight="700" fill="#0F2E54" font-family="sans-serif">{title}</text>'
            f'<text x="24" y="68" font-size="14" fill="#5A6B7E" font-family="sans-serif">{sub}</text>'
            f'<text x="616" y="382" font-size="12" fill="#8A99AC" text-anchor="end" font-family="sans-serif">데모 도식 — 실사진 교체 예정</text></svg>',
            encoding="utf-8")
    svg("design-concept.svg", "컨셉 레이아웃", "6축 로봇 + 비전 체결 · 컨베이어 직결 반송",
        '<rect x="60" y="150" width="200" height="170" rx="10" fill="#DCE9F6" stroke="#2E89D6" stroke-width="2"/>'
        '<text x="160" y="240" font-size="16" text-anchor="middle" fill="#2E6DB0" font-family="sans-serif">6축 로봇 셀</text>'
        '<rect x="300" y="220" width="280" height="46" rx="8" fill="#E6F3EB" stroke="#3E9B6E" stroke-width="2"/>'
        '<text x="440" y="249" font-size="15" text-anchor="middle" fill="#2f7a52" font-family="sans-serif">반송 컨베이어 →</text>'
        '<circle cx="160" cy="130" r="26" fill="#FBF3E6" stroke="#E08600" stroke-width="2"/>'
        '<text x="160" y="136" font-size="13" text-anchor="middle" fill="#B36F0A" font-family="sans-serif">비전</text>')
    svg("design-gripper.svg", "그리퍼 핑거 v1", "림 변형 여유 2mm · 코팅 사양 A",
        '<path d="M240 140 L240 300 L290 300 L290 210 L350 210 L350 300 L400 300 L400 140 Z" fill="#DCE9F6" stroke="#2E89D6" stroke-width="2"/>'
        '<circle cx="320" cy="330" r="34" fill="#F7E5E2" stroke="#C0392B" stroke-width="2" stroke-dasharray="5 4"/>'
        '<text x="320" y="336" font-size="12" text-anchor="middle" fill="#C0392B" font-family="sans-serif">드럼 림</text>')
    svg("asis-manual.svg", "현장 수작업 공정 (As-Is)", "작업자 2인 — 토크 렌치 체결 · 대차 반송 · T/T 18초",
        '<circle cx="150" cy="200" r="22" fill="#DCE9F6" stroke="#5A6B7E" stroke-width="2"/>'
        '<rect x="138" y="226" width="24" height="52" rx="8" fill="#DCE9F6" stroke="#5A6B7E" stroke-width="2"/>'
        '<circle cx="240" cy="200" r="22" fill="#DCE9F6" stroke="#5A6B7E" stroke-width="2"/>'
        '<rect x="228" y="226" width="24" height="52" rx="8" fill="#DCE9F6" stroke="#5A6B7E" stroke-width="2"/>'
        '<text x="195" y="308" font-size="12" text-anchor="middle" fill="#5A6B7E" font-family="sans-serif">작업자 2인 · 토크 렌치</text>'
        '<circle cx="400" cy="240" r="42" fill="#F6EBDA" stroke="#B36F0A" stroke-width="2"/>'
        '<text x="400" y="246" font-size="13" text-anchor="middle" fill="#8a5408" font-family="sans-serif">드럼</text>'
        '<rect x="470" y="270" width="120" height="40" rx="7" fill="#EEF2F7" stroke="#8A99AC" stroke-width="2"/>'
        '<text x="530" y="295" font-size="12" text-anchor="middle" fill="#5A6B7E" font-family="sans-serif">대차 반송</text>'
        '<path d="M448 240 L466 285" stroke="#8A99AC" stroke-width="2" marker-end="none" stroke-dasharray="5 4"/>')
    svg("build-cell.svg", "사외 랩 셀 셋업", "프레임·안전 펜스·비전 조명 설치 완료",
        '<rect x="80" y="120" width="480" height="200" rx="8" fill="none" stroke="#5A6B7E" stroke-width="2" stroke-dasharray="7 5"/>'
        '<rect x="120" y="170" width="150" height="120" rx="8" fill="#DCE9F6" stroke="#2E89D6" stroke-width="2"/>'
        '<rect x="330" y="200" width="190" height="60" rx="8" fill="#E6F3EB" stroke="#3E9B6E" stroke-width="2"/>'
        '<text x="195" y="238" font-size="14" text-anchor="middle" fill="#2E6DB0" font-family="sans-serif">로봇</text>'
        '<text x="425" y="236" font-size="14" text-anchor="middle" fill="#2f7a52" font-family="sans-serif">컨베이어</text>')


def gen_drum_poc():
    pid = "drum-poc"
    (PROJECTS / pid / "raw").mkdir(parents=True, exist_ok=True)
    (PROJECTS / pid / "errors").mkdir(exist_ok=True)
    # 18건 — Critical 2(체결 토크: 1 종결·1 검증중 = 조기 우선 조치), Major 7, Minor 9
    issues = [
        ("비전 오인식", "Major", "구현(SW)", "종결", "2026-06-08", "2026-06-12", "", "저조도 마킹 오인식 → 노출 보정"),
        ("체결 토크 이탈", "Critical", "구현(SW)", "종결", "2026-06-09", "2026-06-11", "", "토크 상한 초과 체결 — 파손 위험, 최우선 조치(레시피·상한 인터록)"),
        ("전장 간섭", "Minor", "시험환경", "종결", "2026-06-09", "2026-06-11", "", "임시 배선 간섭 — 정리 후 미재현"),
        ("비전 오인식", "Major", "구현(SW)", "종결", "2026-06-10", "2026-06-15", "", "역광 한정 재현 → 후드 1차 개선"),
        ("그리퍼 파지 실패", "Minor", "구현(SW)", "종결", "2026-06-11", "2026-06-16", "", "파지 좌표 오프셋 → 티칭 보정"),
        ("비전 오인식", "Minor", "시험환경", "종결", "2026-06-12", "2026-06-14", "", "랩 조명 공사 영향 — 복구 후 미재현"),
        ("통신 지연", "Minor", "설계", "종결", "2026-06-13", "2026-06-19", "", "핸드셰이크 타임아웃 여유 부족 → 설계값 조정"),
        ("체결 토크 이탈", "Major", "구현(SW)", "종결", "2026-06-14", "2026-06-18", "", "과도구간 이탈 → 램프업 로직 수정"),
        ("로그 유실", "Minor", "구현(SW)", "종결", "2026-06-15", "2026-06-17", "", "버퍼 오버플로 → 플러시 주기 수정"),
        ("자재 급송 지연", "Minor", "설계", "종결", "2026-06-16", "2026-06-20", "", "급송 슈트 각도 조정"),
        ("그리퍼 파지 실패", "Minor", "구현(SW)", "종결", "2026-06-17", "2026-06-22", "", "접근 속도 프로파일 수정"),
        ("비전 오인식", "Major", "시험환경", "검증중", "2026-06-20", "", "38/50Cy", "외란광 각도 한정 → 후드 연장, 무발생 감시"),
        ("체결 토크 이탈", "Critical", "설계", "검증중", "2026-06-22", "", "41/50Cy", "반력 편차 상한 근접 — 토크 암 강성 보강(설계), 크리티컬 우선순위 1"),
        ("소음 초과", "Minor", "구현(SW)", "검증중", "2026-06-24", "", "45/50Cy", "고속 공진음 → 속도 프로파일 보정"),
        ("체결 토크 이탈", "Major", "설계", "조치중", "2026-06-30", "", "", "체결 자세 편차 추정 — 지그 재설계 협의"),
        ("비전 오인식", "Major", "설계", "조치중", "2026-07-01", "", "", "고반사 드럼 한정 → 편광 필터 검토"),
        ("그리퍼 파지 실패", "Major", "설계", "조치중", "2026-07-03", "", "", "림 변형품 파지 불가 → 핑거 형상 변경"),
        ("통신 지연", "Minor", "구현(SW)", "조치중", "2026-07-04", "", "", "로그 폭주 시 지연 → 로그 레벨 조정"),
    ]
    rows = [(f"ISS-{i+1:03d}", m, sv, c4, st, d1, d2, vf, det, "")
            for i, (m, sv, c4, st, d1, d2, vf, det) in enumerate(issues)]
    # 런 기록은 평가일 전일까지 끊김 없이 관리 (미가동일 제외) — 2차 시도 누적 52h 유지
    runs = [
        ("2026-06-29", 10, 0, "1차 시도 개시"), ("2026-06-30", 11, 0, ""),
        ("2026-07-01", 10, 1, "31h 시점 비전 오인식 → 조명 후드 개선 · 리셋"),
        ("2026-07-02", 7, 0, "2차 시도 개시"), ("2026-07-03", 6, 0, ""),
        ("2026-07-04", 5, 0, ""), ("2026-07-06", 4, 0, ""), ("2026-07-07", 6, 0, ""),
        ("2026-07-08", 5, 0, ""), ("2026-07-09", 5, 0, "토크 암 보강 반영 후 런 재개"),
        ("2026-07-10", 4, 0, ""), ("2026-07-11", 5, 0, ""), ("2026-07-13", 5, 0, "무발생 검증 병행"),
    ]
    abn = [
        ("비상정지 후 재기동", "45s", "PASS", ""), ("순간 정전 복구", "2.1m", "PASS", ""),
        ("자재 걸림 제거 후 재개", "1.4m", "PASS", ""),
        ("통신 두절 → 자동 재접속", "—", "FAIL", "재접속 로직 설계 변경 후 재시험"),
        ("도어 오픈 인터록", "30s", "PASS", ""), ("외란 조명 변화", "—", "PASS", ""),
        ("이종 자재 투입 감지", "50s", "PASS", ""), ("과부하 정지 복구", "—", "대기", "07-10 예정"),
    ]
    wb = Workbook(); wb.remove(wb.active)
    _sheet(wb, "안내", ["POC 이슈로그 — 필수 5필드 + 선택(종결일·무발생검증). 어휘 v1(코드마스터)은 P2 미팅 산출물. docs/RECORD_SCHEMA.md"], [])
    _sheet(wb, "이슈로그", ["이슈ID", "고장모드", "심각도", "원인분류", "상태", "발생일", "종결일", "무발생검증", "상세", "사진(파일명)"], rows)
    _sheet(wb, "런기록", ["일자", "런시간(h)", "에러수", "비고"], runs)
    _sheet(wb, "비정상평가", ["시나리오", "복구시간", "판정", "비고"], abn)
    _codes_sheet(wb)
    wb.save(PROJECTS / pid / "raw" / "드럼POC_샘플.xlsx")
    _poc_asset_svgs(pid)
    _write_config(pid, {
        "stage": "poc",
        "run": {"target": 72, "unit": "h", "criterion": "무에러", "env": "사외 랩"},
        "tecop": [
            {"k": "T", "status": "warn", "note": "Critical(체결 토크) 보강 검증 중 — 우선순위 1", "risks": [
                {"id": "R-T1", "risk": "체결 토크 반력 편차 — 파손 위험", "level": "High", "mitigation": "토크 암 강성 보강 + 상한 인터록", "owner": "김OO", "due": "2026-07-14", "progress": 80, "status": "완화중", "link": "ISS-013"},
                {"id": "R-T2", "risk": "그리퍼 파지 실패 만성화 가능성", "level": "Medium", "mitigation": "핑거 형상 변경 — Pilot 이관 전 설계 반영", "owner": "박OO", "due": "2026-07-15", "progress": 40, "status": "완화중", "link": "ISS-017"},
            ]},
            {"k": "E", "status": "ok", "note": "투자 효과 개략 확인 (인력·처리량)", "risks": [
                {"id": "R-E1", "risk": "투자 효과 가정(처리량) 실측 미확보", "level": "Medium", "mitigation": "72h 런 실증치로 투자심의 입력 확보", "owner": "PM", "due": "2026-07-20", "progress": 70, "status": "완화중"},
            ]},
            {"k": "C", "status": "ok", "note": "업체 개발 계약 정상 · 잣대 서면 확정", "risks": [
                {"id": "R-C1", "risk": "합격 잣대 변경 요청 발생 가능성", "level": "Low", "mitigation": "서면 동결 유지 — 변경 요청 즉시 에스컬레이션", "owner": "PM", "due": "", "progress": 0, "status": "감시(수용)"},
            ]},
            {"k": "O", "status": "ok", "note": "담당 R&R·업체 대응 체계 정상", "risks": [
                {"id": "R-O1", "risk": "업체 상주 종료 후 대응 공백", "level": "Medium", "mitigation": "원격 대응 SLA — Pilot 계약에 명기", "owner": "김OO", "due": "2026-07-22", "progress": 10, "status": "식별"},
            ]},
            {"k": "P", "status": "warn", "note": "수혜부서 Pilot 지표 정의 참여 필요 (안전인증 컨셉 합의 완료)", "risks": [
                {"id": "R-P1", "risk": "수혜부서 Pilot 지표 정의 미합의", "level": "Medium", "mitigation": "3자 협의체 — 가동 지표 초안 합의 (07-13 회의)", "owner": "PM", "due": "2026-07-15", "progress": 60, "status": "완화중"},
                {"id": "R-P2", "risk": "안전인증 컨셉 부적합 가능성", "level": "Low", "mitigation": "인증기관 사전 컨셉 미팅 — 합의서 확보 (P2)", "owner": "PM", "due": "2026-06-10", "progress": 100, "status": "완화 완료"},
            ]},
        ],
        # phase = 이 기준이 결판나는 세부 단계 번호 (종합 클리어의 단계×기준 통합 행에 사용, 없으면 공통 조건)
        "gate": {"reviewDate": "2026-07-15", "label": "게이트 리뷰(Pilot 이관)", "criteria": [
            {"label": "① 기성능 스펙", "value": "3/3", "status": "pass", "phase": 3},
            {"label": "② 72h 무에러", "value": "auto:run", "status": "prog", "phase": 5},
            {"label": "③ 비정상 시나리오", "value": "auto:abnormal", "status": "prog", "phase": 5},
            {"label": "④ Critical 미해결 0", "value": "현재 1건 · 무발생 41/50Cy", "status": "prog"},
            {"label": "⑤ FMEA 상위 리스크", "value": "조치계획 수립", "status": "pass", "phase": 1},
        ]},
        "project": {"name": "드럼 자동화 (POC)", "department": "인프라 기술팀", "team": "김OO, 박OO",
                    "startDate": "2026-06-05", "endDate": "2026-07-15"},
        "swModules": [
            {"name": "비전 인식", "pct": 90, "group": "로봇"}, {"name": "체결 시퀀스", "pct": 80, "group": "로봇"},
            {"name": "그리퍼 제어", "pct": 75, "group": "로봇"}, {"name": "PLC I/F", "pct": 60, "group": "상위시스템"},
            {"name": "로그/리포트", "pct": 55, "group": "상위시스템"}, {"name": "안전 인터록", "pct": 100, "group": "환경"},
        ],
        # POC 세부 5단계 — 1~2(기획·제작)는 기획/제작 관제, 3~5(평가)는 에러·런 관제로 화면 전환.
        # clear[] = 그 단계에서 클리어해야 하는 항목 (종합 클리어 타일 — 단계 선택 시 교체 표시,
        # 미정의면 게이트 기준 전체로 폴백 = 최종 단계)
        "lifecycle": [
            {"stage": "P1 과제 기획", "status": "done", "note": "목표·평가항목 정의 · 개략 ROI · 컨셉 확정 · 특허 검토 · FMEA 초판(어휘 v1)", "clear": [
                {"label": "목표·평가항목 정의", "value": "판정기준서 v1", "status": "pass",
                 "summary": "합격 기준 서면 합의 — 72h 무에러 · 비정상 8종 · 기성능 3항목 (사후 변경 금지)"},
                {"label": "개략 ROI 산정", "value": "투자심의 입력 초안", "status": "pass",
                 "summary": "인력 2인/교대 절감 + 처리량 +8% → 회수 2.1년 — 실증치로 정밀화 예정"},
                {"label": "컨셉 도출", "value": "2안 비교 — 로봇 안", "status": "pass",
                 "summary": "6축 로봇+비전 vs 겐트리 — 모델 변경 유연성 우위로 로봇 안 채택"},
                {"label": "특허 검토", "value": "저촉 0 · 출원 후보 1", "status": "pass",
                 "summary": "선행 12건 검토 — 저촉 0 · 회피 설계 1건 반영 · 그리퍼 구조 출원 후보"},
                {"label": "안전인증 컨셉", "value": "위험원 12건 반영", "status": "pass",
                 "summary": "위험성 평가 v0 — 협착·낙하 등 12건 식별 → 펜스·인터록 설계 반영 합의"},
                {"label": "FMEA 초판", "value": "어휘 v1 (DRM-01~08)", "status": "pass",
                 "summary": "고장모드 어휘 8종 정의 — 이후 전 단계 에러 분류·라이브러리의 기준"},
            ]},
            {"stage": "P2 설계·제작", "status": "done", "note": "설계 동결 v1 · 사외 랩 셀 셋업 완료 (06-27)", "clear": [
                {"label": "설계 동결", "value": "v1 (06-19)", "status": "pass"},
                {"label": "유닛 제작", "value": "업체 완료 (06-20)", "status": "pass"},
                {"label": "랩 셀 셋업", "value": "사외 설치 (06-27)", "status": "pass"},
                {"label": "SW 1차 구현", "value": "90% — 레시피 편집기 잔여", "status": "prog"},
            ]},
            {"stage": "P3 기/성능 평가", "status": "done", "note": "택트·반복정밀도·체결성공률 3/3 충족 · Critical 2건 조기 발굴→우선 조치", "clear": [
                {"label": "택트 타임", "value": "11.8s / 목표 12s", "status": "pass"},
                {"label": "반복 정밀도", "value": "±0.04mm / 스펙 ±0.05", "status": "pass"},
                {"label": "체결 성공률", "value": "99.2% / 목표 99%", "status": "pass"},
                {"label": "Critical 조기 조치", "value": "2건 발굴 → 우선 조치", "status": "pass"},
            ]},
            {"stage": "P4 SW 체크리스트", "status": "done", "note": "핵심 모듈 점검 — 비전·시퀀스·인터록 (잔여 2건 P5 병행)", "clear": [
                {"label": "비전·시퀀스 검증", "value": "체크 완료", "status": "pass"},
                {"label": "안전 인터록 체크", "value": "전 항목 PASS", "status": "pass"},
                {"label": "레시피 편집기", "value": "잔여 — P5 병행", "status": "prog"},
                {"label": "로그/리포트", "value": "55% 구현", "status": "prog"},
            ]},
            {"stage": "P5 사외 72h 무에러 + 비정상 평가", "status": "current", "note": "2차 시도 52h · 비정상 6/8"},
        ],
        # P1 기획 화면 데이터 — As-Is 현장 수작업 / To-Be 자동화 컨셉 (산출물 정의는 lifecycle P1 clear)
        "pocPlan": {
            "goal": "드럼 체결·반송 자동화 — 수작업 2인 공정 대체, T/T 12초 이내",
            "roi": "인력 2인/교대 절감 + 처리량 +8% — 개략 ROI 2.1년 (투자심의 입력 초안)",
            "concept": "6축 로봇 + 비전 가이드 체결 · 컨베이어 직결 반송 (겐트리 안 대비 유연성 우위로 선정)",
            "asIs": {
                "summary": "드럼 2종을 작업자 2인이 수작업으로 체결·반송 — 토크 렌치 2점 체결 후 대차로 이동",
                "tt": "T/T 18초", "people": "2인/교대",
                "steps": ["드럼 정렬·클램프", "수동 토크 체결 (2점)", "검사·기록", "대차 적재·반송"],
                "pain": ["T/T 18초 — 라인 병목", "수공구 토크 편차 — 품질 산포", "근골격 부하 · 협착 위험"],
                "photo": "asis-manual.svg",
            },
            "toBe": {
                "summary": "6축 로봇 + 비전 가이드 체결 · 컨베이어 직결 반송",
                "targets": ["T/T 12초 이내 (−33%)", "인력 2인/교대 → 감시 1인", "토크 편차 ±3% — 파손 리스크 제거"],
                "why": "겐트리 안 대비 모델 변경 유연성 우위 — 2안 비교 후 채택",
                "photo": "design-concept.svg",
            },
            # P1 부가 정보 — 일정 계획 · 2안 비교 · ROI 내역 (FMEA 어휘는 dashboard codes 사용)
            "schedule": [
                {"from": "2026-06-05", "to": "2026-06-12"},
                {"from": "2026-06-12", "to": "2026-06-27"},
                {"from": "2026-06-27", "to": "2026-07-04"},
                {"from": "2026-07-04", "to": "2026-07-08"},
                {"from": "2026-07-08", "to": "2026-07-15"},
            ],
            "compare": {
                "cols": ["로봇 안 (채택)", "겐트리 안"],
                "rows": [
                    ["모델 변경 유연성", "◎ 티칭 변경으로 대응", "△ 기구 개조 필요"],
                    ["투자비 (상대)", "1.0 (기준)", "0.85"],
                    ["T/T", "12초", "11초"],
                    ["정비·확장", "표준 로봇 — 용이", "전용 기구 — 제약"],
                    ["잔여 리스크", "파지·비전 인식", "정렬 치구 마모"],
                ],
                "verdict": "유연성·확장성 우위로 로봇 안 채택 — T/T 열세는 반송 직결로 상쇄",
            },
            "roiDetail": {
                "invest": [["6축 로봇·비전", "1.8억"], ["제작·설치", "1.2억"], ["SW·통합", "0.6억"]],
                "effect": [["인력 2인/교대 절감", "연 1.4억"], ["처리량 +8%", "연 0.3억"]],
                "payback": "회수 약 2.1년 — POC·Pilot 실증치로 정밀화 후 투자심의 입력",
            },
        },
        # P2 설계·제작 현황 (아이템 진척 + 설계·셋업 사진 assets/)
        "pocBuild": {
            "items": [
                {"name": "기구 설계 — 체결 유닛", "pct": 100, "note": "설계 동결 v1"},
                {"name": "기구 설계 — 반송 유닛", "pct": 100, "note": ""},
                {"name": "전장 설계", "pct": 100, "note": ""},
                {"name": "프레임·유닛 제작", "pct": 100, "note": "업체 제작 완료 (06-20)"},
                {"name": "로봇·비전 셋업", "pct": 100, "note": "사외 랩 설치 (06-27)"},
                {"name": "SW 1차 구현", "pct": 90, "note": "레시피 편집기 잔여 — P5 병행"},
            ],
            "photos": [
                {"file": "design-concept.svg", "cap": "컨셉 레이아웃 — 체결+반송 직결"},
                {"file": "design-gripper.svg", "cap": "그리퍼 핑거 설계 v1 (림 여유 2mm)"},
                {"file": "build-cell.svg", "cap": "사외 랩 셀 셋업 (06-27)"},
            ],
        },
        "ui": {
            "app": {"title": "드럼 자동화 — POC", "brandLogo": "드", "brandName": "드럼 자동화<br>POC",
                    "evalDateLabel": "평가일", "printBtn": "PDF 리포트", "footBrand": "FRACAS-lite · 전수 4분류 · Critical 우선", "updatedPrefix": "업데이트 "},
            "nav": {"overview": "한눈에 보기", "all": "평가 상세 내역"},
            "common": {"stDone": "완료", "stCurrent": "진행 중", "stTodo": "예정", "noteEmpty": "— 메모"},
            "overview": {"stageTitle": "POC 세부 단계", "stageCurrentPrefix": "현재: ", "stageSub": "P1→P4 · 게이트 통과 시 Pilot 이관",
                "discussItems": [
                    {"topic": "Critical: 토크 암 강성 보강 검증 런 (ISS-013)", "tag": "긴급", "group": "안전"},
                    {"topic": "통신 재접속 로직 변경 후 재시험", "tag": "진행", "group": "안전"},
                    {"topic": "편광 필터 설계 변경 (ISS-016)", "tag": "협의", "group": "운영"},
                    {"topic": "Pilot 가동 지표 정의 — 수혜부서 참여", "tag": "협의", "group": "기타"},
                ],
                "goalsMonth": "Critical 미해결 0 달성 (ISS-013 마감)\n72h 무에러 완주",
                "goalsWeek": "2차 시도 잔여 20h\n토크 암 보강 무발생 검증 마감"},
            "modal": {"title": "상세"},
        },
    })
    print(f"[demo] drum-poc: 이슈 {len(issues)} (Critical 2 조기 조치) · 런 {len(runs)}일 · 비정상 {len(abn)}")


def gen_drum_pilot():
    pid = "drum-pilot"
    (PROJECTS / pid / "raw").mkdir(parents=True, exist_ok=True)
    (PROJECTS / pid / "errors").mkdir(exist_ok=True)
    weeks = [(2400, 4), (2480, 2), (1960, 1), (1440, 1), (1440, 1), (1440, 0)]
    start = date(2026, 7, 20)
    daily_rows, err_slots = [], []
    for wi, (cyc, errn) in enumerate(weeks):
        days = [start + timedelta(days=wi * 7 + d) for d in range(6)]
        per = [cyc // 6] * 6
        per[-1] += cyc - sum(per)
        epd = [0] * 6
        for k in range(errn):
            epd[k % 3] += 1
        for d, c, e in zip(days, per, epd):
            daily_rows.append((d.isoformat(), 2, "체결·반송 반복 운전", c, e, 0, 11, ""))
            for _ in range(e):
                err_slots.append(d)
    # 9건 — Critical 0 (POC에서 소진!) · DRM-03 만성 4회
    err_defs = [
        ("DRM-03", "그리퍼 파지 실패", "드럼 림 변형품 파지 미끄러짐", "핑거 형상 여유 부족", "설계", "핑거 Rev B 형상 변경", "v0.9.1", "Rev A"),
        ("DRM-02", "비전 오인식", "고반사 드럼 한정 오인식", "편광 필터 미적용 개체", "부품", "필터 전수 적용", "v0.9.1", "Rev A"),
        ("DRM-04", "통신 지연", "핸드셰이크 응답 지연", "타임아웃 여유 과소", "SW", "타임아웃 로직 변경", "v0.9.1", "Rev A"),
        ("DRM-03", "그리퍼 파지 실패", "동일 모드 재발 — 코팅 편차", "그리퍼 코팅 로트 편차", "부품", "코팅 Rev C 교체", "v0.9.2", "Rev A"),
        ("DRM-07", "소음 초과", "고속 구간 공진 재현", "속도 프로파일 한계", "설계", "프로파일 보정", "v0.9.2", "Rev A"),
        ("DRM-02", "비전 오인식", "야간 조도 저하 오인식", "노출 스케줄 미적용", "SW", "시간대 노출 스케줄", "v0.9.2", "Rev B"),
        ("DRM-03", "그리퍼 파지 실패", "재발 검증 중 재현", "Rev C 적용 전 잔존", "부품", "Rev C 전수 적용", "v0.9.3", "Rev B"),
        ("DRM-04", "통신 지연", "로그 폭주 시 지연", "로그 레벨 과다", "SW", "레벨 조정 배포", "v0.9.3", "Rev B"),
        ("DRM-03", "그리퍼 파지 실패", "고속 모드 한정 재현", "속도-파지 간섭", "설계", "v0.9.4 속도 보정", "v0.9.3", "Rev B"),
    ]
    error_rows = [
        (i + 1, err_slots[i].isoformat(), f"{9 + i % 8}:{10 + i * 5 % 50:02d}", 400 * (i + 1),
         code, mode, det, cause, cls, act, "정상복귀", "김OO", "업체 박OO", sw, hw, "", "")
        for i, (code, mode, det, cause, cls, act, sw, hw) in enumerate(err_defs)
    ]
    wb = Workbook(); wb.remove(wb.active)
    _sheet(wb, "안내", ["Pilot — 에러로그에 코드(DRM)·원인분류·SW/HW버전 필수. POC 대장이 그대로 이어진다."], [])
    _sheet(wb, "일일평가", ["평가일", "입실인원", "주평가내용", "일일평가", "일일에러", "연속성공", "가동시간(h)", "비고"], daily_rows)
    _sheet(wb, "에러로그", ["No", "발생일", "시각", "회차", "코드", "유형", "상세", "원인", "원인분류", "조치", "결과",
                          "삼성 담당자", "업체 담당자", "SW버전", "HW버전", "상세설명", "사진(파일명)"], error_rows)
    wb.save(PROJECTS / pid / "raw" / "드럼Pilot_샘플.xlsx")
    wb2 = Workbook(); wb2.remove(wb2.active)
    _codes_sheet(wb2)
    _sheet(wb2, "조치검증", ["조치ID", "대상코드", "조치내용", "담당", "목표일", "상태", "검증시작"], [
        ("A-001", "DRM-03", "코팅 Rev C 전수 + v0.9.4 속도 보정", "김OO", "2026-09-05", "검증중", "2026-08-25"),
        ("A-002", "DRM-02", "편광 필터 전수 + 노출 스케줄", "박OO", "2026-08-10", "검증완료", "2026-08-01"),
        ("A-003", "DRM-04", "타임아웃·로그 레벨 배포", "박OO", "2026-08-12", "검증완료", "2026-08-05"),
        ("A-004", "DRM-07", "속도 프로파일 보정", "김OO", "2026-08-08", "검증완료", "2026-08-02"),
    ])
    wb2.save(PROJECTS / pid / "REPORT.xlsx")
    _write_config(pid, {
        "stage": "pilot",
        "run": {"target": 300, "unit": "h", "criterion": "무정지", "env": "사내 (공정 연결 없이)", "growthTarget": 1500},
        "tecop": [
            {"k": "T", "status": "ok", "note": "Critical 0 유지 (POC 소진) — MCBF 성장 순항", "risks": [
                {"id": "R-T1", "risk": "파지 실패(DRM-03) 검증 실패 시 게이트 지연", "level": "Medium", "mitigation": "Rev C·v0.9.4 무발생 검증 병행 — 주간 점검", "owner": "박OO", "due": "2026-09-01", "progress": 60, "status": "완화중", "link": "E-009"},
            ]},
            {"k": "E", "status": "warn", "note": "그리퍼 코팅 단가 절감안 검토 (TCO 가정)", "risks": [
                {"id": "R-E1", "risk": "코팅 Rev C 단가 — TCO 가정 초과", "level": "Medium", "mitigation": "대체 코팅 2안 견적 — 8월 초 단가 확정", "owner": "PM", "due": "2026-08-08", "progress": 50, "status": "완화중", "link": "E-004"},
            ]},
            {"k": "C", "status": "ok", "note": "공급계약 초안 합의 · 잣대 동결 유지", "risks": [
                {"id": "R-C1", "risk": "양산 이관 잣대(360Cy) 변경 요청 가능성", "level": "Low", "mitigation": "서면 동결 유지 — 변경 요청 즉시 에스컬레이션", "owner": "PM", "due": "", "progress": 0, "status": "감시(수용)"},
            ]},
            {"k": "O", "status": "ok", "note": "업체 개선 대응 정상"},
            {"k": "P", "status": "ok", "note": "수혜부서 지표 정의 참여 완료 · 안전인증 심사 일정 정상", "risks": [
                {"id": "R-P1", "risk": "임시 사용 안전 승인(L3) 서류 지연", "level": "Low", "mitigation": "안전인증 심사 서류 선제 제출 — 완료", "owner": "PM", "due": "2026-08-20", "progress": 100, "status": "완화 완료"},
            ]},
        ],
        "gate": {"reviewDate": "2026-09-15", "label": "게이트 리뷰(양산시범 이관)", "criteria": [
            {"label": "① MCBF 성장 목표", "value": "auto:growth", "status": "prog"},
            {"label": "② 만성(재발) 고장", "value": "1건(DRM-03) — 마감 필요", "status": "fail"},
            {"label": "③ 시정조치 검증마감", "value": "auto:actions", "status": "prog"},
            {"label": "④ Critical 재발", "value": "0건 — POC 소진 유지", "status": "pass"},
            {"label": "⑤ 무정지 300h", "value": "설계 동결(9/10) 후", "status": "wait"},
        ]},
        "project": {"name": "드럼 자동화 (Pilot)", "department": "인프라 기술팀", "team": "김OO, 박OO",
                    "startDate": "2026-07-20", "endDate": "2026-10-31"},
        "swModules": [
            {"name": "비전 인식", "pct": 95, "group": "로봇"}, {"name": "체결 시퀀스", "pct": 90, "group": "로봇"},
            {"name": "그리퍼 제어", "pct": 85, "group": "로봇"}, {"name": "PLC I/F", "pct": 80, "group": "상위시스템"},
            {"name": "로그/리포트", "pct": 70, "group": "상위시스템"}, {"name": "안전 인터록", "pct": 100, "group": "환경"},
        ],
        "lifecycle": [
            {"stage": "L1 가동 지표 정의", "status": "done", "note": "MCBF/MTTR/가동률 — 양산시범과 동일 산식"},
            {"stage": "L2 신뢰성 평가/검증 (TAAF)", "status": "current", "note": "성장 추적 · DRM-03 만성 마감 중"},
            {"stage": "L3 임시 사용 안전 승인 · 퀄 사양", "status": "current", "note": "v0.9 협의"},
            {"stage": "L4 무정지 300h", "status": "todo", "note": "설계 동결(9/10) 후 본 런"},
        ],
        "ui": {
            "app": {"title": "드럼 자동화 — Pilot", "brandLogo": "드", "brandName": "드럼 자동화<br>Pilot",
                    "evalDateLabel": "평가일", "printBtn": "PDF 리포트", "footBrand": "FRACAS + MCBF 성장 — POC 대장 승계", "updatedPrefix": "업데이트 "},
            "nav": {"overview": "한눈에 보기", "all": "평가 상세 내역"},
            "common": {"stDone": "완료", "stCurrent": "진행 중", "stTodo": "예정", "noteEmpty": "— 메모"},
            "overview": {"stageTitle": "Pilot 세부 단계", "stageCurrentPrefix": "현재: ", "stageSub": "L1→L4 · POC 어휘(DRM) 승계",
                "discussItems": [
                    {"topic": "DRM-03 만성 마감 — Rev C 전수 검증", "tag": "긴급", "group": "운영"},
                    {"topic": "설계 동결(9/10) 전 변경 등급 매트릭스 합의", "tag": "협의", "group": "운영"},
                    {"topic": "안전인증 심사 서류 (~9/1)", "tag": "진행", "group": "안전"},
                ],
                "goalsMonth": "만성(DRM-03) 마감 → 설계 동결 → 300h 본 런",
                "goalsWeek": "Rev C 무발생 검증\n안전인증 서류 제출"},
            "modal": {"title": "상세"},
        },
    })
    n = sum(r[4] for r in daily_rows)
    assert n == len(error_rows)
    print(f"[demo] drum-pilot: daily {len(daily_rows)}일 · 에러 {len(error_rows)}건 (Critical 0 — POC 소진)")


def gen_drum_mass():
    pid = "drum-mass"
    (PROJECTS / pid / "raw").mkdir(parents=True, exist_ok=True)
    (PROJECTS / pid / "errors").mkdir(exist_ok=True)
    days, d = [], date(2026, 11, 2)
    while len(days) < 24:
        if d.weekday() < 6:
            days.append(d)
        d += timedelta(days=1)
    err_at = {7: 1, 13: 1, 19: 2}
    daily_rows, err_cycles = [], []
    streak = cum = 0
    for i, dt in enumerate(days):
        e = err_at.get(i, 0)
        streak = 0 if e else streak + 15
        for k in range(e):
            err_cycles.append((dt, cum + 7 + k * 4))
        cum += 15
        daily_rows.append((dt.isoformat(), 2, "체결·반송 실부하 운전", 15, e, streak, "에러 — 리셋" if e else ""))
    # 4건 — Critical 0 · DRM-03 재발 0 (조기 소진 효과) · 외생(자재) 혼입 시작
    err_defs = [
        ("DRM-02", "비전 오인식", "신규 로트 표면 반사율 편차", "자재 로트 편차", "시험환경·자재", "로트 수입검사 추가", "v1.0.0"),
        ("DRM-04", "통신 지연", "라인 PLC 응답 지연", "라인측 설정", "운영·조작", "라인 설정 표준화", "v1.0.0"),
        ("DRM-06", "자재 급송 지연", "급송 슈트 이물", "라인 청소 주기", "운영·조작", "청소 SOP 반영", "v1.0.1"),
        ("DRM-06", "자재 급송 지연", "동일 모드 — 슈트 마모", "슈트 코팅 마모", "부품", "코팅 사양 변경", "v1.0.1"),
    ]
    error_rows = [
        (i + 1, dt.isoformat(), f"{9 + i * 2}:{15 + i * 7 % 40:02d}", cyc,
         code, mode, det, cause, cls, act, "정상복귀", "김OO", "업체 박OO", sw, "Rev C", "", "")
        for i, ((dt, cyc), (code, mode, det, cause, cls, act, sw)) in enumerate(zip(err_cycles, err_defs))
    ]
    wb = Workbook(); wb.remove(wb.active)
    _sheet(wb, "안내", ["양산 시범 — 일일평가 + 에러로그(원인분류·버전). 판정·처분은 REPORT.xlsx."], [])
    _sheet(wb, "일일평가", ["평가일", "입실인원", "주평가내용", "일일평가", "일일에러", "연속성공", "비고"], daily_rows)
    _sheet(wb, "에러로그", ["No", "발생일", "시각", "회차", "코드", "유형", "상세", "원인", "원인분류", "조치", "결과",
                          "삼성 담당자", "업체 담당자", "SW버전", "HW버전", "상세설명", "사진(파일명)"], error_rows)
    wb.save(PROJECTS / pid / "raw" / "드럼양산평가_샘플.xlsx")
    wb2 = Workbook(); wb2.remove(wb2.active)
    _codes_sheet(wb2)
    _sheet(wb2, "조치검증", ["조치ID", "대상코드", "조치내용", "담당", "목표일", "상태", "검증시작"], [
        ("A-001", "DRM-02", "로트 수입검사 항목 추가", "김OO", "2026-11-20", "검증완료", "2026-11-12"),
        ("A-002", "DRM-04", "라인 PLC 설정 표준화", "박OO", "2026-11-25", "검증완료", "2026-11-18"),
        ("A-003", "DRM-06", "슈트 코팅 사양 변경", "김OO", "2026-12-10", "검증중", "2026-12-01"),
    ])
    _sheet(wb2, "판정대장", ["사건ID", "대상에러No", "판정", "귀책분류", "증거", "합의상태", "판정일"], [
        ("JD-01", 1, "비관련", "자재 (로트 편차)", "로트 성적서 + 재현시험", "합의완료", "2026-11-12"),
        ("JD-02", 2, "비관련", "운영 (라인 설정)", "PLC 로그", "합의완료", "2026-11-19"),
        ("JD-03", 3, "관련", "설비 (운영 연동)", "슈트 점검 기록", "합의완료", "2026-11-28"),
        ("JD-04", 4, "관련", "설비 (부품)", "마모 측정 + 재현", "합의완료", "2026-12-01"),
    ])
    _sheet(wb2, "처분대장", ["처분ID", "대상ID", "처분", "사유", "기한", "오너", "합의"], [
        ("DSP-01", "DRM-06", "종결예정", "코팅 사양 변경 검증 중 — 심의 전 종결 예상", "2026-12-20", "김OO", "합의완료"),
    ])
    wb2.save(PROJECTS / pid / "REPORT.xlsx")
    cfg = json.loads((PROJECTS / "chem" / "config.json").read_text(encoding="utf-8"))
    for k in ("_stage_readme", "_swModules_readme", "_ui_readme", "_edit_guide"):
        cfg.pop(k, None)
    cfg["project"] = {"name": "드럼 자동화 (양산 시범 평가)", "department": "인프라 기술팀",
                      "team": "김OO, 박OO", "startDate": "2026-11-02", "endDate": "2027-01-31"}
    cfg["gate"] = {"reviewDate": "2027-01-15", "label": "가동인증"}
    cfg["tecop"] = [
        {"k": "T", "status": "ok", "note": "Critical·만성 재발 0 — 조기 소진 효과로 완주 순항", "risks": [
            {"id": "R-T1", "risk": "외생 요인(자재·운영) 관련 판정 재유입 — 에러 버짓 소모", "level": "Low", "mitigation": "비관련 건도 시정 오너 지정 (CRITERIA §4)", "owner": "PM", "due": "", "progress": 0, "status": "감시(수용)", "link": "E-003"},
        ]},
        {"k": "E", "status": "ok", "note": "실증치 축적 — 투자심의 입력 준비", "risks": [
            {"id": "R-E1", "risk": "투자심의 입력치(MCBF·가동률) 미달 가능성", "level": "Medium", "mitigation": "360Cy 완주 실증치 확보 — 월간 점검", "owner": "PM", "due": "2027-01-15", "progress": 60, "status": "완화중"},
        ]},
        {"k": "C", "status": "ok", "note": "유지보수 계약 조건 합의 · 잣대 동결 유지"},
        {"k": "O", "status": "ok", "note": "운영 인수 준비(교육 계획) 진행"},
        {"k": "P", "status": "ok", "note": "수혜부서 합동판정 참여 정상 · 설치 위험성 평가 완료", "risks": [
            {"id": "R-P1", "risk": "가동인증 심의 일정 리스크", "level": "Low", "mitigation": "산출물 체크리스트 사전 준비 — 일정 정상", "owner": "PM", "due": "2027-01-15", "progress": 100, "status": "완화 완료"},
        ]},
    ]
    cfg["swModules"] = [
        {"name": "비전 인식", "pct": 100, "group": "로봇"}, {"name": "체결 시퀀스", "pct": 100, "group": "로봇"},
        {"name": "그리퍼 제어", "pct": 95, "group": "로봇"}, {"name": "라인 I/F", "pct": 85, "group": "상위시스템"},
        {"name": "리포트 연동", "pct": 70, "group": "상위시스템"}, {"name": "안전 인터록", "pct": 100, "group": "환경"},
    ]
    cfg["ui"]["app"].update({"title": "드럼 자동화 — 양산 시범 평가", "brandLogo": "드",
                             "brandName": "드럼 자동화<br>양산 시범 평가"})
    ov = cfg["ui"].setdefault("overview", {})
    ov.update({
        "goalsMonth": "연속 360Cy 완주 (에러버짓 관련만)\nDRM-06 검증 종결",
        "goalsWeek": "판정 완료분 대장 정리\n가동인증 산출물 준비",
        "discussItems": [
            {"topic": "DRM-06 슈트 코팅 검증 런", "tag": "진행", "group": "운영"},
            {"topic": "자재 로트 수입검사 기준 — 구매 협의", "tag": "협의", "group": "운영"},
            {"topic": "가동인증 산출물 체크리스트 사전 점검", "tag": "검토", "group": "기타"},
        ],
        "lineCaption": "현재 평가 <b>체결 셀 · {cum}/{target}</b> · 반송부 통과 · 출하 I/F 대기",
    })
    _write_config(pid, cfg)
    src_img = PROJECTS / "chem" / "assets" / "line_layout.png"
    if src_img.exists():
        (PROJECTS / pid / "assets").mkdir(exist_ok=True)
        shutil.copy(src_img, PROJECTS / pid / "assets" / "line_layout.png")
    print(f"[demo] drum-mass: daily {len(daily_rows)}일 · 에러 {len(error_rows)}건 (Critical 0·재발은 DRM-06뿐)")


def gen_drum_spread():
    pid = "drum-spread"
    (PROJECTS / pid / "raw").mkdir(parents=True, exist_ok=True)
    (PROJECTS / pid / "errors").mkdir(exist_ok=True)
    units = [
        ("1호기", "L1", "2027-02-03", "PASS", 48, 48, "퀄 완료", ""),
        ("2호기", "L1", "2027-02-05", "PASS", 48, 48, "퀄 완료", ""),
        ("3호기", "L2", "2027-02-17", "PASS", 48, 48, "퀄 완료", ""),
        ("4호기", "L2", "2027-02-19", "PASS", 48, 48, "퀄 완료", ""),
        ("5호기", "L3", "2027-03-03", "PASS", 31, 48, "런 진행", ""),
        ("6호기", "L3", "2027-03-05", "진행", 0, 48, "SAT 진행", ""),
        ("7호기", "L4", "2027-03-17", "—", 0, 48, "설치 중", ""),
        ("8호기", "L4", "2027-03-24", "—", 0, 48, "설치 예정", ""),
    ]
    # 6건 — 설계성 0 · Critical 0 (여정의 결실) · 설치/운영 기인 위주 · 신규 모드 2건(어휘 v2.1 추가)
    issues = [
        ("비전 오인식", "Minor", "운영·환경", "종결", "2027-02-10", "2027-02-12", "", "1호기", "라인 조명 편차 — 임계 재설정 (DRM-02)"),
        ("레일 수평도 초과", "Major", "설치·시공", "종결", "2027-02-06", "2027-02-08", "", "2호기", "신규 모드(v2.1 추가) — 재시공 후 SAT 재수행"),
        ("AP 음영 끊김", "Major", "설치·시공", "검증중", "2027-03-06", "", "12/20일", "5호기", "신규 모드(v2.1) — AP 증설 후 무발생 감시"),
        ("통신 지연", "Minor", "설치·시공", "종결", "2027-02-20", "2027-02-22", "", "4호기", "배선 경로 간섭 — 재배선 (DRM-04)"),
        ("커넥터 미압착", "Minor", "제작·조립", "종결", "2027-03-04", "2027-03-05", "", "5호기", "조립 체크리스트 보강"),
        ("수동 개입 절차 미준수", "Minor", "운영·환경", "조치중", "2027-03-10", "", "", "3호기", "운영 SOP 재교육"),
    ]
    rows = [(f"DIS-{i+1:03d}", m, sv, c, st, d1, d2, vf, u, det, "")
            for i, (m, sv, c, st, d1, d2, vf, u, det) in enumerate(issues)]
    wb = Workbook(); wb.remove(wb.active)
    _sheet(wb, "안내", ["확산 — 이슈로그(원인계층·호기 필수) + 호기퀄. 설계성 고장 = 전 함대 리스크 즉시 에스컬레이션."], [])
    _sheet(wb, "이슈로그", ["이슈ID", "고장모드", "심각도", "원인분류", "상태", "발생일", "종결일", "무발생검증", "호기", "상세", "사진(파일명)"], rows)
    _sheet(wb, "호기퀄", ["호기", "라인", "설치일", "SAT", "축약런(h)", "목표(h)", "상태", "비고"], units)
    _codes_sheet(wb)
    wb.save(PROJECTS / pid / "raw" / "드럼확산_샘플.xlsx")
    _write_config(pid, {
        "stage": "spread",
        "run": {"target": 48, "unit": "호기", "criterion": "축약 런", "env": "각 적용 라인 (호기별 SAT)"},
        "tecop": [
            {"k": "T", "status": "ok", "note": "설계성 0 · Critical 0 — 개발기 대장의 결실", "risks": [
                {"id": "R-T1", "risk": "신규 모드(AP 음영) 재발 시 확산 지연", "level": "Medium", "mitigation": "AP 증설 후 무발생 감시 — 어휘 v2.1 등재", "owner": "박OO", "due": "2027-03-26", "progress": 60, "status": "완화중", "link": "DIS-003"},
            ]},
            {"k": "E", "status": "ok", "note": "확산 투자 승인 조건(효과 실증치) 유지"},
            {"k": "C", "status": "ok", "note": "호기 추가 발주 정상", "risks": [
                {"id": "R-C1", "risk": "호기 추가 발주 단가 변동", "level": "Low", "mitigation": "물량 계약 단가 고정 조항 — 발주 완료분 적용", "owner": "PM", "due": "", "progress": 0, "status": "감시(수용)"},
            ]},
            {"k": "O", "status": "ok", "note": "라인별 운영 인수 교육 진행", "risks": [
                {"id": "R-O1", "risk": "5~8호기 라인 인수 교육 지연", "level": "Medium", "mitigation": "교육 일정 라인 협의 확정 — 3/4 완료", "owner": "김OO", "due": "2027-04-30", "progress": 75, "status": "완화중"},
            ]},
            {"k": "P", "status": "ok", "note": "수혜 라인 인수 합의 정상 · 호기별 가동인증 일정 정상"},
        ],
        "gate": {"reviewDate": "2027-04-15", "label": "확산 완료 리뷰", "criteria": [
            {"label": "① 호기별 퀄 (SAT+런)", "value": "auto:fleet", "status": "prog"},
            {"label": "② 설계성 고장 0", "value": "0건 유지", "status": "pass"},
            {"label": "③ Critical 0", "value": "0건 — 전 여정 유지", "status": "pass"},
            {"label": "④ 기준 구성 동결", "value": "편차 0", "status": "pass"},
            {"label": "⑤ 횡전개 문서", "value": "설치 표준 v2.1", "status": "prog"},
        ]},
        "project": {"name": "드럼 자동화 (양산 적용·확산)", "department": "인프라 기술팀", "team": "박OO, 정OO",
                    "startDate": "2027-02-01", "endDate": "2027-05-31"},
        "swModules": [
            {"name": "체결 제어 (동결 v1.1)", "pct": 100, "group": "로봇"},
            {"name": "호기 파라미터 셋", "pct": 80, "group": "로봇"},
            {"name": "관제 연동", "pct": 90, "group": "상위시스템"},
            {"name": "설치 자동 점검", "pct": 70, "group": "상위시스템"},
            {"name": "라인별 안전 인터록", "pct": 100, "group": "환경"},
        ],
        "lifecycle": [
            {"stage": "S1 기준 구성 동결", "status": "done", "note": "v1.1 동결 — 편차 0"},
            {"stage": "S2 호기 설치·SAT", "status": "current", "note": "8호기 중 SAT 5 통과"},
            {"stage": "S3 축약 무고장 런 (48h)", "status": "current", "note": "4호기 통과 · 5호기 진행"},
            {"stage": "S4 가동지표 검증·횡전개", "status": "todo", "note": "설치 표준 v2.1 개정 중"},
        ],
        "ui": {
            "app": {"title": "드럼 자동화 — 양산 적용(확산)", "brandLogo": "드", "brandName": "드럼 자동화<br>확산",
                    "evalDateLabel": "평가일", "printBtn": "PDF 리포트", "footBrand": "원인계층 + 호기 층화 — 어휘 v2.1", "updatedPrefix": "업데이트 "},
            "nav": {"overview": "한눈에 보기", "all": "평가 상세 내역"},
            "common": {"stDone": "완료", "stCurrent": "진행 중", "stTodo": "예정", "noteEmpty": "— 메모"},
            "overview": {"stageTitle": "확산 세부 단계", "stageCurrentPrefix": "현재: ", "stageSub": "S1→S4 · 호기별 양산 퀄",
                "discussItems": [
                    {"topic": "AP 음영 — 신규 라인 시공 사양 반영 (DIS-003)", "tag": "협의", "group": "운영"},
                    {"topic": "7·8호기 설치 슬롯 협의", "tag": "진행", "group": "운영"},
                    {"topic": "신규 모드 2건 → 차기 과제 FMEA 환류", "tag": "검토", "group": "기타"},
                ],
                "goalsMonth": "5호기 런 완료 → 6/8 퀄\n설치 표준 v2.1 개정",
                "goalsWeek": "AP 증설 무발생 감시\n6호기 SAT"},
            "modal": {"title": "상세"},
        },
    })
    print(f"[demo] drum-spread: 호기 {len(units)} · 이슈 {len(issues)}건 (설계성 0 · Critical 0)")


if __name__ == "__main__":
    # 드럼 자동화 — 한 과제의 4단계 여정 (다른 가상 과제 생성은 중단: gen_drum/gen_sort/... 는 보존용)
    gen_drum_poc()
    gen_drum_pilot()
    gen_drum_mass()
    gen_drum_spread()
    print("[demo] 드럼 자동화 4단계 여정 생성 완료 — python3 scripts/build_dashboard_json.py 로 빌드")
