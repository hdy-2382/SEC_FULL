"""
generate_mgmt_template.py
-------------------------
PM가 유지하는 관리 데이터 양식(data/SEC_REPORT.xlsx)을 생성한다. (업체에 보내지 않음)

시트:
  - 안내       : 작성 방법
  - 코드마스터  : 에러 코드 → 유형 / 등급(Critical·Major·Minor) / 설명   (등급=심각도 S)
  - 조치검증    : 조치ID / 대상코드 / 조치내용 / 담당 / 목표일 / 상태 / 검증시작(날짜 또는 누적Cycle)

build_dashboard_json.py 가 이 파일을 읽어 업체 데이터와 병합한다.
업체 양식(일일평가/에러로그)은 이 파일과 무관하게 그대로 유지된다.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT_PATH = ROOT / "data" / "SEC_REPORT.xlsx"

INK = "0F2E54"; NAVY = "1E4A7A"; SKY = "2E89D6"; SKY_SOFT = "EAF2FB"
ZEBRA = "F5F9FE"; LINE = "C9DAEE"; WHITE = "FFFFFF"
FONT = "맑은 고딕"

HEADER_FONT = Font(name=FONT, size=11, bold=True, color=WHITE)
TITLE_FONT  = Font(name=FONT, size=16, bold=True, color=INK)
BODY_FONT   = Font(name=FONT, size=10, color="2B3A4F")
NOTE_FONT   = Font(name=FONT, size=10, color="3D4147")
EX_FONT     = Font(name=FONT, size=10, italic=True, color="8092A8")

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
EX_FILL     = PatternFill("solid", fgColor=SKY_SOFT)
ZEBRA_FILL  = PatternFill("solid", fgColor=ZEBRA)
WHITE_FILL  = PatternFill("solid", fgColor=WHITE)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
_thin  = Side(border_style="thin", color=LINE)
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _header(ws, cols, widths):
    for i, (name, w) in enumerate(zip(cols, widths), start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def _rows(ws, start, rows, n_cols, *, example=False, height=22):
    for r, vals in enumerate(rows, start=start):
        for i in range(1, n_cols + 1):
            c = ws.cell(row=r, column=i, value=vals[i - 1] if i - 1 < len(vals) else "")
            c.font = EX_FONT if example else BODY_FONT
            c.fill = EX_FILL if example else (ZEBRA_FILL if (r - start) % 2 else WHITE_FILL)
            c.alignment = LEFT; c.border = BORDER
        ws.row_dimensions[r].height = height


def build_guide(wb):
    ws = wb.create_sheet("안내", 0)
    ws.sheet_properties.tabColor = SKY
    ws.column_dimensions["A"].width = 16; ws.column_dimensions["B"].width = 96
    ws["A1"] = "관리 데이터(PM 유지) — 신뢰성 대시보드 입력"
    ws["A1"].font = TITLE_FONT; ws.merge_cells("A1:B1"); ws.row_dimensions[1].height = 30
    notes = [
        ("", ""),
        ("코드마스터", "에러 코드별 유형·등급(심각도)을 한 번씩 정의. 등급은 결과 기준: Critical(안전/라인정지·복구난), Major(자동정지·복구가능), Minor(자가복구/경고)."),
        ("", "→ 치명도 분포·S×O 위험 매트릭스·빈발 Top5 등급색에 사용."),
        ("조치검증", "에러에 대한 시정조치를 1건씩 기록. 검증시작 = 조치 적용 시점."),
        ("", "→ 날짜(예: 2026-06-10) 또는 누적 Cycle 숫자 둘 다 입력 가능. 날짜를 적으면 그날까지의 누적 Cycle로 자동 환산."),
        ("", "→ 조치 후 200 Cycle(설정값) 동안 같은 코드가 안 나오면 '검증 완료'로 자동 판정."),
        ("", ""),
        ("주의", "시트명(코드마스터/조치검증)·헤더는 변경하지 마세요. 이 파일은 업체에 보내지 않습니다."),
    ]
    for off, (a, b) in enumerate(notes, start=2):
        ca = ws.cell(row=off, column=1, value=a); cb = ws.cell(row=off, column=2, value=b)
        ca.font = Font(name=FONT, size=11, bold=True, color=INK) if a else BODY_FONT
        cb.font = NOTE_FONT; cb.alignment = LEFT


def build_codes(wb):
    ws = wb.create_sheet("코드마스터")
    ws.sheet_properties.tabColor = NAVY
    cols = ["코드", "유형", "등급", "설명"]
    _header(ws, cols, [14, 22, 14, 50])
    ex = [
        ["ERR-001", "Vision Timeout", "Critical", "비전 좌표 인식 실패, 로봇 정지"],
        ["ERR-002", "Gripper Slip", "Major", "그리퍼 그립 실패"],
        ["ERR-003", "Path Error", "Major", "경로 계획 오류"],
        ["ERR-005", "Sensor Noise", "Minor", "센서 노이즈/EMI"],
    ]
    _rows(ws, 2, ex, len(cols), example=True, height=24)
    _rows(ws, 2 + len(ex), [[""] * len(cols)] * 12, len(cols))


def build_actions(wb):
    ws = wb.create_sheet("조치검증")
    ws.sheet_properties.tabColor = NAVY
    cols = ["조치ID", "대상코드", "조치내용", "담당", "목표일", "상태", "검증시작(날짜/Cycle)"]
    _header(ws, cols, [10, 12, 40, 12, 13, 12, 18])
    # 검증시작 칸: 날짜(YYYY-MM-DD) 또는 누적 Cycle 숫자 — 아래는 두 방식을 섞어 예시.
    ex = [
        ["A-001", "ERR-001", "카메라 버퍼 설정 변경", "양희두", "2026-06-20", "진행중", 358],
        ["A-002", "ERR-003", "경로 플래너 펌웨어 업데이트", "김현일", "2026-06-14", "완료", "2026-06-14"],
        ["A-003", "ERR-002", "그리퍼 일지 구조 개선", "양희두", "2026-06-22", "진행중", 312],
        ["A-004", "ERR-005", "EMI 필터 추가", "김현일", "2026-06-12", "완료", 160],
    ]
    _rows(ws, 2, ex, len(cols), example=True, height=24)
    _rows(ws, 2 + len(ex), [[""] * len(cols)] * 14, len(cols))


def main():
    wb = Workbook(); wb.remove(wb.active)
    build_guide(wb); build_codes(wb); build_actions(wb)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"[mgmt] 생성 완료: {OUT_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
